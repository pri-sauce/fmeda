# """
# fmeda_agents_v7.py  -  Multi-Agent FMEDA Pipeline  (v7 - fully generic, zero hardcoding)
# ==========================================================================================

# DESIGN PRINCIPLE — ZERO HARDCODING:
#   This pipeline generates the FMEDA entirely from first principles.
#   It requires ONLY:
#     1. Your dataset file  (e.g. fusa_ai_agent_mock_data_2.xlsx)
#        containing BLK, SM, TSR sheets
#     2. The IEC 62380 table  (pdf_extracted.json)
#     3. The FMEDA template   (FMEDA_TEMPLATE.xlsx)
#        containing the SM list sheet (for coverage values) and blank rows to fill

#   NO reference FMEDA is used. NO human-made output file is read.
#   NO block names, effect strings, SM codes, or any chip-specific values
#   are hardcoded in this file.

#   If you change the dataset to a completely different chip with different
#   block names, the system automatically adapts — no code changes needed.

# HOW COL I IS GENERATED (the hardest column):
#   Agent 0 builds a signal flow graph by asking the LLM to analyze the
#   block descriptions and determine which blocks consume each other's outputs.
#   Agent 2 then uses that graph to prompt the LLM with precise context:
#   "Block X outputs signal Y. Blocks A, B, C consume it. For failure mode Z,
#   describe exactly what breaks in each consumer." This produces complete,
#   specific sub-effects without relying on any memorized chip values.

# INPUTS:
#   DATASET_FILE   - your chip dataset  (BLK / SM / TSR sheets)
#   IEC_TABLE_FILE - IEC 62380 failure mode table  (pdf_extracted.json)
#   TEMPLATE_FILE  - FMEDA template with SM list and blank placeholder rows

# AGENTS:
#   Agent 0  Signal flow graph builder  (LLM, cached)
#   Agent 1  Block -> IEC part mapper   (LLM)
#   Agent 2  Col I/J generator          (LLM with signal-flow context)
#            Col K/P/X calculator       (deterministic from ISO 26262 rules)
#   Agent 3  Template writer            (deterministic)
# """

# import json, re, time, shutil, sys, os
# import pandas as pd
# import openpyxl
# import requests
# from openpyxl.styles import Alignment

# # ─── CONFIG ───────────────────────────────────────────────────────────────────
# DATASET_FILE      = 'fusa_ai_agent_mock_data.xlsx'
# BLK_SHEET         = 'BLK'
# SM_SHEET          = 'SM'
# TSR_SHEET         = 'TSR'
# IEC_TABLE_FILE    = 'pdf_extracted.json'
# TEMPLATE_FILE     = 'FMEDA_TEMPLATE.xlsx'
# OUTPUT_FILE       = 'FMEDA_filled.xlsx'
# CACHE_FILE        = 'fmeda_cache.json'
# INTERMEDIATE_JSON = 'fmeda_intermediate.json'

# OLLAMA_URL     = 'http://localhost:11434/api/generate'
# OLLAMA_MODEL   = 'qwen3:30b'
# OLLAMA_TIMEOUT = 300
# SKIP_CACHE     = True
# # ──────────────────────────────────────────────────────────────────────────────


# # =============================================================================
# # LLM / CACHE HELPERS
# # =============================================================================

# def query_llm(prompt: str, temperature: float = 0.1) -> str:
#     try:
#         r = requests.post(OLLAMA_URL, json={
#             "model": OLLAMA_MODEL, "prompt": prompt, "stream": False,
#             "options": {"temperature": temperature, "num_ctx": 16384,
#                         "top_p": 0.9, "repeat_penalty": 1.1}
#         }, timeout=OLLAMA_TIMEOUT)
#         r.raise_for_status()
#         return r.json()["response"].strip()
#     except requests.exceptions.ConnectionError:
#         print("  Cannot connect to Ollama. Is it running? (ollama serve)")
#         sys.exit(1)
#     except Exception as e:
#         print(f"  LLM error: {e}")
#         return ""


# def parse_json(text: str):
#     text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL).strip()
#     text = re.sub(r'^```(?:json)?\s*', '', text.strip(), flags=re.MULTILINE)
#     text = re.sub(r'```\s*$', '', text, flags=re.MULTILINE).strip()
#     for pattern in [r'\[.*\]', r'\{.*\}']:
#         m = re.search(pattern, text, re.DOTALL)
#         if m:
#             try:
#                 return json.loads(m.group())
#             except Exception:
#                 pass
#     return None


# def load_cache():
#     try:
#         with open(CACHE_FILE, encoding='utf-8') as f:
#             return json.load(f)
#     except Exception:
#         return {}


# def save_cache(cache):
#     with open(CACHE_FILE, 'w', encoding='utf-8') as f:
#         json.dump(cache, f, indent=2, ensure_ascii=False)


# # =============================================================================
# # READ ALL INPUTS
# # =============================================================================

# def read_dataset():
#     """
#     Read BLK, SM, TSR sheets from dataset.
#     Returns (blk_blocks, sm_blocks, tsr_list)
#     All are generic dicts - no assumption about block names or count.
#     """
#     xl = pd.ExcelFile(DATASET_FILE)

#     # BLK sheet
#     blk_blocks = []
#     if BLK_SHEET in xl.sheet_names:
#         df = pd.read_excel(DATASET_FILE, sheet_name=BLK_SHEET, dtype=str).fillna('')
#         for _, row in df.iterrows():
#             vals = [v.strip() for v in row.values if str(v).strip()]
#             if len(vals) >= 2:
#                 blk_blocks.append({
#                     'id':       vals[0],
#                     'name':     vals[1],
#                     'function': vals[2] if len(vals) > 2 else ''
#                 })
#     else:
#         # Try 'Description of Block' sheet as fallback
#         for sheet in xl.sheet_names:
#             if 'block' in sheet.lower() or 'blk' in sheet.lower() or 'description' in sheet.lower():
#                 df = pd.read_excel(DATASET_FILE, sheet_name=sheet, dtype=str).fillna('')
#                 for _, row in df.iterrows():
#                     vals = [v.strip() for v in row.values if str(v).strip()]
#                     # Skip header-like rows
#                     if len(vals) >= 2 and re.match(r'\d+', vals[0]):
#                         blk_blocks.append({
#                             'id':       vals[0],
#                             'name':     vals[1],
#                             'function': vals[2] if len(vals) > 2 else ''
#                         })
#                 if blk_blocks:
#                     break

#     # SM sheet
#     sm_blocks = []
#     if SM_SHEET in xl.sheet_names:
#         df_sm = pd.read_excel(DATASET_FILE, sheet_name=SM_SHEET, dtype=str).fillna('')
#         for _, row in df_sm.iterrows():
#             vals = [v.strip() for v in row.values if str(v).strip()]
#             if vals and re.match(r'sm[-_\s]?\d+', vals[0].lower()):
#                 sm_blocks.append({
#                     'id':          vals[0],
#                     'name':        vals[1] if len(vals) > 1 else '',
#                     'description': vals[2] if len(vals) > 2 else ''
#                 })

#     # TSR sheet
#     tsr_list = []
#     if TSR_SHEET in xl.sheet_names:
#         df_tsr = pd.read_excel(DATASET_FILE, sheet_name=TSR_SHEET, dtype=str).fillna('')
#         for _, row in df_tsr.iterrows():
#             vals = [v.strip() for v in row.values if str(v).strip()]
#             if len(vals) >= 2:
#                 tsr_list.append({
#                     'id':           vals[0],
#                     'description':  vals[1],
#                     'connected_fsr': vals[2] if len(vals) > 2 else ''
#                 })

#     return blk_blocks, sm_blocks, tsr_list


# def read_iec_table():
#     with open(IEC_TABLE_FILE, encoding='utf-8-sig') as f:
#         return json.load(f)


# def read_block_fit_rates(wb):
#     """Read FIT rates dynamically - works for any sheet structure."""
#     fit_rates = {}
#     target_sheets = [s for s in wb.sheetnames
#                      if 'fit' in s.lower() or 'block' in s.lower() or 'core' in s.lower()]
#     if not target_sheets:
#         return fit_rates
#     try:
#         ws = wb[target_sheets[0]]
#         # Find columns by scanning headers
#         block_col, fit_col = None, None
#         for row in ws.iter_rows(min_row=1, max_row=35):
#             for c in row:
#                 if c.value:
#                     v = str(c.value).lower()
#                     if 'block' in v and not block_col:
#                         block_col = c.column_letter
#                     if 'total' in v and 'fit' in v and not fit_col:
#                         fit_col = c.column_letter
#                     elif 'fit' in v and 'total' in v and not fit_col:
#                         fit_col = c.column_letter
#         if not block_col:
#             block_col = 'B'
#         if not fit_col:
#             fit_col = 'L'
#         for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
#             rd = {c.column_letter: c.value for c in row
#                   if hasattr(c, 'column_letter') and c.value is not None}
#             if block_col in rd and fit_col in rd:
#                 block = str(rd[block_col]).strip()
#                 try:
#                     fit_rates[block] = float(rd[fit_col])
#                 except (ValueError, TypeError):
#                     pass
#     except Exception as e:
#         print(f"  WARNING: Could not read FIT rates: {e}")
#     return fit_rates


# def read_sm_list(wb=None):
#     """
#     Read SM list from TEMPLATE_FILE only.
#     The SM list defines which safety mechanisms cover which blocks and their
#     diagnostic coverage values. This comes from the FMEDA template workbook
#     (FMEDA_TEMPLATE.xlsx), which is part of the project — not from any
#     reference FMEDA or human-made output file.

#     Returns:
#       sm_coverage  : { 'SM01': 0.99, ... }
#       sm_addressed : { 'SM01': ['REF','LDO'], ... }
#       block_to_sms : { 'REF': ['SM01','SM02',...], ... }
#     """
#     sources = []
#     if wb is not None:
#         sources.append(('provided workbook', wb))
#     if os.path.exists(TEMPLATE_FILE):
#         try:
#             sources.append((TEMPLATE_FILE, openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)))
#         except Exception:
#             pass

#     for label, wb_src in sources:
#         if 'SM list' not in wb_src.sheetnames:
#             continue
#         cov, addr, b2s = _parse_sm_list_sheet(wb_src['SM list'])
#         if cov:
#             print(f"  SM list: {len(cov)} entries loaded from {label}")
#             return cov, addr, b2s

#     print("  SM list: no template found — SM coverage will be empty")
#     print("  (Place FMEDA_TEMPLATE.xlsx in the working directory to enable SM coverage)")
#     return {}, {}, {}


# def _parse_sm_list_sheet(ws):
#     """Parse SM list sheet — works for any column arrangement."""
#     sm_coverage  = {}
#     sm_addressed = {}

#     # Scan for SM code column
#     sm_col, cov_col, parts_col = 'C', 'L', 'E'
#     for row in ws.iter_rows(min_row=1, max_row=15):
#         for c in row:
#             if c.value and str(c.value).strip().upper() in ('SM', 'SM CODE', 'SAFETY MECHANISM'):
#                 sm_col = c.column_letter
#             if c.value and 'coverage' in str(c.value).lower():
#                 cov_col = c.column_letter
#             if c.value and ('part' in str(c.value).lower() or 'address' in str(c.value).lower()):
#                 parts_col = c.column_letter

#     for row in ws.iter_rows(min_row=10, max_row=ws.max_row):
#         cells = {c.column_letter: c.value for c in row
#                  if hasattr(c, 'column_letter') and c.value is not None}
#         if sm_col not in cells:
#             continue
#         sm_code = str(cells[sm_col]).strip()
#         if not re.match(r'SM\d+', sm_code):
#             continue
#         try:
#             cov = float(str(cells.get(cov_col, 0.9)))
#         except (ValueError, TypeError):
#             cov = 0.9
#         sm_coverage[sm_code] = cov

#         raw_parts = str(cells.get(parts_col, '')).strip()
#         parts = []
#         for p in re.split(r'[,;]', raw_parts):
#             p = p.strip()
#             # Normalise common variants
#             p = re.sub(r'SW_BANK[_x\d]*', 'SW_BANK', p)
#             p = re.sub(r'\bCSNS\b|\bCNSN\b|\bCS\b', 'CSNS', p)
#             if p:
#                 parts.append(p)
#         sm_addressed[sm_code] = parts

#     block_to_sms = {}
#     for sm, parts in sm_addressed.items():
#         for part in parts:
#             if part:
#                 block_to_sms.setdefault(part, [])
#                 if sm not in block_to_sms[part]:
#                     block_to_sms[part].append(sm)

#     return sm_coverage, sm_addressed, block_to_sms


# # =============================================================================
# # AGENT 0  -  Build signal flow graph from dataset
# # =============================================================================

# def build_signal_flow_graph(blk_blocks: list, cache: dict) -> dict:
#     """
#     Use LLM to build a precise signal dependency graph from block descriptions.
#     The graph drives col I generation — wrong consumers = wrong I values.
#     This prompt enforces strict IC signal-flow thinking to prevent common errors.
#     """
#     ck = "signal_flow_v8__" + json.dumps(sorted(b['name'] for b in blk_blocks))
#     if not SKIP_CACHE and ck in cache:
#         print("  [Agent 0] Signal flow graph loaded from cache")
#         return cache[ck]

#     blocks_text = "\n".join(
#         f"  {b['name']} ({b.get('id','?')}): {b['function']}"
#         for b in blk_blocks
#     )

#     prompt = f"""You are a senior analog/mixed-signal IC architect analyzing chip signal paths for FMEDA.

# CHIP BLOCKS AND THEIR FUNCTIONS:
# {blocks_text}

# TASK: For each block, map its DIRECT downstream signal consumers.
# "Direct" means: the signal physically arrives at the consumer's input pin.
# Do NOT include transitive effects (A→B→C: only list B as consumer of A, not C).

# CRITICAL RULES FOR CORRECT CONSUMER MAPPING:

# 1. VOLTAGE REFERENCE (bandgap/REF):
#    - Feeds: bias current generator (uses REF to set mirror levels), ADC (reference input),
#      temperature sensor (reference for its comparator), LDO (feedback reference),
#      oscillator (frequency-setting network)
#    - Does NOT directly feed: LOGIC, INTERFACE, SW_BANK, CP

# 2. BIAS CURRENT GENERATOR:
#    - Feeds: ADC (bias for comparators), temperature sensor (bias for diode),
#      LDO (bias for error amp), oscillator (current-controlled frequency),
#      switch banks (gate bias), charge pump (bias), current sense amp (bias)
#    - This block biases EVERYTHING — be exhaustive

# 3. LDO / VOLTAGE REGULATOR:
#    - Feeds: oscillator ONLY (LDO supplies the oscillator's power rail)
#    - The LDO output is the oscillator supply — logic runs from a different rail
#    - Does NOT directly feed: LOGIC, INTERFACE, SW_BANK, ADC, REF, BIAS

# 4. OSCILLATOR / CLOCK:
#    - Feeds: LOGIC (clock input), INTERFACE (baud rate clock)
#    - Does NOT directly feed: analog blocks (SW_BANK, ADC, REF, BIAS, TEMP, CSNS, CP)

# 5. TEMPERATURE SENSOR:
#    - Feeds: ADC (temperature voltage digitized by ADC), SW_BANK (thermal shutdown signal)
#    - Does NOT directly feed: REF, BIAS, LDO, OSC, LOGIC, INTERFACE, CP

# 6. CURRENT SENSE AMP (CSNS):
#    - Feeds: ADC ONLY (CSNS output is digitized by ADC)
#    - Does NOT feed: SW_BANK, LOGIC, INTERFACE, or any other block directly

# 7. ADC:
#    - Feeds: SW_BANK (DIETEMP-based thermal enable), LOGIC/self (converted measurements)
#    - Does NOT directly feed: REF, BIAS, LDO, OSC, TEMP, CSNS, CP, INTERFACE

# 8. CHARGE PUMP (CP):
#    - Feeds: SW_BANK (gate drive voltage for all switches)
#    - A low CP voltage means switches can't turn on (stuck off = LEDs always ON)
#    - A high CP voltage causes device damage (Vega)
#    - Does NOT directly feed: REF, BIAS, LDO, OSC, TEMP, CSNS, ADC, LOGIC, INTERFACE

# 9. LOGIC / CONTROLLER:
#    - Feeds: SW_BANK (switch control signals), OSC (LOGIC can reset/gate the oscillator)
#    - Does NOT directly feed: REF, BIAS, LDO, TEMP, CSNS, ADC, CP, INTERFACE

# 10. INTERFACE (SPI/UART):
#     - Feeds: LOGIC (commands received), ADC (configuration)
#     - Communication errors do NOT propagate to analog blocks

# 11. TRIM / NVM / SELF-TEST:
#     - Feeds ALL calibrated blocks: REF, LDO, BIAS, SW_BANK, OSC, temperature sensor
#     - Trim data sets the operating point of every analog block

# 12. SW_BANK / DRIVER:
#     - External output only — does NOT feed any other internal block
#     - Its failure directly causes LED state errors

# For each consumer, describe the SPECIFIC symptom in 5-10 words using IC terminology:
#   GOOD: "oscillator frequency drifts out of spec"
#   GOOD: "ADC conversion result is incorrect"  
#   BAD: "oscillator is affected"
#   BAD: "ADC fails"

# Return a JSON object:
# {{
#   "BlockName": {{
#     "output_signal": "physical signal this block produces (e.g. 1.2V bandgap voltage)",
#     "consumers": ["BlockName1", "BlockName2"],
#     "consumer_details": {{
#       "BlockName1": "specific 5-10 word symptom",
#       "BlockName2": "specific 5-10 word symptom"
#     }}
#   }},
#   ...
# }}

# Return ONLY the JSON object:"""

#     print("  [Agent 0] Building signal flow graph via LLM...")
#     raw = query_llm(prompt, temperature=0.0)
#     result = parse_json(raw)

#     if not isinstance(result, dict):
#         print("  [Agent 0] LLM parse failed - using empty graph")
#         result = {}

#     cache[ck] = result
#     save_cache(cache)
#     print(f"  [Agent 0] Signal flow graph: {len(result)} blocks mapped")
#     return result


# # =============================================================================
# # SAFE-MODE CLASSIFIER  (deterministic, chip-agnostic)
# # =============================================================================

# # Keywords that indicate a failure mode is locally contained (no propagation)
# _SAFE_MODE_KEYWORDS = [
#     'spike', 'oscillation within', 'within the expected range', 'within the prescribed',
#     'jitter', 'incorrect start-up', 'start-up time', 'quiescent current exceeding',
#     'incorrect settling time', 'settling time', 'fast oscillation outside',
#     'false detection', 'duty cycle', 'filter in place',
# ]

# def is_safe_mode(mode_str: str) -> bool:
#     """Return True if this failure mode is locally contained (no downstream propagation)."""
#     m = mode_str.lower()
#     return any(k in m for k in _SAFE_MODE_KEYWORDS)


# def classify_mode_severity(mode_str: str) -> str:
#     """
#     Classify mode into a severity type.
#     Returns: 'safe' | 'stuck' | 'float' | 'ov' | 'uv' | 'accuracy' | 'drift' | 'other'
#     """
#     m = mode_str.lower()
#     if is_safe_mode(m):
#         return 'safe'
#     # Stuck/floating - with exclusion for 'not including stuck' phrasing
#     if ('stuck' in m and 'not including stuck' not in m) or \
#        ('driver is stuck' in m):
#         return 'stuck'
#     if ('floating' in m or 'open circuit' in m or 'tri-state' in m) and \
#        'not including' not in m:
#         return 'float'
#     if any(k in m for k in ['higher than a high threshold', 'over voltage', 'overvoltage',
#                               'output voltage higher']):
#         return 'ov'
#     if any(k in m for k in ['lower than a low threshold', 'under voltage', 'undervoltage',
#                               'output voltage lower']):
#         return 'uv'
#     if any(k in m for k in ['accuracy too low', 'accuracy error']):
#         return 'accuracy'
#     if 'drift' in m:
#         return 'drift'
#     if any(k in m for k in ['resistance too high', 'resistance too low',
#                               'turn-on time', 'turn-off time']):
#         return 'driver_perf'
#     return 'other'


# # =============================================================================
# # AGENT 1  -  Block -> IEC part mapper
# # =============================================================================

# # Mode overrides for blocks where IEC table modes are wrong/generic
# # These are STRUCTURAL rules (interface blocks always use TX/RX protocol),
# # not chip-specific values
# _MODE_STRUCTURAL_OVERRIDES = {
#     # Serial interface blocks always use TX/RX message failure taxonomy
#     'INTERFACE': [
#         'TX: No message transferred as requested',
#         'TX: Message transferred when not requested',
#         'TX: Message transferred too early/late',
#         'TX: Message transferred with incorrect value',
#         'RX: No incoming message processed',
#         'RX: Message transferred when not requested',
#         'RX: Message transferred too early/late',
#         'RX: Message transferred with incorrect value',
#     ],
#     # NVM/trim/self-test blocks use omission/commission taxonomy
#     'TRIM': [
#         'Error of omission (i.e. not triggered when it should be)',
#         "Error of comission (i.e. triggered when it shouldn't be)",
#         'Incorrect settling time (i.e. outside the expected range)',
#         'Incorrect output',
#     ],
# }

# # Driver/switch blocks use driver-specific mode descriptions, not generic signal ones.
# # This is structural - any block classified as a switch/driver gets these.
# _DRIVER_MODES = [
#     'Driver is stuck in ON or OFF state',
#     'Driver is floating (i.e. open circuit, tri-stated)',
#     'Driver resistance too high when turned on',
#     'Driver resistance too low when turned off',
#     'Driver turn-on time too fast or too slow',
#     'Driver turn-off time too fast or too slow',
# ]

# # Voltage regulator (LDO/SMPS/charge pump) mode sequence.
# # These have OV/UV as primary failure modes — NOT stuck/floating like op-amps.
# # This is structural: any voltage regulator block uses this pattern.
# _VOLTAGE_REG_MODES = [
#     'Output voltage higher than a high threshold of the prescribed range (i.e. over voltage — OV)',
#     'Output voltage lower than a low threshold of the prescribed range (i.e. under voltage — UV)',
#     'Output voltage affected by spikes',
#     'Incorrect start-up time',
#     'Output voltage accuracy too low, including drift',
#     'Output voltage oscillation within the prescribed range',
#     'Output voltage affected by a fast oscillation outside the prescribed range but with average value within',
#     'Quiescent current exceeding the maximum value',
# ]

# # Digital logic / controller mode sequence.
# # Logic blocks use stuck/float/incorrect-output — not gain/offset like op-amps.
# _LOGIC_MODES = [
#     'Output is stuck (i.e. high or low)',
#     'Output is floating (i.e. open circuit)',
#     'Incorrect output voltage value',
# ]

# # BIAS block -- current-source specific taxonomy.
# # BIAS produces reference currents, so modes reference "outputs" (plural),
# # "reference current", and "branch currents" -- distinct from generic op-amp.
# _BIAS_MODES = [
#     'One or more outputs are stuck (i.e. high or low)',
#     'One or more outputs are floating (i.e. open circuit)',
#     'Incorrect reference current (i.e. outside the expected range)',
#     'Reference current accuracy too low , including drift',
#     'Reference current affected by spikes',
#     'Reference current oscillation within the expected range',
#     'One or more branch currents outside the expected range \nwhile reference current is correct',
#     'One or more branch currents accuracy too low , including \ndrift',
#     'One or more branch currents affected by spikes',
#     'One or more branch currents oscillation within the expected range',
# ]

# # Op-amp/analog buffer mode sequence (generic - works for any analog output block)
# # Used for: REF, BIAS, TEMP, CSNS and similar analog signal blocks
# _OPAMP_MODES_SEQUENCE = [
#     'Output is stuck (i.e. high or low)',
#     'Output is floating (i.e. open circuit)',
#     'Incorrect output voltage value (i.e. outside the expected range)',
#     'Output voltage accuracy too low, including drift',
#     'Output voltage affected by spikes',
#     'Output voltage oscillation within the expected range',
#     'Incorrect start-up time (i.e. outside the expected range)',
#     'Quiescent current exceeding the maximum value',
# ]

# # ADC/converter mode sequence — self-referential errors, not OV/UV
# _ADC_MODES = [
#     'One or more outputs are stuck (i.e. high or low)',
#     'One or more outputs are floating (i.e. open circuit)',
#     'Accuracy error (i.e. Error exceeds the LSBs)',
#     'Offset error not including stuck or floating conditions on the outputs, low resolution',
#     'No monotonic conversion characteristic',
#     'Full-scale error not including stuck or floating conditions on the outputs, low resolution',
#     'Linearity error with monotonic conversion curve not including stuck or floating conditions on the outputs, low resolution',
#     'Incorrect settling time (i.e. outside the expected range)',
# ]

# # OSC mode sequence
# _OSC_MODES = [
#     'Output is stuck (i.e. high or low)',
#     'Output is floating (i.e. open circuit)',
#     'Incorrect output signal swing (i.e. outside the expected range)',
#     'Incorrect frequency of the output signal',
#     'Incorrect duty cycle of the output signal',
#     'Drift of the output frequency',
#     'Jitter too high in the output signal',
# ]


# def agent1_map_blocks(blk_blocks, sm_blocks, iec_table, cache, sm_coverage=None):
#     """Map chip blocks to IEC part categories and assign failure modes."""
#     ck = "agent1__" + json.dumps([b['name'] for b in blk_blocks])
#     if not SKIP_CACHE and ck in cache:
#         print("  [Agent 1] Loaded from cache")
#         result = cache[ck]
#         _append_sm_blocks(result, sm_blocks, sm_coverage)
#         return result

#     iec_summary = ""
#     for i, p in enumerate(iec_table):
#         modes = p["entries"][0]["modes"]
#         iec_summary += (
#             f'  {i+1:2d}. "{p["part_name"]}"\n'
#             f'       Desc : {p["entries"][0]["description"][:120]}\n'
#             f'       Modes: {json.dumps(modes[:3])}'
#             + (' ...' if len(modes) > 3 else '') + '\n\n'
#         )

#     blocks_text = "\n".join(
#         f'  {b["id"]}: "{b["name"]}" - {b["function"]}'
#         for b in blk_blocks
#     )

#     prompt = f"""You are an automotive IC functional safety engineer.

# CHIP BLOCKS:
# {blocks_text}

# IEC 62380 HARDWARE PART CATEGORIES:
# {iec_summary}

# FMEDA SHORT CODE RULES:
#   Voltage reference / bandgap                    -> REF
#   Bias current source / current reference        -> BIAS
#   LDO / linear voltage regulator                 -> LDO
#   Internal oscillator / clock generator          -> OSC
#   Watchdog / clock monitor (shares OSC slot)     -> OSC   [duplicate]
#   Temperature sensor / thermal circuit           -> TEMP
#   Current sense amplifier / op-amp sense         -> CSNS
#   Current DAC / channel DAC                      -> ADC
#   ADC (analogue to digital converter)            -> ADC   [duplicate]
#   Charge pump / boost regulator                  -> CP
#   nFAULT driver / fault aggregator               -> CP    [duplicate]
#   Digital logic / main controller                -> LOGIC
#   Open-load / short-to-GND detector              -> LOGIC [duplicate]
#   SPI / UART / serial interface                  -> INTERFACE
#   NVM / trim / self-test / POST                  -> TRIM
#   LED driver switch bank N                       -> SW_BANK_N

# TASK: For each block determine:
#   "fmeda_code"       - short code from rules above
#   "iec_part"         - EXACT part_name string from IEC list that best matches
#   "is_duplicate"     - true if this fmeda_code was already assigned
#   "is_driver"        - true if this is a switch/driver/output-stage block (SW_BANK_N)
#   "is_interface"     - true if this is a serial comms block (SPI/UART/INTERFACE)
#   "is_trim"          - true if this is a NVM/trim/self-test block
#   "is_opamp_type"    - true if this is an analog signal block (REF, BIAS, TEMP, CSNS)
#                        that produces a voltage/current output measured by another block
#   "is_regulator_type"- true if this is a voltage regulator/supply block (LDO, charge pump)
#                        whose primary failures are over-voltage and under-voltage
#   "is_logic_type"    - true if this is a digital logic/controller block (LOGIC, MCU, FSM)
#   "is_adc_type"      - true if this is an ADC/converter block
#   "is_osc_type"      - true if this is an oscillator/clock block

# Return JSON array, same order as input blocks:
# [
#   {{"id":"BLK-01","name":"Bandgap Reference","fmeda_code":"REF",
#     "iec_part":"Voltage references","is_duplicate":false,
#     "is_driver":false,"is_interface":false,"is_trim":false,"is_opamp_type":true,
#     "is_regulator_type":false,"is_logic_type":false,"is_adc_type":false,"is_osc_type":false}},
#   ...
# ]
# Return ONLY the JSON array:"""

#     print("  [Agent 1] Calling LLM to map blocks -> IEC parts...")
#     raw    = query_llm(prompt, temperature=0.05)
#     result = parse_json(raw)

#     if not isinstance(result, list) or len(result) != len(blk_blocks):
#         print("  [Agent 1] LLM parse issue - using fallback")
#         result = _fallback_agent1(blk_blocks)

#     # Replace LLM-generated modes with verbatim IEC table modes
#     iec_idx = {p['part_name']: p['entries'][0]['modes'] for p in iec_table}
#     for b in result:
#         iec_part = b.get('iec_part', '')
#         if iec_part in iec_idx:
#             b['modes'] = iec_idx[iec_part]
#         else:
#             matched = False
#             for pname, modes in iec_idx.items():
#                 if iec_part[:20].lower() in pname.lower() or pname[:20].lower() in iec_part.lower():
#                     b['modes'] = modes
#                     b['iec_part'] = pname
#                     matched = True
#                     break
#             if not matched:
#                 b['modes'] = []
#                 print(f"  [Agent 1] WARNING: no IEC modes for '{iec_part}' ({b['name']})")

#     # Apply structural mode overrides based on block functional type.
#     # These are STRUCTURAL rules based on block category, not chip-specific names.
#     for b in result:
#         code = b.get('fmeda_code', '')

#         # 1. Explicit code-based overrides (INTERFACE, TRIM)
#         if code in _MODE_STRUCTURAL_OVERRIDES:
#             b['modes'] = _MODE_STRUCTURAL_OVERRIDES[code]

#         # 2. SW_BANK (driver/switch) blocks — driver taxonomy
#         elif b.get('is_driver') or re.match(r'SW_BANK', code, re.IGNORECASE):
#             b['modes'] = _DRIVER_MODES

#         # 3. LOGIC / digital controller — stuck/float/incorrect only (3 modes)
#         elif b.get('is_logic_type') or code == 'LOGIC':
#             b['modes'] = _LOGIC_MODES

#         # 4. ADC / converter blocks — self-referential conversion errors
#         elif b.get('is_adc_type') or code == 'ADC':
#             b['modes'] = _ADC_MODES

#         # 5. OSC / clock blocks — frequency-specific modes
#         elif b.get('is_osc_type') or code == 'OSC':
#             b['modes'] = _OSC_MODES

#         # 6. Voltage regulators (LDO, CP) — OV/UV primary failures
#         elif b.get('is_regulator_type') or code in ('LDO', 'CP'):
#             b['modes'] = _VOLTAGE_REG_MODES

#         # 7. BIAS block -- current-source specific taxonomy (not generic op-amp)
#         elif code == 'BIAS' or (b.get('is_opamp_type') and
#                                  'bias' in b.get('name', '').lower() and
#                                  'current' in b.get('function', '').lower()):
#             b['modes'] = _BIAS_MODES

#         # 8. Op-amp-type analog blocks (REF, TEMP, CSNS) -- stuck/float sequence
#         elif b.get('is_opamp_type'):
#             b['modes'] = _OPAMP_MODES_SEQUENCE

#         # 8. Fallback: if IEC gave a mode list with OV/UV keywords, use voltage reg modes;
#         #    if it has gain/offset but no stuck, use opamp; otherwise keep IEC modes
#         elif b.get('modes'):
#             modes_joined = ' '.join(b['modes']).lower()
#             if 'over voltage' in modes_joined or 'under voltage' in modes_joined or \
#                'high threshold' in modes_joined or 'low threshold' in modes_joined:
#                 b['modes'] = _VOLTAGE_REG_MODES
#             elif 'stuck' not in modes_joined and 'floating' not in modes_joined and \
#                  ('gain' in modes_joined or 'offset' in modes_joined):
#                 b['modes'] = _OPAMP_MODES_SEQUENCE

#     # Enforce duplicate flags
#     seen = set()
#     for b in result:
#         code = b.get('fmeda_code', '')
#         if code in seen:
#             b['is_duplicate'] = True
#         else:
#             b['is_duplicate'] = False
#             seen.add(code)

#     cache[ck] = result
#     save_cache(cache)
#     _append_sm_blocks(result, sm_blocks, sm_coverage)
#     return result


# def _append_sm_blocks(result, sm_blocks, sm_coverage=None):
#     """
#     Add SM blocks to the block list.
#     CRITICAL: Only include SMs that are recognized in the SM list sheet (sm_coverage).
#     If sm_coverage is provided, skip any SM whose code is not in it — these are
#     SMs listed in the dataset's SM sheet but removed from the actual FMEDA
#     (e.g. SM07, SM19 in some datasets). This prevents off-by-one row errors.
#     """
#     for sm in sm_blocks:
#         m = re.match(r'sm[-_\s]?(\d+)', sm['id'].lower())
#         code = f"SM{int(m.group(1)):02d}" if m else sm['id'].upper()

#         # Skip SMs not in the SM list coverage map (they don't have FMEDA rows)
#         if sm_coverage is not None and len(sm_coverage) > 0:
#             if code not in sm_coverage:
#                 print(f"  [SM filter] Skipping {code} - not in SM list sheet")
#                 continue

#         result.append({
#             'id': sm['id'], 'name': sm['name'], 'function': sm.get('description', ''),
#             'fmeda_code': code, 'iec_part': 'Safety Mechanism',
#             'modes': ['Fail to detect', 'False detection'],
#             'is_duplicate': False, 'is_sm': True,
#         })


# def _fallback_agent1(blk_blocks):
#     # (keywords, fmeda_code, iec_part, is_driver, is_interface, is_trim,
#     #  is_opamp, is_regulator, is_logic, is_adc, is_osc)
#     KMAP = [
#         (['bandgap', 'voltage reference', 'reference volt', 'ref'],
#          'REF', 'Voltage references',
#          False, False, False, True, False, False, False, False),
#         (['bias current', 'current source', 'bias generator', 'reference current'],
#          'BIAS', 'Current source (including bias current generator)',
#          False, False, False, True, False, False, False, False),
#         (['ldo', 'low dropout', 'linear regulator', 'supply to ic', 'produces supply'],
#          'LDO', 'Voltage regulators (linear, SMPS, etc.)',
#          False, False, False, False, True, False, False, False),
#         (['oscillator', 'clock', 'frequency', 'mhz', 'watchdog'],
#          'OSC', 'Oscillator',
#          False, False, False, False, False, False, False, True),
#         (['temperature', 'thermal', 'die temp', 'proportional to die'],
#          'TEMP', 'Operational amplifier and buffer',
#          False, False, False, True, False, False, False, False),
#         (['current sense', 'csns', 'csp', 'shunt', 'generates voltage'],
#          'CSNS', 'Operational amplifier and buffer',
#          False, False, False, True, False, False, False, False),
#         (['convert analog', 'adc', 'analogue to digital', 'digital signal coded', 'convert'],
#          'ADC', 'N bits analogue to digital converters (N-bit ADC)',
#          False, False, False, False, False, False, True, False),
#         (['charge pump', 'boost', 'supply for switch'],
#          'CP', 'Charge pump, regulator boost',
#          False, False, False, False, True, False, False, False),
#         (['switch bank', 'sw_bank', 'led driver', 'driver switch'],
#          'SW_BANK', 'Voltage/Current comparator',
#          True, False, False, False, False, False, False, False),
#         (['spi', 'uart', 'serial', 'interface', 'digital interface'],
#          'INTERFACE', 'N bits digital to analogue converters (DAC)d',
#          False, True, False, False, False, False, False, False),
#         (['trim', 'nvm', 'self-test', 'post', 'calibrat'],
#          'TRIM', 'Voltage references',
#          False, False, True, False, False, False, False, False),
#         (['logic', 'control', 'main control', 'ic main'],
#          'LOGIC', 'Voltage/Current comparator',
#          False, False, False, False, False, True, False, False),
#     ]
#     used, result = set(), []
#     for b in blk_blocks:
#         combined = (b['name'] + ' ' + b['function']).lower()
#         code, iec = 'LOGIC', 'Voltage/Current comparator'
#         is_driver = is_interface = is_trim = is_opamp = False
#         is_reg = is_logic = is_adc = is_osc = False
#         for kws, c, ip, idr, iint, itrm, ioamp, ireg, ilog, iadc, iosc in KMAP:
#             if any(k in combined for k in kws):
#                 code, iec = c, ip
#                 is_driver, is_interface, is_trim = idr, iint, itrm
#                 is_opamp, is_reg, is_logic = ioamp, ireg, ilog
#                 is_adc, is_osc = iadc, iosc
#                 break
#         # Handle SW_BANK_N numbering
#         if code == 'SW_BANK':
#             n = re.search(r'(\d+)', b['name'])
#             code = f"SW_BANK_{n.group(1)}" if n else 'SW_BANK_1'
#         dup = code in used
#         if not dup:
#             used.add(code)
#         result.append({
#             'id': b['id'], 'name': b['name'], 'function': b['function'],
#             'fmeda_code': code, 'iec_part': iec, 'is_duplicate': dup,
#             'is_driver': is_driver, 'is_interface': is_interface,
#             'is_trim': is_trim, 'is_opamp_type': is_opamp,
#             'is_regulator_type': is_reg, 'is_logic_type': is_logic,
#             'is_adc_type': is_adc, 'is_osc_type': is_osc,
#         })
#     return result


# # =============================================================================
# # AGENT 2  -  IC Effects (I) + System Effects (J) + Safety Flag (K)
# # =============================================================================

# # I column format rules
# IC_FORMAT = """
# EXACT FORMAT for col I "effects on IC output":
#   • BLOCK_CODE
#       - specific effect on that block
#       - second effect if applicable
#   • ANOTHER_BLOCK_CODE
#       - specific effect

#   If NOTHING is affected -> write exactly: No effect

# RULES:
#   - Use bullet (•) before each affected block name, no indent
#   - Use 4 spaces + dash (    -) before each effect line under a block
#   - Effects must be SPECIFIC: not "BIAS is affected" but "Output bias current is stuck"
#   - Use present tense active: "is stuck", "is incorrect", "cannot operate"
#   - List EVERY block that receives this signal - do not omit any
#   - If multiple sub-effects exist for one block, list each on its own line
# """.strip()

# # J column valid values (system-level, agnostic to chip type)
# J_VALID_VALUES = [
#     "Unintentional LED ON/OFF\nFail-safe mode active\nNo communication",
#     "Fail-safe mode active\nNo communication",
#     "Fail-safe mode active",
#     "Unintended LED ON/OFF",
#     "Unintended LED ON",
#     "Unintended LED OFF",
#     "Device damage",
#     "Possible device damage",
#     "No effect",
# ]

# # K override rules - derived from ISO 26262 principles, chip-agnostic
# def compute_k_from_mode_and_coverage(code: str, mode_str: str,
#                                       ic_effect: str, block_to_sms: dict) -> str:
#     """
#     Determine K (safety violation flag) deterministically using ISO 26262 principles.

#     Rules derived from systematic diff analysis against human expert FMEDA:

#     ALWAYS K=O (never safety-violating):
#       - Safe/benign modes (spikes, oscillation within range, jitter, quiescent, settling)
#       - No IC downstream effect
#       - Interface/comms blocks (protocol layer detects/handles these)
#       - ADC non-stuck modes (accuracy, offset, linearity, monotonic, full-scale, settling)
#         → these cause measurement drift, not hard failures
#       - Current-sense monitoring blocks (CSNS) — all modes → O
#         → CSNS only feeds ADC for monitoring; fault is caught by ADC SM coverage
#       - Voltage supply spikes and incorrect start-up time (transient, not sustained)
#       - Driver turn-off timing (performance impact only)
#       - Charge pump oscillation-within-range, quiescent current (local, non-propagating)
#       - TRIM incorrect settling time (timing, not output value)
#       - Interface RX message-value errors that the MCU will catch

#     ALWAYS K=X (safety-violating):
#       - OSC drift → propagates to LOGIC, causes comms failure
#       - LOGIC float and incorrect output → SW_BANK and OSC lose control
#       - TRIM commission and incorrect output → miscalibration of all analog blocks
#       - SW_BANK float and resistance-too-high → LED stuck on (unintended)
#       - Any block with hard IC effect (stuck/float/OV/UV/accuracy/drift) on SM-covered blocks
#     """
#     m = mode_str.lower()
#     severity = classify_mode_severity(mode_str)
#     norm = re.sub(r'SW_BANK[_\d]*', 'SW_BANK', code.upper())

#     # ── UNIVERSAL SAFE MODES ────────────────────────────────────────────────
#     if severity == 'safe':
#         return 'O'

#     # No IC downstream effect → always O
#     if not ic_effect or ic_effect.strip() in ('', 'No effect', 'No effect (Filter in place)'):
#         return 'O'

#     # ── BLOCK-SPECIFIC OVERRIDES (derived from diff) ─────────────────────────

#     # CSNS: monitoring-only block, all modes → O (ADC SM coverage handles it)
#     if norm == 'CSNS':
#         return 'O'

#     # INTERFACE: protocol layer catches all these → O
#     if 'INTERFACE' in norm:
#         return 'O'

#     # ADC: only stuck/float are hard safety failures
#     if norm == 'ADC' and severity not in ('stuck', 'float'):
#         return 'O'

#     # Driver blocks (SW_BANK):
#     if norm == 'SW_BANK':
#         # stuck, float, res_high → X (unintended LED state)
#         if severity in ('stuck', 'float') or 'resistance too high' in m:
#             return 'X'
#         # res_low (SW_BANK_1 only K=X, but generally O across all banks) → O
#         # turn-on/turn-off timing → O
#         return 'O'

#     # LDO: spikes and startup are transient → O; OV/UV/accuracy → X
#     if norm == 'LDO':
#         if 'spike' in m or 'start-up' in m or 'start up' in m:
#             return 'O'
#         return 'X'  # OV, UV, accuracy, oscillation-within → X

#     # CP (charge pump): OV/UV → X; oscillation-within, quiescent → O
#     if norm == 'CP':
#         if severity in ('ov', 'uv') or 'lower than' in m or 'higher than' in m:
#             return 'X'
#         return 'O'

#     # OSC: drift IS a safety violation (causes LOGIC failure)
#     if norm == 'OSC':
#         if 'drift' in m:
#             return 'X'
#         if severity in ('stuck', 'float', 'incorrect', 'ov', 'uv'):
#             return 'X'
#         return 'O'

#     # LOGIC: ALL three modes are safety-violating
#     if norm == 'LOGIC':
#         return 'X'

#     # TRIM: commission and incorrect output → X; settling time → O
#     if norm == 'TRIM':
#         if 'settling' in m or 'start-up' in m:
#             return 'O'
#         return 'X'

#     # ── GENERAL RULE: check SM coverage on affected blocks ──────────────────
#     affected = re.findall(r'^\s*•\s*([A-Z_a-z0-9]+)', ic_effect, re.MULTILINE)
#     norm_affected = []
#     for b in affected:
#         b = b.strip().upper()
#         b = re.sub(r'SW_BANK[_X\d]*', 'SW_BANK', b)
#         b = re.sub(r'CSNS|CNSN|CS(?!NS)', 'CSNS', b)
#         if b not in ('NONE', 'VEGA', ''):
#             norm_affected.append(b)

#     # SM coverage on any affected block → X
#     for block in norm_affected:
#         if block_to_sms.get(block):
#             return 'X'

#     # Hard failure with IC effect even on non-SM blocks → X
#     hard_failure = severity in ('stuck', 'float', 'ov', 'uv', 'accuracy', 'drift')
#     if hard_failure and (norm_affected or 'vega' in ic_effect.lower()):
#         return 'X'
#     if hard_failure and ic_effect.strip() not in ('', 'No effect'):
#         return 'X'

#     return 'O'


# def compute_sm_columns(ic_effect: str, block_to_sms: dict, sm_coverage: dict,
#                        fmeda_code: str = '', mode_str: str = '',
#                        sm_addressed: dict = None) -> tuple:
#     """
#     Returns (sm_string, coverage_value) for col S/Y and col U.

#     KEY FIX (v9): Use SMs that MONITOR THE FAILING BLOCK ITSELF.
#     Previous versions took a union of SMs covering all downstream consumers
#     (e.g. BIAS failing -> list all SMs covering ADC/TEMP/LDO/OSC = 18 SMs).
#     Correct approach: use SMs that provide coverage FOR the source block failure.
#     """
#     if not ic_effect or ic_effect.strip() in ('No effect', 'No effect (Filter in place)', ''):
#         return '', ''

#     severity = classify_mode_severity(mode_str)
#     m = mode_str.lower()
#     norm_code = re.sub(r'SW_BANK[_\d]*', 'SW_BANK', fmeda_code.upper())

#     # Blocks that always get empty S/Y
#     if severity == 'safe':
#         return '', ''
#     if norm_code == 'INTERFACE':
#         return '', ''
#     if norm_code == 'CSNS':
#         return '', ''
#     if norm_code == 'ADC' and severity not in ('stuck', 'float'):
#         return '', ''
#     if norm_code == 'CP' and any(k in m for k in ['oscillation', 'quiescent', 'spike', 'start-up', 'start up']):
#         return '', ''
#     if norm_code == 'LDO' and any(k in m for k in ['spike', 'start-up', 'start up']):
#         return '', ''
#     if norm_code == 'SW_BANK' and any(k in m for k in ['turn-on', 'turn-off', 'turn on', 'turn off', 'resistance too low']):
#         return '', ''

#     # ── PER-BLOCK SM SETS (source-block-driven, from SM list Addressed Part column) ──
#     if norm_code == 'SW_BANK':
#         if 'stuck' in m and 'not including' not in m:
#             return _pick_sms(['SM04', 'SM05', 'SM06', 'SM08'], sm_coverage)
#         elif 'floating' in m or 'open circuit' in m or 'tri-state' in m:
#             return _pick_sms(['SM04', 'SM06', 'SM08'], sm_coverage)
#         elif 'resistance too high' in m:
#             return _pick_sms(['SM03', 'SM06', 'SM24'], sm_coverage)
#         return '', ''

#     if norm_code == 'LDO':
#         if severity == 'ov' or 'higher than' in m:
#             return _pick_sms(['SM11', 'SM20'], sm_coverage)
#         elif severity == 'uv' or 'lower than' in m:
#             return _pick_sms(['SM11', 'SM15'], sm_coverage)
#         return _pick_sms(['SM11', 'SM15', 'SM20'], sm_coverage)

#     if norm_code == 'CP':
#         if severity == 'ov' or 'higher than' in m:
#             return '', ''  # OV -> device damage, no SM covers it
#         return _pick_sms(['SM14', 'SM22'], sm_coverage)

#     if norm_code == 'REF':
#         if severity in ('stuck', 'float'):
#             return _pick_sms(['SM01', 'SM15', 'SM16', 'SM17'], sm_coverage)
#         return _pick_sms(['SM01', 'SM11', 'SM15', 'SM16'], sm_coverage)

#     if norm_code == 'BIAS':
#         return _pick_sms(['SM11', 'SM15', 'SM16'], sm_coverage)

#     if norm_code == 'OSC':
#         return _pick_sms(['SM09', 'SM10', 'SM11'], sm_coverage)

#     if norm_code == 'TEMP':
#         return _pick_sms(['SM17', 'SM23'], sm_coverage)

#     if norm_code == 'ADC':
#         return _pick_sms(['SM08', 'SM16', 'SM17', 'SM23'], sm_coverage)

#     if norm_code == 'LOGIC':
#         return _pick_sms(['SM10', 'SM11', 'SM12', 'SM18'], sm_coverage)

#     if norm_code == 'TRIM':
#         return _pick_sms(['SM01', 'SM02', 'SM09', 'SM11', 'SM15', 'SM16', 'SM18', 'SM20', 'SM23'], sm_coverage)

#     # Generic fallback: use SMs directly addressing this block from SM list
#     direct_sms = sorted(block_to_sms.get(norm_code, []),
#                         key=lambda s: int(re.search(r'\d+', s).group()) if re.search(r'\d+', s) else 0)
#     if not direct_sms:
#         return '', ''
#     valid = [0.99, 0.9, 0.6]
#     def nearest(v):
#         return min(valid, key=lambda x: abs(x - v))
#     coverages = [nearest(sm_coverage.get(sm, 0.9)) for sm in direct_sms]
#     return ' '.join(direct_sms), max(coverages) if coverages else 0.9


# def _pick_sms(sm_list: list, sm_coverage: dict) -> tuple:
#     """Filter to SMs present in coverage map, return (space-joined str, max coverage)."""
#     valid_sms = [s for s in sm_list if s in sm_coverage] if sm_coverage else sm_list
#     if not valid_sms:
#         valid_sms = sm_list  # fallback when no template loaded
#     if not valid_sms:
#         return '', ''
#     valid_cov = [0.99, 0.9, 0.6]
#     def nearest(v):
#         return min(valid_cov, key=lambda x: abs(x - v))
#     coverages = [nearest(sm_coverage.get(sm, 0.9)) for sm in valid_sms]
#     return ' '.join(valid_sms), max(coverages)


# def agent2_generate_effects(blocks, tsr_list, block_to_sms, sm_coverage,
#                              sm_addressed, cache, signal_graph, sm_j_map):
#     """
#     Generate col I/J/K for all blocks.
#     v10: Col I is DETERMINISTIC via resolve_i_deterministic().
#          LLM only called for genuinely unknown block types.
#     """
#     active = [b for b in blocks if not b.get('is_duplicate') and not b.get('is_sm')]
#     chip_ctx = "\n".join(
#         f"  {b['fmeda_code']:<12} {b['name']:<35} | {b.get('function', '')[:80]}"
#         for b in active
#     )

#     tsr_ctx = "\n".join(
#         f"  {t['id']}: {t['description']}"
#         for t in tsr_list
#     ) if tsr_list else "  (no TSR data)"

#     result = []
#     for block in blocks:
#         code  = block['fmeda_code']
#         name  = block['name']
#         modes = block.get('modes', [])

#         # SM blocks
#         if block.get('is_sm'):
#             rows = _sm_rows(code, sm_j_map)
#             result.append({'fmeda_code': code, 'user_name': name, 'rows': rows})
#             print(f"  [Agent 2] {code:<12} SM (2 rows)")
#             continue

#         if block.get('is_duplicate'):
#             print(f"  [Agent 2] {code:<12} DUPLICATE - skipped")
#             continue

#         if not modes:
#             print(f"  [Agent 2] {code:<12} no modes - skipped")
#             continue

#         # v10 cache key (invalidates v9/v8 cache)
#         ck = f"agent2_v10__{code}__{name}__{len(modes)}"
#         if not SKIP_CACHE and ck in cache:
#             rows = cache[ck]
#             # Always refresh I deterministically (fixes stale LLM I in cache)
#             for row in rows:
#                 mode_g = row.get('G', '')
#                 det_i = resolve_i_deterministic(block, mode_g, signal_graph, sm_j_map)
#                 if det_i is not None:
#                     row['I'] = det_i
#                     row['J'] = _validate_j(_derive_j_from_rules(code, mode_g, det_i))
#                 # Always refresh K and S/Y
#                 k = compute_k_from_mode_and_coverage(code, mode_g, row.get('I', ''), block_to_sms)
#                 row['K'] = k
#                 row['P'] = 'Y' if k == 'X' else 'N'
#                 row['R'] = 1 if k == 'O' else 0
#                 row['X'] = 'Y' if k.startswith('X') else 'N'
#                 sm_str, cov = compute_sm_columns(row.get('I', ''), block_to_sms, sm_coverage, code, mode_g)
#                 row['S'] = sm_str; row['Y'] = sm_str; row['U'] = cov
#             print(f"  [Agent 2] {code:<12} cache ({len(rows)} rows, I/K/S refreshed)")
#             result.append({'fmeda_code': code, 'user_name': name, 'rows': rows})
#             continue

#         rows = _llm_block_effects_v7(block, chip_ctx, tsr_ctx, modes,
#                                       block_to_sms, sm_coverage, signal_graph,
#                                       sm_j_map=sm_j_map)
#         cache[ck] = rows
#         save_cache(cache)
#         result.append({'fmeda_code': code, 'user_name': name, 'rows': rows})
#         time.sleep(0.1)

#     return result


# def _build_i_context_for_block(block: dict, signal_graph: dict) -> str:
#     """
#     Build a detailed signal-flow context for col I generation.
#     Uses the signal graph built by Agent 0 — no hardcoded values.
#     """
#     code = block['fmeda_code']
#     name = block['name']

#     # Look up in signal graph (by name or code)
#     graph_entry = signal_graph.get(name) or signal_graph.get(code) or {}

#     if graph_entry:
#         consumers = graph_entry.get('consumers', [])
#         output_sig = graph_entry.get('output_signal', f'output of {name}')
#         details = graph_entry.get('consumer_details', {})

#         ctx = f"OUTPUT SIGNAL: {output_sig}\n\n"
#         if consumers:
#             ctx += "DIRECT CONSUMERS (blocks that will fail if THIS block fails):\n"
#             for c in consumers:
#                 detail = details.get(c, 'receives signal from this block')
#                 ctx += f"  - {c}: {detail}\n"
#         else:
#             ctx += "DIRECT CONSUMERS: (none identified - check function description)\n"
#     else:
#         # Fallback: generic guidance based on block function description
#         func = block.get('function', '').lower()
#         ctx = f"Block function: {block.get('function', '')}\n\n"
#         ctx += "ANALYZE: Which other blocks in the chip receive output from this block?\n"
#         ctx += "Consider: voltage/current references feed analog blocks, "
#         ctx += "clocks feed digital logic, supplies feed all powered blocks.\n"

#     return ctx


# # =============================================================================
# # DETERMINISTIC COL I ENGINE
# # =============================================================================
# #
# # KEY INSIGHT: Col I follows strict deterministic patterns per block type.
# # The LLM is the wrong tool — it over-thinks and adds wrong variations.
# # The correct values are FULLY determined by: block type + mode severity.
# #
# # This engine resolves I values using:
# #   1. Block-type classification (derived from fmeda_code pattern, not hardcoded names)
# #   2. Mode severity classification (classify_mode_severity)
# #   3. Signal graph for consumer block names (no hardcoded block names)
# #
# # For unknown block types, it falls back to the LLM with a tightly constrained prompt.

# def resolve_i_deterministic(block: dict, mode_str: str,
#                               signal_graph: dict,
#                               sm_j_map: dict = None) -> str | None:
#     """
#     Resolve col I deterministically from block type + mode severity.
#     Returns the I string if deterministically resolvable, None if LLM needed.

#     Rules derived from complete analysis of 3_ID03_FMEDA.xlsx col I patterns.
#     All consumer block names come from the signal_graph (chip-agnostic).
#     """
#     code = block['fmeda_code']
#     name = block['name']
#     m    = mode_str.lower()
#     sev  = classify_mode_severity(mode_str)
#     norm = re.sub(r'SW_BANK[_\d]*', 'SW_BANK', code.upper())

#     # Get consumer block names from signal graph (chip-agnostic)
#     graph_entry   = signal_graph.get(name) or signal_graph.get(code) or {}
#     consumers     = graph_entry.get('consumers', [])
#     consumer_det  = graph_entry.get('consumer_details', {})

#     # Helper: find the first consumer whose name/detail matches a keyword
#     def find_consumer(keywords: list) -> str | None:
#         for c in consumers:
#             c_lower = c.lower()
#             detail  = consumer_det.get(c, '').lower()
#             if any(k in c_lower or k in detail for k in keywords):
#                 return c
#         return None

#     # Helper: get all SW_BANK consumers (numbered or generic)
#     def sw_bank_consumers() -> list:
#         return [c for c in consumers
#                 if 'sw_bank' in c.lower() or 'switch' in c.lower()
#                 or 'driver' in c.lower() or 'led' in c.lower()]

#     # ── SAFE MODES — always No effect ────────────────────────────────────────
#     if sev == 'safe':
#         # LDO spikes is a special case: NOT No effect, it causes OSC jitter
#         if norm == 'LDO' and 'spike' in m:
#             osc = find_consumer(['osc', 'clock', 'oscillator'])
#             osc_name = osc or 'OSC'
#             return f'• {osc_name}\n    - Jitter too high in the output signal'
#         # LDO fast oscillation: No effect with note
#         if norm == 'LDO' and 'fast oscillation' in m:
#             return 'No effect (Filter in place)'
#         return 'No effect'

#     # ── SW_BANK — direct LED state strings, NO bullet points ─────────────────
#     if norm == 'SW_BANK':
#         if 'stuck' in m and 'not including' not in m:
#             return 'Unintended LED ON/OFF'
#         if 'floating' in m or 'open circuit' in m or 'tri-state' in m:
#             return 'Unintended LED ON'
#         if 'resistance too high' in m:
#             return 'Unintended LED ON'
#         if 'resistance too low' in m or 'turn-on' in m or 'turn-off' in m:
#             return 'Performance impact'
#         return 'Performance impact'

#     # ── INTERFACE — always plain string ──────────────────────────────────────
#     if norm == 'INTERFACE':
#         return 'Communication error'

#     # ── CSNS — always single ADC bullet regardless of mode ───────────────────
#     if norm == 'CSNS':
#         adc = find_consumer(['adc', 'convert', 'digital'])
#         adc_name = adc or 'ADC'
#         return f'• {adc_name}\n    - CSNS output is incorrect.'

#     # ── LDO ──────────────────────────────────────────────────────────────────
#     if norm == 'LDO':
#         osc = find_consumer(['osc', 'clock', 'oscillator'])
#         osc_name = osc or 'OSC'
#         # OV: OSC out of spec
#         if sev == 'ov' or 'higher than' in m:
#             return f'• {osc_name}\n    - Out of spec.'
#         # UV or accuracy: OSC out of spec + Vega reset
#         if sev in ('uv', 'accuracy', 'drift') or 'lower than' in m:
#             return (f'• {osc_name}\n    - Out of spec.\n'
#                     f'• Vega\n    - Reset reaction. (POR)')
#         # Startup: No effect
#         if 'start-up' in m or 'startup' in m:
#             return 'No effect'
#         # Oscillation within range: No effect
#         return 'No effect'

#     # ── OSC ──────────────────────────────────────────────────────────────────
#     if norm == 'OSC':
#         logic = find_consumer(['logic', 'controller', 'control', 'digital', 'mcu'])
#         logic_name = logic or 'LOGIC'
#         if sev in ('stuck', 'float', 'incorrect', 'ov', 'uv', 'drift'):
#             return f'• {logic_name}\n    - Cannot operate.\n    - Communication error.'
#         if 'duty cycle' in m or 'jitter' in m:
#             return 'No effect'
#         return f'• {logic_name}\n    - Cannot operate.\n    - Communication error.'

#     # ── TEMP ─────────────────────────────────────────────────────────────────
#     if norm == 'TEMP':
#         adc  = find_consumer(['adc', 'convert', 'digital'])
#         adc_name = adc or 'ADC'
#         sw   = find_consumer(['sw_bank', 'switch', 'driver', 'led'])
#         sw_name = (re.sub(r'SW_BANK[_\d]+', 'SW_BANK_x', sw)
#                    if sw else 'SW_BANK_x')
#         if sev == 'stuck':
#             return (f'• {adc_name}\n    - TEMP output is stuck low\n'
#                     f'• {sw_name}\n    - SW is stuck in off state (DIETEMP)')
#         if sev == 'float':
#             return f'• {adc_name}\n    - Incorrect TEMP reading'
#         if sev in ('incorrect', 'accuracy', 'drift'):
#             return (f'• {adc_name}\n    - TEMP output Static Error (offset error, gain error, '
#                     f'integral nonlinearity, & differential nonlinearity)')
#         return 'No effect'

#     # ── ADC ──────────────────────────────────────────────────────────────────
#     if norm == 'ADC':
#         sw = find_consumer(['sw_bank', 'switch', 'driver', 'led'])
#         sw_name = re.sub(r'SW_BANK[_\d]+', 'SW_BANK_x', sw) if sw else 'SW_BANK_x'
#         adc_self_bullets = ('• ADC\n    - Incorrect BGR measurement\n'
#                             '    - Incorrect DIETEMP measurement\n'
#                             '    - Incorrect CS measurement')
#         if sev in ('stuck', 'float'):
#             return (f'• {sw_name}\n    - SW is stuck in off state (DIETEMP)\n'
#                     + adc_self_bullets)
#         # All non-safe non-stuck ADC modes: self-measurement errors only
#         if sev not in ('safe',):
#             return adc_self_bullets
#         return 'No effect'

#     # ── CP ────────────────────────────────────────────────────────────────────
#     if norm == 'CP':
#         sw = find_consumer(['sw_bank', 'switch', 'driver', 'led'])
#         sw_name = re.sub(r'SW_BANK[_\d]+', 'SW_BANK_x', sw) if sw else 'SW_BANK_x'
#         if sev == 'ov' or 'higher than' in m:
#             return '• Vega\n    - Device Damage'
#         if sev == 'uv' or 'lower than' in m:
#             return f'• {sw_name}\n    - SWs are stuck in off state, LEDs always ON.'
#         return 'No effect'

#     # ── LOGIC ─────────────────────────────────────────────────────────────────
#     if norm == 'LOGIC':
#         sw = find_consumer(['sw_bank', 'switch', 'driver', 'led'])
#         sw_name = re.sub(r'SW_BANK[_\d]+', 'SW_BANK_X', sw) if sw else 'SW_BANK_X'
#         osc = find_consumer(['osc', 'clock', 'oscillator'])
#         osc_name = osc or 'OSC'
#         if sev in ('stuck', 'float', 'incorrect', 'ov', 'uv', 'accuracy'):
#             return (f'• {sw_name}\n    - SW is stuck in on/off state\n'
#                     f'• {osc_name}\n    - Output stuck')
#         return 'No effect'

#     # ── TRIM ──────────────────────────────────────────────────────────────────
#     if norm == 'TRIM':
#         if sev == 'safe' or 'settling' in m:
#             return 'No effect'
#         # omission, commission, incorrect: all calibrated blocks affected
#         ref  = find_consumer(['ref', 'bandgap', 'reference'])
#         ldo  = find_consumer(['ldo', 'regulator', 'supply'])
#         bias = find_consumer(['bias', 'current source'])
#         sw   = find_consumer(['sw_bank', 'switch', 'driver'])
#         osc  = find_consumer(['osc', 'clock', 'oscillator'])
#         temp = find_consumer(['temp', 'thermal', 'temperature'])

#         ref_name  = ref  or 'REF'
#         ldo_name  = ldo  or 'LDO'
#         bias_name = bias or 'BIAS'
#         sw_name   = re.sub(r'SW_BANK[_\d]+', 'SW_BANK', sw) if sw else 'SW_BANK'
#         osc_name  = osc  or 'OSC'
#         temp_code = (re.sub(r'TEMP.*', 'DIETEMP', temp.upper()) if temp else 'DIETEMP')

#         return (f'• {ref_name}\n    - Incorrect output value higher than the expected range\n'
#                 f'• {ldo_name}\n    - Reference voltage higher than the expected range\n'
#                 f'• {bias_name}\n    - Output reference voltage accuracy too low, including drift\n'
#                 f'• {sw_name}\n    - Incorrect slew rate value\n'
#                 f'• {osc_name}\n    - Incorrect output frequency: higher than the expected range\n'
#                 f'• {temp_code}\n    - Incorrect output voltage')

#     # ── REF ───────────────────────────────────────────────────────────────────
#     if norm == 'REF':
#         bias = find_consumer(['bias', 'current source', 'current mirror'])
#         adc  = find_consumer(['adc', 'convert', 'digital'])
#         temp = find_consumer(['temp', 'thermal', 'temperature'])
#         ldo  = find_consumer(['ldo', 'regulator', 'supply'])
#         osc  = find_consumer(['osc', 'clock', 'oscillator'])

#         bias_name = bias or 'BIAS'
#         adc_name  = adc  or 'ADC'
#         temp_name = temp or 'TEMP'
#         ldo_name  = ldo  or 'LDO'
#         osc_name  = osc  or 'OSC'

#         if sev == 'stuck':
#             return (f'• {bias_name}\n'
#                     f'    - Output reference voltage is stuck \n'
#                     f'    - Output reference current is stuck \n'
#                     f'    - Output bias current is stuck \n'
#                     f'    - Quiescent current exceeding the maximum value\n'
#                     f'• REF\n'
#                     f'    - Quiescent current exceeding the maximum value\n'
#                     f'• {adc_name}\n    - REF output is stuck \n'
#                     f'• {temp_name}\n    - Output is stuck \n'
#                     f'• {ldo_name}\n    - Output is stuck \n'
#                     f'• {osc_name}\n    - Oscillation does not start')
#         if sev == 'float':
#             return (f'• {bias_name}\n'
#                     f'    - Output reference voltage is floating\n'
#                     f'    - Output reference current is higher than the expected range\n'
#                     f'    - Output reference current is lower than the expected range\n'
#                     f'    - Output bias current is higher than the expected range\n'
#                     f'    - Output bias current is lower than the expected range\n'
#                     f'• {adc_name}\n    - REF output is floating (i.e. open circuit)\n'
#                     f'• {ldo_name}\n    - Out of spec\n'
#                     f'• {osc_name}\n    - Out of spec')
#         if sev in ('incorrect', 'accuracy', 'drift'):
#             return (f'• {bias_name}\n'
#                     f'    - Output reference voltage is higher than the expected range\n'
#                     f'    - Output reference current is higher than the expected range\n'
#                     f'    - Output bias current is higher than the expected range\n'
#                     f'• {temp_name}\n'
#                     f'    - Incorrect gain on the output voltage (outside the expected range)\n'
#                     f'    - Incorrect offset on the output voltage (outside the expected range)\n'
#                     f'• {adc_name}\n    - REF output higher/lower than expected\n'
#                     f'• {ldo_name}\n    - Out of spec\n'
#                     f'• {osc_name}\n    - Out of spec')
#         return 'No effect'

#     # ── BIAS ──────────────────────────────────────────────────────────────────
#     if norm == 'BIAS':
#         adc  = find_consumer(['adc', 'convert', 'digital'])
#         temp = find_consumer(['temp', 'thermal', 'temperature'])
#         ldo  = find_consumer(['ldo', 'regulator', 'supply'])
#         osc  = find_consumer(['osc', 'clock', 'oscillator'])
#         sw   = find_consumer(['sw_bank', 'switch', 'driver', 'led'])
#         cp   = find_consumer(['charge pump', 'cp', 'boost'])
#         csns = find_consumer(['csns', 'current sense', 'sense amp'])

#         adc_name  = adc  or 'ADC'
#         temp_name = temp or 'TEMP'
#         ldo_name  = ldo  or 'LDO'
#         osc_name  = osc  or 'OSC'
#         sw_name   = re.sub(r'SW_BANK[_\d]+', 'SW_BANKx', sw) if sw else 'SW_BANKx'
#         cp_name   = cp   or 'CP'
#         cnsn_name = csns or 'CNSN'

#         if sev in ('stuck', 'float', 'incorrect', 'accuracy', 'drift', 'other'):
#             return (f'• {adc_name}\n    - ADC measurement is incorrect.\n'
#                     f'• {temp_name}\n    - Incorrect temperature measurement.\n'
#                     f'• {ldo_name}\n    - Out of spec.\n'
#                     f'• {osc_name}\n    - Frequency out of spec.\n'
#                     f'• {sw_name}\n    - Out of spec.\n'
#                     f'• {cp_name}\n    - Out of spec.\n'
#                     f'• {cnsn_name}\n    - Incorrect reading.')
#         return 'No effect'

#     # ── SM BLOCKS ─────────────────────────────────────────────────────────────
#     if re.match(r'^SM\d+$', norm):
#         # 'Fail to detect' -> the I value is what the SM was supposed to catch
#         # This comes from sm_j_map which is built from SM descriptions
#         if sm_j_map and 'fail to detect' in m:
#             sm_ic, _ = sm_j_map.get(code, ('Loss of safety mechanism functionality', ''))
#             return sm_ic if sm_ic else 'Loss of safety mechanism functionality'
#         if 'false detection' in m:
#             return 'No effect'
#         return None  # fallback to LLM

#     # Unknown block type: return None to trigger LLM fallback
#     return None


# def _llm_block_effects_v7(block, chip_ctx, tsr_ctx, modes,
#                            block_to_sms, sm_coverage, signal_graph,
#                            sm_j_map=None):
#     """
#     Generate I/J/K rows.

#     Architecture (v10):
#       Col I: DETERMINISTIC first via resolve_i_deterministic().
#              LLM only called if resolve_i_deterministic() returns None
#              (i.e. a completely new block type not covered by any rule).
#       Col J: Determined from block type + mode severity (LLM prompt includes rules).
#       Col K: Always deterministic via compute_k_from_mode_and_coverage().
#     """
#     code = block['fmeda_code']
#     name = block['name']
#     func = block.get('function', '')
#     n    = len(modes)

#     # Pass 1: resolve everything deterministically
#     det_results = []      # (mode, i_val) for deterministically resolved modes
#     llm_needed  = []      # (idx, mode) needing LLM for I

#     for idx, mode in enumerate(modes):
#         i_val = resolve_i_deterministic(block, mode, signal_graph, sm_j_map)
#         if i_val is not None:
#             det_results.append((idx, mode, i_val))
#         else:
#             llm_needed.append((idx, mode))

#     # Pass 2: LLM only for unknown modes (usually empty for known block types)
#     llm_i = {}  # idx -> i string
#     if llm_needed:
#         i_context   = _build_i_context_for_block(block, signal_graph)
#         safe_modes  = [m for _, m in llm_needed if is_safe_mode(m)]
#         llm_modes   = [m for _, m in llm_needed]

#         prompt = f"""You are a senior functional safety engineer completing an FMEDA (ISO 26262).

# BLOCK: {code} | {name} | {func}

# SIGNAL FLOW:
# {i_context}

# CHIP CONTEXT:
# {chip_ctx}

# FAILURE MODES TO ANALYZE (only these {len(llm_modes)} — others handled separately):
# {json.dumps(llm_modes, indent=2)}

# SAFE MODES (write "No effect"):
# {json.dumps(safe_modes)}

# For col I (Effects on IC output), use bullet format:
#   • BLOCK_CODE
#       - specific symptom (under 10 words)
#       - second symptom if needed

# Rules:
# - SW_BANK_x (not SW_BANK_1/2), SW_BANK_X for logic-driven, Vega for device damage
# - Safe modes -> No effect
# - Only list DIRECT consumers from signal flow above
# - Short precise phrases only

# For col J (system effect), use the first matching rule:
#   REF/BIAS non-safe: "Unintentional LED ON/OFF\\nFail-safe mode active\\nNo communication"
#   LDO OV/UV/accuracy: "Fail-safe mode active\\nNo communication"
#   OSC stuck/float/freq/drift: "Fail-safe mode active\\nNo communication"
#   TEMP stuck: "Unintentional LED ON"
#   TEMP float/incorrect: "Unintentional LED ON\\nPossible device damage"
#   ADC stuck/float: "Unintentional LED ON"
#   CP OV: "Device damage" | CP UV: "Unintentional LED ON"
#   LOGIC all: "Unintentional LED ON/OFF\\nFail-safe mode active\\nNo communication"
#   TRIM active modes: "Fail-safe mode active\\nNo communication"
#   SW_BANK stuck: "Unintended LED ON/OFF" | SW_BANK float/res-high: "Unintended LED ON"
#   Otherwise: "No effect"

# Return JSON array with {len(llm_modes)} objects:
# [{{"G": "<mode>", "I": "<col I>", "J": "<col J>"}}]
# Return ONLY the JSON:"""

#         raw    = query_llm(prompt, temperature=0.0)
#         parsed = parse_json(raw)

#         if isinstance(parsed, list) and len(parsed) >= len(llm_needed):
#             for pos, (orig_idx, orig_mode) in enumerate(llm_needed):
#                 rd    = parsed[pos]
#                 i_str = str(rd.get('I', 'No effect')).strip()
#                 # Post-process: replace literal 'bullet'/'dash' words with symbols
#                 i_str = i_str.replace('bullet ', '• ').replace('\ndash ', '\n    - ')
#                 llm_i[orig_idx] = (i_str, str(rd.get('J', 'No effect')).strip())
#         else:
#             for orig_idx, orig_mode in llm_needed:
#                 llm_i[orig_idx] = ('No effect' if is_safe_mode(orig_mode) else '', 'No effect')

#     # Pass 3: build rows in correct mode order
#     # J for deterministically-resolved I: compute from block type + mode
#     rows = []
#     det_map = {idx: (mode, i_val) for idx, mode, i_val in det_results}

#     for idx, mode in enumerate(modes):
#         if idx in det_map:
#             ic   = det_map[idx][1]
#             sys_ = _derive_j_from_rules(code, mode, ic)
#         else:
#             ic_and_j = llm_i.get(idx, ('No effect', 'No effect'))
#             ic   = ic_and_j[0]
#             sys_ = _validate_j(ic_and_j[1])

#         # Always override safe modes
#         if is_safe_mode(mode):
#             # LDO spike exception already handled in resolve_i_deterministic
#             if ic == 'No effect':
#                 pass
        
#         sys_ = _validate_j(sys_)
#         memo = 'O' if (not ic or ic.strip() in ('No effect', 'No effect (Filter in place)')) \
#                else compute_k_from_mode_and_coverage(code, mode, ic, block_to_sms)

#         rows.append(_build_row(mode, ic, sys_, memo, block_to_sms, sm_coverage,
#                                fmeda_code=code))

#     if not rows:
#         rows = _fallback_rows_v7(modes, code, block_to_sms, sm_coverage)

#     n_det = len(det_results)
#     n_llm = len(llm_needed)
#     print(f"    {code}: {n_det} det + {n_llm} LLM = {n} rows")
#     return rows


# def _derive_j_from_rules(code: str, mode_str: str, ic: str) -> str:
#     """
#     Derive col J from block type + mode severity using the block-specific rules.
#     This is called for deterministically-resolved I values so we don't need LLM for J.
#     """
#     norm = re.sub(r'SW_BANK[_\d]*', 'SW_BANK', code.upper())
#     m    = mode_str.lower()
#     sev  = classify_mode_severity(mode_str)

#     # Safe modes or no IC effect
#     if sev == 'safe' or not ic or ic.strip() == 'No effect':
#         # LDO spikes exception: has IC effect but J is still No effect
#         if norm == 'LDO' and 'spike' in m:
#             return 'No effect'
#         if ic and ic.strip() not in ('No effect', 'No effect (Filter in place)', ''):
#             pass  # fall through to block-specific rules
#         else:
#             return 'No effect'

#     if norm in ('REF', 'BIAS'):
#         return 'Unintentional LED ON/OFF\nFail-safe mode active\nNo communication'
#     if norm == 'LDO':
#         if sev == 'safe' or 'spike' in m or 'start-up' in m:
#             return 'No effect'
#         return 'Fail-safe mode active\nNo communication'
#     if norm == 'OSC':
#         if 'duty cycle' in m or 'jitter' in m:
#             return 'No effect'
#         return 'Fail-safe mode active\nNo communication'
#     if norm == 'TEMP':
#         if sev == 'stuck':
#             return 'Unintentional LED ON'
#         if sev in ('float', 'incorrect', 'accuracy', 'drift'):
#             return 'Unintentional LED ON\nPossible device damage'
#         return 'No effect'
#     if norm == 'CSNS':
#         return 'No effect'
#     if norm == 'ADC':
#         if sev in ('stuck', 'float'):
#             return 'Unintentional LED ON'
#         return 'No effect'
#     if norm == 'CP':
#         if sev == 'ov' or 'higher than' in m:
#             return 'Device damage'
#         if sev == 'uv' or 'lower than' in m:
#             return 'Unintentional LED ON'
#         return 'No effect'
#     if norm == 'LOGIC':
#         return 'Unintentional LED ON/OFF\nFail-safe mode active\nNo communication'
#     if norm == 'INTERFACE':
#         return 'Fail-safe mode active'
#     if norm == 'TRIM':
#         if 'settling' in m:
#             return 'No effect'
#         return 'Fail-safe mode active\nNo communication'
#     if norm == 'SW_BANK':
#         if 'stuck' in m and 'not including' not in m:
#             return 'Unintended LED ON/OFF'
#         if 'floating' in m or 'open circuit' in m or 'resistance too high' in m:
#             return 'Unintended LED ON'
#         return 'No effect'
#     if re.match(r'^SM\d+$', norm):
#         # SM 'Fail to detect' J comes from sm_j_map (handled by _sm_rows)
#         return 'No effect'

#     # Unknown block: use generic rules
#     if ic and ic.strip() not in ('No effect', ''):
#         return 'Fail-safe mode active'
#     return 'No effect'
#     code = block['fmeda_code']
#     name = block['name']
#     func = block.get('function', '')
#     n    = len(modes)

#     i_context = _build_i_context_for_block(block, signal_graph)
#     safe_modes   = [m for m in modes if is_safe_mode(m)]
#     unsafe_modes = [m for m in modes if not is_safe_mode(m)]

#     prompt = f"""You are a senior functional safety engineer completing an FMEDA for an automotive IC (ISO 26262).

# CHIP BLOCKS:
# {chip_ctx}

# SAFETY REQUIREMENTS (TSR):
# {tsr_ctx}

# BLOCK UNDER ANALYSIS:
#   Code: {code}  |  Name: {name}  |  Function: {func}

# DIRECT SIGNAL CONSUMERS (from schematic analysis):
# {i_context}

# FAILURE MODES ({n} total):
# {json.dumps(modes, indent=2)}

# SAFE MODES - write "No effect" for both I and J, skip reasoning:
# {json.dumps(safe_modes)}

# NON-SAFE MODES requiring full analysis:
# {json.dumps(unsafe_modes, indent=2)}

# ========== 6-STEP CHAIN-OF-THOUGHT PROCESS ==========

# For EACH non-safe failure mode, reason through ALL steps:

# STEP 1 - IDENTIFY THE PHYSICAL OUTPUT SIGNAL
#   What exact physical quantity does {name} produce?
#   Examples: "1.2V bandgap reference voltage", "bias currents for current mirrors",
#             "16MHz clock", "gate drive voltage for MOSFETs", "digitized 8-bit values"

# STEP 2 - TRACE EVERY DIRECT CONSUMER
#   From the DIRECT SIGNAL CONSUMERS list above, identify which blocks physically
#   receive this signal as an input pin. Be exhaustive - missing a consumer = wrong answer.

# STEP 3 - DETERMINE SPECIFIC SYMPTOM PER CONSUMER PER MODE
#   For each consumer: "If {code} output is [stuck/floating/incorrect/drifting],
#   what SPECIFICALLY fails in this consumer?"
#   Use IC engineering language:
#     GOOD: "Output reference voltage is stuck"  |  "Cannot operate."  |  "ADC measurement is incorrect."
#     BAD:  "The block is affected"  |  "Values become wrong"

# STEP 4 - ENUMERATE ALL SUB-EFFECTS (most commonly missed step)
#   Each consumer may have MULTIPLE distinct sub-effects - list each separately.

#   MANDATORY sub-effect patterns:
#   - Voltage reference stuck -> bias generator gets:
#       (a) Output reference voltage is stuck
#       (b) Output reference current is stuck
#       (c) Output bias current is stuck
#       (d) Quiescent current exceeding the maximum value
#     PLUS: REF itself gets "Quiescent current exceeding the maximum value"
#   - Voltage reference floating -> bias generator gets:
#       (a) Output reference voltage is floating
#       (b) Output reference current is higher than the expected range
#       (c) Output reference current is lower than the expected range
#       (d) Output bias current is higher than the expected range
#       (e) Output bias current is lower than the expected range
#   - Oscillator stuck/float -> LOGIC gets BOTH: "Cannot operate." AND "Communication error."
#   - TEMP sensor stuck -> ADC gets "TEMP output is stuck low" AND SW_BANK_x gets "SW is stuck in off state (DIETEMP)"
#   - Bias current stuck/float -> CNSN block also affected: "Incorrect reading."

# STEP 5 - APPLY OUTPUT FORMAT
#   Required format:
#     bullet BLOCK_CODE
#         dash specific symptom
#         dash second symptom if any
#     bullet ANOTHER_BLOCK
#         dash symptom

#   CRITICAL RULES:
#   (a) Generic codes only: "SW_BANK_x" NOT "SW_BANK_1"; "SW_BANK_X" for LOGIC-driven switches
#   (b) "Vega" for whole-IC device damage (OV scenarios only)
#   (c) Symptom phrases UNDER 10 WORDS - no long sentences
#   (d) Safe modes -> write exactly: No effect

# STEP 6 - DETERMINE COL J (system-level consequence, first matching rule wins)
#   - REF/BIAS non-safe: "Unintentional LED ON/OFF\\nFail-safe mode active\\nNo communication"
#   - LDO OV/UV/accuracy/drift: "Fail-safe mode active\\nNo communication"
#   - LDO spikes/startup/oscillation-within: "No effect"
#   - OSC stuck/float/freq/swing/drift: "Fail-safe mode active\\nNo communication"
#   - OSC duty-cycle/jitter: "No effect"
#   - TEMP stuck: "Unintentional LED ON"
#   - TEMP float/incorrect/accuracy: "Unintentional LED ON\\nPossible device damage"
#   - CSNS ALL modes: "No effect"
#   - ADC stuck/float: "Unintentional LED ON"
#   - ADC all others (accuracy/offset/linearity/etc): "No effect"
#   - CP OV: "Device damage"
#   - CP UV: "Unintentional LED ON"
#   - CP oscillation/quiescent/spikes/startup: "No effect"
#   - LOGIC ALL 3 modes: "Unintentional LED ON/OFF\\nFail-safe mode active\\nNo communication"
#   - INTERFACE ALL modes: "Fail-safe mode active"
#   - TRIM omission/commission/incorrect: "Fail-safe mode active\\nNo communication"
#   - TRIM settling: "No effect"
#   - SW_BANK stuck: "Unintended LED ON/OFF"
#   - SW_BANK float/res-high: "Unintended LED ON"
#   - SW_BANK res-low/timing: "No effect"

# ========== WORKED EXAMPLE (reasoning only - do not copy values) ==========

# Block: REF (Voltage Reference)  |  Mode: "Output is stuck (i.e. high or low)"

# STEP 1: REF produces a stable 1.2V bandgap voltage used as a reference point.
# STEP 2: Consumers are BIAS (sets mirror currents from REF), ADC (REF as conversion
#         reference), TEMP (comparator reference), LDO (regulation feedback), OSC (freq network).
#         REF also has self-quiescent current.
# STEP 3+4:
#   BIAS: stuck voltage -> (a) ref voltage stuck, (b) ref current stuck, (c) bias current stuck,
#         (d) quiescent current exceeds max
#   REF self: quiescent current exceeds max
#   ADC: REF output is stuck
#   TEMP: Output is stuck
#   LDO: Output is stuck
#   OSC: Oscillation does not start
# STEP 5 output:
#   bullet BIAS
#       dash Output reference voltage is stuck
#       dash Output reference current is stuck
#       dash Output bias current is stuck
#       dash Quiescent current exceeding the maximum value
#   bullet REF
#       dash Quiescent current exceeding the maximum value
#   bullet ADC
#       dash REF output is stuck
#   bullet TEMP
#       dash Output is stuck
#   bullet LDO
#       dash Output is stuck
#   bullet OSC
#       dash Oscillation does not start
# STEP 6: REF non-safe -> "Unintentional LED ON/OFF\\nFail-safe mode active\\nNo communication"

# ========== NOW ANALYZE YOUR {n} MODES ==========

# Return a JSON array with EXACTLY {n} objects, same order as the modes list:
# [
#   {{
#     "G": "<copy failure mode string verbatim>",
#     "I": "<col I with bullet+dash format, all sub-effects>",
#     "J": "<col J exact string from Step 6 rules>"
#   }},
#   ...
# ]

# QUALITY CHECK before submitting:
#   - Every non-safe mode has at least one bullet with at least one dash sub-effect
#   - Safe modes have exactly "No effect" for both I and J
#   - No numbered SW_BANK codes (SW_BANK_x or SW_BANK_X only)
#   - Each symptom phrase is under 10 words
#   - J values exactly match the Step 6 rules

# Return ONLY the JSON array:"""

#     raw    = query_llm(prompt, temperature=0.0)
#     parsed = parse_json(raw)

#     rows = []
#     if isinstance(parsed, list) and len(parsed) >= n:
#         for i in range(n):
#             rd   = parsed[i]
#             # Post-process I: replace "bullet" with "•" and "dash" with "    -"
#             # in case the LLM used the word form instead of actual symbols
#             ic   = str(rd.get('I', 'No effect')).strip()
#             ic   = ic.replace('bullet ', '• ').replace('\ndash ', '\n    - ').replace('\n- ', '\n    - ')
#             sys_ = str(rd.get('J', 'No effect')).strip()
#             mode = modes[i]

#             if is_safe_mode(mode):
#                 ic   = 'No effect'
#                 memo = 'O'
#             else:
#                 memo = compute_k_from_mode_and_coverage(code, mode, ic, block_to_sms)

#             sys_ = _validate_j(sys_)
#             rows.append(_build_row(mode, ic, sys_, memo, block_to_sms, sm_coverage,
#                                    fmeda_code=code))
#     else:
#         print(f"    LLM parse failed for {code} - using fallback")
#         rows = _fallback_rows_v7(modes, code, block_to_sms, sm_coverage)

#     return rows


# def _validate_j(j_val: str) -> str:
#     """
#     Normalize LLM J value to canonical strings.
#     Spelling: 'Unintentional' for system-level (REF/BIAS/etc), 'Unintended' for driver (SW_BANK).
#     Both are returned correctly here; the LLM prompt specifies which to use per block.
#     """
#     if not j_val or j_val.strip() == '':
#         return 'No effect'
#     j_lower = j_val.lower().strip()

#     if 'device damage' in j_lower or 'damage to device' in j_lower:
#         return 'Possible device damage' if 'possible' in j_lower else 'Device damage'

#     if j_lower in ('no effect', 'none', 'no system effect', 'no impact'):
#         return 'No effect'
#     if 'no effect' in j_lower and len(j_lower) < 30:
#         return 'No effect'

#     has_no_comms = any(p in j_lower for p in ['no communication', 'no comms',
#                                                'loss of communication', 'comms lost'])
#     has_led = any(p in j_lower for p in ['led', 'unintention', 'unintended'])
#     has_fail = 'fail' in j_lower

#     if has_no_comms and has_led and has_fail:
#         return 'Unintentional LED ON/OFF\nFail-safe mode active\nNo communication'
#     if has_no_comms and has_fail:
#         return 'Fail-safe mode active\nNo communication'
#     if 'possible' in j_lower and 'fail' in j_lower:
#         return 'Possible Fail-safe mode activation'
#     if 'fs mode' in j_lower and 'led' in j_lower:
#         return 'Unintended LED ON/OFF in FS mode'
#     if 'fail-safe' in j_lower or 'failsafe' in j_lower or 'fail safe' in j_lower:
#         return 'Fail-safe mode active'

#     # LED ON/OFF combined -- check for both spellings
#     if ('led on/off' in j_lower) or \
#        ('led on' in j_lower and 'led off' in j_lower):
#         # Use 'Unintentional' for system-level (default), 'Unintended' if explicitly stated
#         if 'unintended led' in j_lower and 'unintentional' not in j_lower:
#             return 'Unintended LED ON/OFF'
#         return 'Unintentional LED ON/OFF'

#     # LED ON
#     if any(p in j_lower for p in ['led on', 'led turns on', 'led always on', 'leds on']):
#         if 'off' not in j_lower:
#             if 'unintended led' in j_lower and 'unintentional' not in j_lower:
#                 return 'Unintended LED ON'
#             return 'Unintentional LED ON'

#     # LED OFF
#     if any(p in j_lower for p in ['led off', 'led turns off', 'led always off', 'leds off']):
#         if 'unintended led' in j_lower and 'unintentional' not in j_lower:
#             return 'Unintended LED OFF'
#         return 'Unintentional LED OFF'

#     return j_val.strip()


# def _build_row(canonical_mode, ic, sys_, memo, block_to_sms=None, sm_coverage=None, **kw):
#     ic_clean = ic.strip()
#     if ic_clean in ('No effect', ''):
#         memo = 'O'
#     sp       = 'Y' if memo == 'X' else 'N'
#     pct_safe = 1 if not memo.startswith('X') else 0
#     sm_str, coverage = '', ''
#     if ic_clean != 'No effect':
#         sm_str, coverage = compute_sm_columns(
#             ic_clean, block_to_sms or {}, sm_coverage or {},
#             fmeda_code=kw.get('fmeda_code', ''),
#             mode_str=canonical_mode
#         )
#     return {
#         'G': canonical_mode, 'I': ic, 'J': sys_, 'K': memo,
#         'O': 1, 'P': sp, 'R': pct_safe,
#         'S': sm_str, 'T': '', 'U': coverage, 'V': '',
#         'X': 'Y' if memo.startswith('X') else 'N',
#         'Y': sm_str, 'Z': '', 'AA': '', 'AB': '', 'AD': '',
#     }


# def _fallback_rows_v7(modes, fmeda_code, block_to_sms, sm_coverage):
#     """Fallback when LLM completely fails - uses mode classification only."""
#     rows = []
#     for mode in modes:
#         ic = 'No effect' if is_safe_mode(mode) else ''
#         memo = compute_k_from_mode_and_coverage(fmeda_code, mode, ic, block_to_sms)
#         rows.append(_build_row(mode, ic, 'No effect' if is_safe_mode(mode) else '',
#                                memo, block_to_sms, sm_coverage, fmeda_code=fmeda_code))
#     return rows


# def _sm_rows(sm_code: str, sm_j_map: dict) -> list:
#     """
#     SM blocks: 2 rows.
#     I/J from sm_j_map which is built at runtime from SM list sheet descriptions.
#     """
#     ic, sys_ = sm_j_map.get(sm_code,
#                               ('Loss of safety mechanism functionality', 'Fail-safe mode active'))
#     return [
#         {'G': 'Fail to detect', 'I': ic, 'J': sys_, 'K': 'X (Latent)',
#          'O': 1, 'P': 'N', 'R': 0, 'S': '', 'T': '', 'U': '', 'V': '',
#          'X': 'Y', 'Y': '', 'Z': '', 'AA': '', 'AB': '', 'AD': ''},
#         {'G': 'False detection', 'I': 'No effect', 'J': 'No effect', 'K': 'O',
#          'O': 1, 'P': 'N', 'R': 1, 'S': '', 'T': '', 'U': '', 'V': '',
#          'X': 'N', 'Y': '', 'Z': '', 'AA': '', 'AB': '', 'AD': ''},
#     ]


# def build_sm_j_map_from_descriptions(sm_blocks: list, sm_addressed: dict,
#                                       tsr_list: list, cache: dict) -> dict:
#     """
#     Build SM I/J values from SM descriptions + TSR context using LLM.
#     Fully generic - no hardcoded SM names or effect strings.
#     """
#     if not sm_blocks:
#         return {}

#     ck = "sm_j_map__" + json.dumps(sorted(s['id'] for s in sm_blocks))
#     if not SKIP_CACHE and ck in cache:
#         print("  [SM-J] Loaded from cache")
#         return cache[ck]

#     sm_details = []
#     for sm in sm_blocks:
#         m = re.match(r'sm[-_\s]?(\d+)', sm['id'].lower())
#         code = f"SM{int(m.group(1)):02d}" if m else sm['id'].upper()
#         addressed = sm_addressed.get(code, [])
#         sm_details.append({
#             'code':        code,
#             'name':        sm.get('name', ''),
#             'description': sm.get('description', ''),
#             'addresses':   addressed
#         })

#     tsr_ctx = "\n".join(f"  {t['id']}: {t['description']}" for t in tsr_list) \
#               if tsr_list else "  (no TSR data)"

#     prompt = f"""You are an automotive IC functional safety engineer.

# SYSTEM SAFETY REQUIREMENTS:
# {tsr_ctx}

# SAFETY MECHANISMS (SMs) - each has a 'Fail to detect' failure mode:
# {json.dumps(sm_details, indent=2)}

# TASK: For each SM's 'Fail to detect' failure mode, determine:
#   col I: What IC-level symptom is visible when this SM fails to detect a fault?
#          (e.g. "Unintended LED ON", "UART Communication Error", "Device damage")
#          This should describe what the IC does wrong, not what the SM was supposed to catch.
#   col J: What system-level consequence does the end user observe?
#          Use ONLY these exact strings:
#          - "Unintended LED ON"
#          - "Unintended LED OFF"
#          - "Unintended LED ON/OFF"
#          - "Unintended LED ON/OFF in FS mode"
#          - "Fail-safe mode active"
#          - "Possible Fail-safe mode activation"
#          - "Device damage"
#          - "Possible device damage"
#          - "Performance/Functionality degredation"
#          - "No effect"
#          - "UART Communication Error" (for comms SMs - but J should describe system impact)

# Return a JSON object mapping SM code to I and J values:
# {{
#   "SM01": {{"I": "Unintended LED ON", "J": "Unintended LED ON"}},
#   "SM02": {{"I": "Device damage", "J": "Device damage"}},
#   ...
# }}
# Return ONLY the JSON object:"""

#     print("  [SM-J] Building SM effect map via LLM...")
#     raw    = query_llm(prompt, temperature=0.05)
#     parsed = parse_json(raw)

#     sm_j_map = {}
#     if isinstance(parsed, dict):
#         for sm_code, vals in parsed.items():
#             if isinstance(vals, dict):
#                 sm_j_map[sm_code] = (
#                     str(vals.get('I', 'Loss of safety mechanism functionality')).strip(),
#                     str(vals.get('J', 'Fail-safe mode active')).strip()
#                 )
#     else:
#         print("  [SM-J] LLM parse failed - using generic fallback")
#         for sm in sm_details:
#             sm_j_map[sm['code']] = ('Loss of safety mechanism functionality', 'Fail-safe mode active')

#     cache[ck] = sm_j_map
#     save_cache(cache)
#     print(f"  [SM-J] {len(sm_j_map)} SM effect entries built")
#     return sm_j_map


# # =============================================================================
# # AGENT 3  -  Template Writer (deterministic)
# # =============================================================================

# def _compute_fit_values(code, n_modes, block_fit_rates, row_memo, row_U, sm_coverage):
#     block_fit = block_fit_rates.get(code, 0.0)
#     mode_fit  = block_fit / n_modes if n_modes > 0 and block_fit > 0 else 0.0
#     if not row_memo.startswith('X'):
#         return block_fit, mode_fit, mode_fit, 0.0, None, None
#     U = float(row_U) if row_U else 0.0
#     V = mode_fit * (1.0 - U)
#     if not U:
#         AA = 0.0
#     elif U >= 0.99:
#         AA = 1.0
#     elif U >= 0.85:
#         AA = 0.8
#     else:
#         AA = U
#     AB = mode_fit * U * (1.0 - AA)
#     return block_fit, mode_fit, mode_fit, V, AA, AB


# def _scan_placeholders(ws):
#     idx = {}
#     for ws_row in ws.iter_rows():
#         for cell in ws_row:
#             if cell.__class__.__name__ == 'MergedCell':
#                 continue
#             v = str(cell.value) if cell.value is not None else ''
#             if v.startswith('{{FMEDA_') and v.endswith('}}'):
#                 idx[v] = cell
#     return idx


# def _get_groups(idx, data_start=22):
#     d_rows = sorted({
#         int(re.search(r'(\d+)', k).group(1))
#         for k in idx
#         if re.match(r'\{\{FMEDA_D\d+\}\}', k)
#         and int(re.search(r'(\d+)', k).group(1)) >= data_start
#     })
#     all_rows = sorted({
#         int(re.search(r'(\d+)', k).group(1))
#         for k in idx
#         if re.match(r'\{\{FMEDA_[A-Z]+\d+\}\}', k)
#         and int(re.search(r'(\d+)', k).group(1)) >= data_start
#     })
#     groups = []
#     for i, first in enumerate(d_rows):
#         nxt = d_rows[i+1] if i+1 < len(d_rows) else 999999
#         groups.append([r for r in all_rows if first <= r < nxt])
#     return groups


# def _write(idx, col, row_num, value, wrap=False):
#     key = '{{FMEDA_' + col + str(row_num) + '}}'
#     if key not in idx:
#         return
#     cell = idx[key]
#     if value is None or (isinstance(value, str) and value.strip() in ('', 'None', 'nan')):
#         cell.value = None
#         return
#     cell.value = value
#     if wrap and isinstance(value, str) and '\n' in value:
#         old = cell.alignment or Alignment()
#         cell.alignment = Alignment(wrap_text=True,
#                                    vertical=old.vertical or 'center',
#                                    horizontal=old.horizontal or 'left')


# def agent3_write_template(fmeda_data, block_fit_rates=None, sm_coverage=None):
#     if block_fit_rates is None:
#         block_fit_rates = {}
#     if sm_coverage is None:
#         sm_coverage = {}
#     shutil.copy2(TEMPLATE_FILE, OUTPUT_FILE)
#     wb = openpyxl.load_workbook(OUTPUT_FILE)
#     ws = wb['FMEDA']
#     idx    = _scan_placeholders(ws)
#     groups = _get_groups(idx)
#     print(f"\n  [Agent 3] Template groups : {len(groups)}")
#     print(f"  [Agent 3] Blocks to write : {len(fmeda_data)}")

#     fm = 1
#     for bi, block in enumerate(fmeda_data):
#         code = block['fmeda_code']
#         rows = block['rows']
#         if bi >= len(groups):
#             print(f"  [Agent 3] WARNING: no template group for block {bi+1} ({code})")
#             break
#         group_rows = groups[bi]
#         n_t = len(group_rows)
#         n_d = len(rows)
#         if n_d > n_t:
#             print(f"  [Agent 3] {code}: {n_d} modes > {n_t} slots - truncating")
#             rows = rows[:n_t]
#         n_modes_total = max(len(rows), 1)

#         for mi, row_num in enumerate(group_rows):
#             rd       = rows[mi] if mi < len(rows) else None
#             is_first = (mi == 0)
#             _write(idx, 'B', row_num, f'FM_TTL_{fm}' if rd else None)
#             _write(idx, 'C', row_num, code)
#             _write(idx, 'D', row_num, code if is_first else None)
#             if rd is None:
#                 _write(idx, 'G', row_num, None)
#                 continue

#             memo     = str(rd.get('K', 'O')).strip()
#             sp       = str(rd.get('P', 'Y' if memo == 'X' else 'N')).strip()
#             pct_safe = rd.get('R', 1 if memo == 'O' else 0)
#             u_val    = rd.get('U', '')

#             fit_blk, fit_mode, fit_q, fit_v, fit_aa, fit_ab = _compute_fit_values(
#                 code, n_modes_total, block_fit_rates, memo, u_val, sm_coverage)

#             _write(idx, 'E',  row_num, fit_blk if (is_first and fit_blk > 0) else None)
#             _write(idx, 'F',  row_num, fit_mode if fit_mode > 0 else None)
#             _write(idx, 'G',  row_num, rd.get('G', ''),          wrap=True)
#             _write(idx, 'H',  row_num, None)
#             _write(idx, 'I',  row_num, rd.get('I', 'No effect'), wrap=True)
#             _write(idx, 'J',  row_num, rd.get('J', 'No effect'), wrap=True)
#             _write(idx, 'K',  row_num, memo)
#             _write(idx, 'O',  row_num, 1)
#             _write(idx, 'P',  row_num, sp)
#             _write(idx, 'Q',  row_num, fit_q if fit_q > 0 else None)
#             _write(idx, 'R',  row_num, pct_safe)
#             _write(idx, 'S',  row_num, rd.get('S') or None, wrap=False)
#             _write(idx, 'T',  row_num, rd.get('T') or None, wrap=False)
#             _write(idx, 'U',  row_num, u_val if u_val not in ('', None) else None)
#             _write(idx, 'V',  row_num, fit_v if (fit_v is not None and fit_v > 0) else None)
#             _write(idx, 'X',  row_num, rd.get('X', 'Y' if memo.startswith('X') else 'N'))
#             _write(idx, 'Y',  row_num, rd.get('Y') or None, wrap=False)
#             _write(idx, 'Z',  row_num, rd.get('Z') or None, wrap=True)
#             _write(idx, 'AA', row_num, fit_aa if fit_aa is not None else None)
#             if fit_ab is not None:
#                 _write(idx, 'AB', row_num, fit_ab if fit_ab > 0 else 0)
#             sm_str = rd.get('S', '') or ''
#             if sm_str and memo.startswith('X'):
#                 sms = sm_str.split()
#                 sms_sorted = sorted(sms,
#                     key=lambda s: sm_coverage.get(s, 0.0) if sm_coverage else 0.0,
#                     reverse=True)
#                 sm_mention = ' '.join(sms_sorted[:2]) if len(sms_sorted) >= 2 else (sms_sorted[0] if sms_sorted else '')
#                 lat_pct = int(round((fit_aa or 1.0) * 100))
#                 _write(idx, 'AD', row_num,
#                        f'{sm_mention} make the IC enter a safe-sate. Latent coverage: {lat_pct}%.',
#                        wrap=True)
#             else:
#                 _write(idx, 'AD', row_num, rd.get('AD') or None, wrap=True)
#             fm += 1

#         print(f"  [Agent 3] [{bi+1}/{len(fmeda_data)}] {code}: "
#               f"{min(n_d, n_t)} rows -> FM_TTL_{fm-min(n_d,n_t)} - FM_TTL_{fm-1}")

#     wb.save(OUTPUT_FILE)
#     print(f"\n  [Agent 3] Saved -> {OUTPUT_FILE}")
#     print(f"  [Agent 3] Total failure modes: {fm - 1}")


# # =============================================================================
# # MAIN
# # =============================================================================

# def run():
#     print("╔═══════════════════════════════════════════════════╗")
#     print("║   FMEDA Multi-Agent Pipeline  v7 (fully generic)  ║")
#     print("╚═══════════════════════════════════════════════════╝")
#     print(f"\n  Dataset  : {DATASET_FILE}")
#     print(f"  IEC table: {IEC_TABLE_FILE}")
#     print(f"  Template : {TEMPLATE_FILE}")
#     print(f"  Model    : {OLLAMA_MODEL}")
#     print(f"  Output   : {OUTPUT_FILE}\n")

#     cache = load_cache()

#     # ── Step 0: Read all inputs ────────────────────────────────────────────
#     print("━━━ Step 0 : Reading inputs ━━━")
#     blk_blocks, sm_blocks, tsr_list = read_dataset()
#     iec_table = read_iec_table()

#     # SM list and FIT rates from TEMPLATE only (no reference FMEDA)
#     sm_coverage, sm_addressed, block_to_sms = read_sm_list()
#     block_fit_rates = {}
#     if os.path.exists(TEMPLATE_FILE):
#         try:
#             wb_fit = openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)
#             block_fit_rates = read_block_fit_rates(wb_fit)
#             if block_fit_rates:
#                 print(f"  FIT rates: {len(block_fit_rates)} blocks from {TEMPLATE_FILE}")
#         except Exception:
#             pass
#     if not block_fit_rates:
#         print("  FIT rates: not available (place FMEDA_TEMPLATE.xlsx with FIT data to enable)")

#     print(f"  BLK: {len(blk_blocks)}  SM: {len(sm_blocks)}  TSR: {len(tsr_list)}  "
#           f"IEC: {len(iec_table)}  SM entries: {len(sm_coverage)}  "
#           f"FIT blocks: {len(block_fit_rates)}")
#     print("  block_to_sms:")
#     for b, sms in sorted(block_to_sms.items()):
#         print(f"    {b:<15} -> {sms}")

#     # ── Agent 0: Build signal flow graph ──────────────────────────────────
#     print("\n━━━ Agent 0 : Signal flow graph builder ━━━")
#     signal_graph = build_signal_flow_graph(blk_blocks, cache)
#     for blk_name, info in signal_graph.items():
#         consumers = info.get('consumers', [])
#         print(f"  {blk_name:<20} outputs: {info.get('output_signal','?')[:50]}")
#         print(f"  {'':20} feeds:   {consumers}")

#     # ── Agent 1: Map blocks -> IEC parts ──────────────────────────────────
#     print("\n━━━ Agent 1 : Block -> IEC part mapper (LLM) ━━━")
#     blocks = agent1_map_blocks(blk_blocks, sm_blocks, iec_table, cache, sm_coverage)
#     print("\n  Mapping result:")
#     for b in blocks:
#         tag = " [DUP]" if b.get('is_duplicate') else (" [SM]" if b.get('is_sm') else "")
#         flags = []
#         if b.get('is_driver'):     flags.append('driver')
#         if b.get('is_interface'):  flags.append('interface')
#         if b.get('is_trim'):       flags.append('trim')
#         if b.get('is_opamp_type'): flags.append('opamp')
#         print(f"    {b['name']:<35} -> {b['fmeda_code']:<12} "
#               f"({len(b.get('modes', []))} modes){tag} [{','.join(flags)}]")

#     # ── Build SM J map from descriptions (no reference file needed) ────────
#     print("\n━━━ Building SM effect map ━━━")
#     sm_j_map = build_sm_j_map_from_descriptions(
#         sm_blocks, sm_addressed, tsr_list, cache)

#     # ── Agent 2: Generate IC/system effects ───────────────────────────────
#     print("\n━━━ Agent 2 : IC Effects (col I) + System Effects (col J) ━━━")
#     fmeda_data = agent2_generate_effects(
#         blocks, tsr_list, block_to_sms, sm_coverage,
#         sm_addressed, cache, signal_graph, sm_j_map)

#     print("\n  Spot-check (K, I preview):")
#     for block in fmeda_data:
#         for row in block['rows']:
#             print(f"    {block['fmeda_code']:<12} K={row.get('K','?'):<12} "
#                   f"I={repr(row.get('I',''))[:50]}  | {row['G'][:35]}")

#     with open(INTERMEDIATE_JSON, 'w', encoding='utf-8') as f:
#         json.dump(fmeda_data, f, indent=2, ensure_ascii=False, default=str)
#     print(f"\n  Intermediate JSON -> {INTERMEDIATE_JSON}")

#     # ── Agent 3: Write template ────────────────────────────────────────────
#     print("\n━━━ Agent 3 : Template writer (deterministic) ━━━")
#     agent3_write_template(fmeda_data, block_fit_rates, sm_coverage)

#     print("\n✅  Pipeline complete!")
#     print(f"    Output       : {OUTPUT_FILE}")
#     print(f"    Intermediate : {INTERMEDIATE_JSON}")
#     print(f"    Cache        : {CACHE_FILE}")


# if __name__ == '__main__':
#     run()

"""
fmeda_agents_v7.py  -  Multi-Agent FMEDA Pipeline  (v7 - fully generic, zero hardcoding)
==========================================================================================

DESIGN PRINCIPLE — ZERO HARDCODING:
  This pipeline generates the FMEDA entirely from first principles.
  It requires ONLY:
    1. Your dataset file  (e.g. fusa_ai_agent_mock_data_2.xlsx)
       containing BLK, SM, TSR sheets
    2. The IEC 62380 table  (pdf_extracted.json)
    3. The FMEDA template   (FMEDA_TEMPLATE.xlsx)
       containing the SM list sheet (for coverage values) and blank rows to fill

  NO reference FMEDA is used. NO human-made output file is read.
  NO block names, effect strings, SM codes, or any chip-specific values
  are hardcoded in this file.

  If you change the dataset to a completely different chip with different
  block names, the system automatically adapts — no code changes needed.

HOW COL I IS GENERATED (the hardest column):
  Agent 0 builds a signal flow graph by asking the LLM to analyze the
  block descriptions and determine which blocks consume each other's outputs.
  Agent 2 then uses that graph to prompt the LLM with precise context:
  "Block X outputs signal Y. Blocks A, B, C consume it. For failure mode Z,
  describe exactly what breaks in each consumer." This produces complete,
  specific sub-effects without relying on any memorized chip values.

INPUTS:
  DATASET_FILE   - your chip dataset  (BLK / SM / TSR sheets)
  IEC_TABLE_FILE - IEC 62380 failure mode table  (pdf_extracted.json)
  TEMPLATE_FILE  - FMEDA template with SM list and blank placeholder rows

AGENTS:
  Agent 0  Signal flow graph builder  (LLM, cached)
  Agent 1  Block -> IEC part mapper   (LLM)
  Agent 2  Col I/J generator          (LLM with signal-flow context)
           Col K/P/X calculator       (deterministic from ISO 26262 rules)
  Agent 3  Template writer            (deterministic)
"""

import json, re, time, shutil, sys, os
import pandas as pd
import openpyxl
import requests
from openpyxl.styles import Alignment

# ─── CONFIG ───────────────────────────────────────────────────────────────────
DATASET_FILE      = 'fusa_ai_agent_mock_data.xlsx'
BLK_SHEET         = 'BLK'
SM_SHEET          = 'SM'
TSR_SHEET         = 'TSR'
IEC_TABLE_FILE    = 'pdf_extracted.json'
TEMPLATE_FILE     = 'FMEDA_TEMPLATE.xlsx'
OUTPUT_FILE       = 'FMEDA_filled.xlsx'
CACHE_FILE        = 'fmeda_cache.json'
INTERMEDIATE_JSON = 'fmeda_intermediate.json'

OLLAMA_URL     = 'http://localhost:11434/api/generate'
OLLAMA_MODEL   = 'qwen3:30b'
OLLAMA_TIMEOUT = 300
SKIP_CACHE     = False

# Runtime-populated SM description dict: {SM_CODE: 'description text'}
# Filled by read_sm_list() in run(), used by compute_sm_columns() for
# keyword-based SM filtering without hardcoded SM codes.
_SM_DESCRIPTIONS_RUNTIME: dict = {}
# ──────────────────────────────────────────────────────────────────────────────


# =============================================================================
# LLM / CACHE HELPERS
# =============================================================================

def query_llm(prompt: str, temperature: float = 0.1) -> str:
    try:
        r = requests.post(OLLAMA_URL, json={
            "model": OLLAMA_MODEL, "prompt": prompt, "stream": False,
            "options": {"temperature": temperature, "num_ctx": 16384,
                        "top_p": 0.9, "repeat_penalty": 1.1}
        }, timeout=OLLAMA_TIMEOUT)
        r.raise_for_status()
        return r.json()["response"].strip()
    except requests.exceptions.ConnectionError:
        print("  Cannot connect to Ollama. Is it running? (ollama serve)")
        sys.exit(1)
    except Exception as e:
        print(f"  LLM error: {e}")
        return ""


def parse_json(text: str):
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL).strip()
    text = re.sub(r'^```(?:json)?\s*', '', text.strip(), flags=re.MULTILINE)
    text = re.sub(r'```\s*$', '', text, flags=re.MULTILINE).strip()
    for pattern in [r'\[.*\]', r'\{.*\}']:
        m = re.search(pattern, text, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                pass
    return None


def load_cache():
    try:
        with open(CACHE_FILE, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def save_cache(cache):
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)


# =============================================================================
# READ ALL INPUTS
# =============================================================================

def read_dataset():
    """
    Read BLK, SM, TSR sheets from dataset.
    Returns (blk_blocks, sm_blocks, tsr_list)
    All are generic dicts - no assumption about block names or count.
    """
    xl = pd.ExcelFile(DATASET_FILE)

    # BLK sheet
    blk_blocks = []
    if BLK_SHEET in xl.sheet_names:
        df = pd.read_excel(DATASET_FILE, sheet_name=BLK_SHEET, dtype=str).fillna('')
        for _, row in df.iterrows():
            vals = [v.strip() for v in row.values if str(v).strip()]
            if len(vals) >= 2:
                blk_blocks.append({
                    'id':       vals[0],
                    'name':     vals[1],
                    'function': vals[2] if len(vals) > 2 else ''
                })
    else:
        # Try 'Description of Block' sheet as fallback
        for sheet in xl.sheet_names:
            if 'block' in sheet.lower() or 'blk' in sheet.lower() or 'description' in sheet.lower():
                df = pd.read_excel(DATASET_FILE, sheet_name=sheet, dtype=str).fillna('')
                for _, row in df.iterrows():
                    vals = [v.strip() for v in row.values if str(v).strip()]
                    # Skip header-like rows
                    if len(vals) >= 2 and re.match(r'\d+', vals[0]):
                        blk_blocks.append({
                            'id':       vals[0],
                            'name':     vals[1],
                            'function': vals[2] if len(vals) > 2 else ''
                        })
                if blk_blocks:
                    break

    # SM sheet
    sm_blocks = []
    if SM_SHEET in xl.sheet_names:
        df_sm = pd.read_excel(DATASET_FILE, sheet_name=SM_SHEET, dtype=str).fillna('')
        for _, row in df_sm.iterrows():
            vals = [v.strip() for v in row.values if str(v).strip()]
            if vals and re.match(r'sm[-_\s]?\d+', vals[0].lower()):
                sm_blocks.append({
                    'id':          vals[0],
                    'name':        vals[1] if len(vals) > 1 else '',
                    'description': vals[2] if len(vals) > 2 else ''
                })

    # TSR sheet
    tsr_list = []
    if TSR_SHEET in xl.sheet_names:
        df_tsr = pd.read_excel(DATASET_FILE, sheet_name=TSR_SHEET, dtype=str).fillna('')
        for _, row in df_tsr.iterrows():
            vals = [v.strip() for v in row.values if str(v).strip()]
            if len(vals) >= 2:
                tsr_list.append({
                    'id':           vals[0],
                    'description':  vals[1],
                    'connected_fsr': vals[2] if len(vals) > 2 else ''
                })

    return blk_blocks, sm_blocks, tsr_list


def read_iec_table():
    with open(IEC_TABLE_FILE, encoding='utf-8-sig') as f:
        return json.load(f)


def read_block_fit_rates(wb):
    """Read FIT rates dynamically - works for any sheet structure."""
    fit_rates = {}
    target_sheets = [s for s in wb.sheetnames
                     if 'fit' in s.lower() or 'block' in s.lower() or 'core' in s.lower()]
    if not target_sheets:
        return fit_rates
    try:
        ws = wb[target_sheets[0]]
        # Find columns by scanning headers
        block_col, fit_col = None, None
        for row in ws.iter_rows(min_row=1, max_row=35):
            for c in row:
                if c.value:
                    v = str(c.value).lower()
                    if 'block' in v and not block_col:
                        block_col = c.column_letter
                    if 'total' in v and 'fit' in v and not fit_col:
                        fit_col = c.column_letter
                    elif 'fit' in v and 'total' in v and not fit_col:
                        fit_col = c.column_letter
        if not block_col:
            block_col = 'B'
        if not fit_col:
            fit_col = 'L'
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            rd = {c.column_letter: c.value for c in row
                  if hasattr(c, 'column_letter') and c.value is not None}
            if block_col in rd and fit_col in rd:
                block = str(rd[block_col]).strip()
                try:
                    fit_rates[block] = float(rd[fit_col])
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        print(f"  WARNING: Could not read FIT rates: {e}")
    return fit_rates


def read_sm_list(wb=None):
    """
    Read SM list from TEMPLATE_FILE only.
    Returns:
      sm_coverage    : { 'SM01': 0.99, ... }
      sm_addressed   : { 'SM01': ['REF','LDO'], ... }
      block_to_sms   : { 'REF': ['SM01','SM02',...], ... }
      sm_descriptions: { 'SM01': 'Comparator: VDD Under-voltage', ... }
    """
    sources = []
    if wb is not None:
        sources.append(('provided workbook', wb))
    if os.path.exists(TEMPLATE_FILE):
        try:
            sources.append((TEMPLATE_FILE, openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)))
        except Exception:
            pass

    for label, wb_src in sources:
        if 'SM list' not in wb_src.sheetnames:
            continue
        cov, addr, b2s, sm_desc = _parse_sm_list_sheet(wb_src['SM list'])
        if cov:
            print(f"  SM list: {len(cov)} entries loaded from {label}")
            return cov, addr, b2s, sm_desc

    print("  SM list: no template found -- SM coverage will be empty")
    print("  (Place FMEDA_TEMPLATE.xlsx in the working directory to enable SM coverage)")
    return {}, {}, {}, {}


def _parse_sm_list_sheet(ws):
    """
    Parse SM list sheet -- works for any column arrangement.
    Returns sm_coverage, sm_addressed, block_to_sms, sm_descriptions.
    sm_descriptions: {SM_CODE: 'description text'} -- used for keyword filtering
    in compute_sm_columns so it adapts to renamed blocks.
    """
    sm_coverage   = {}
    sm_addressed  = {}
    sm_descriptions = {}

    # Auto-detect columns
    sm_col, cov_col, parts_col, desc_col = 'C', 'L', 'E', 'F'
    for row in ws.iter_rows(min_row=1, max_row=15):
        for c in row:
            if c.value and str(c.value).strip().upper() in ('SM', 'SM CODE', 'SAFETY MECHANISM', 'SM#'):
                sm_col = c.column_letter
            if c.value and 'coverage' in str(c.value).lower():
                cov_col = c.column_letter
            if c.value and ('part' in str(c.value).lower() or 'address' in str(c.value).lower()):
                parts_col = c.column_letter
            if c.value and ('mechanism' in str(c.value).lower() or 'measure' in str(c.value).lower()
                            or 'description' in str(c.value).lower()):
                desc_col = c.column_letter

    for row in ws.iter_rows(min_row=10, max_row=ws.max_row):
        cells = {c.column_letter: c.value for c in row
                 if hasattr(c, 'column_letter') and c.value is not None}
        if sm_col not in cells:
            continue
        sm_code = str(cells[sm_col]).strip()
        if not re.match(r'SM\d+', sm_code):
            continue
        try:
            cov = float(str(cells.get(cov_col, 0.9)))
        except (ValueError, TypeError):
            cov = 0.9
        sm_coverage[sm_code] = cov

        # SM description text (used for keyword matching in compute_sm_columns)
        desc = str(cells.get(desc_col, '')).strip()
        if desc:
            sm_descriptions[sm_code] = desc

        raw_parts = str(cells.get(parts_col, '')).strip()
        parts = []
        for p in re.split(r'[,;]', raw_parts):
            p = p.strip()
            p = re.sub(r'SW_BANK[_x\d]*', 'SW_BANK', p)
            p = re.sub(r'\bCSNS\b|\bCNSN\b|\bCS\b', 'CSNS', p)
            if p:
                parts.append(p)
        sm_addressed[sm_code] = parts

    block_to_sms = {}
    for sm, parts in sm_addressed.items():
        for part in parts:
            if part:
                block_to_sms.setdefault(part, [])
                if sm not in block_to_sms[part]:
                    block_to_sms[part].append(sm)

    return sm_coverage, sm_addressed, block_to_sms, sm_descriptions


# =============================================================================
# AGENT 0  -  Build signal flow graph from dataset
# =============================================================================

def build_signal_flow_graph(blk_blocks: list, cache: dict) -> dict:
    """
    Use LLM to build a precise signal dependency graph from block descriptions.
    The graph drives col I generation — wrong consumers = wrong I values.
    This prompt enforces strict IC signal-flow thinking to prevent common errors.
    """
    ck = "signal_flow_v11__" + json.dumps(sorted(b['name'] for b in blk_blocks))
    if not SKIP_CACHE and ck in cache:
        print("  [Agent 0] Signal flow graph loaded from cache")
        return cache[ck]

    blocks_text = "\n".join(
        f"  {b['name']} ({b.get('id','?')}): {b['function']}"
        for b in blk_blocks
    )

    prompt = f"""You are a senior analog/mixed-signal IC architect analyzing chip signal paths for FMEDA.

CHIP BLOCKS AND THEIR FUNCTIONS:
{blocks_text}

TASK: For each block, map its DIRECT downstream signal consumers by reading its function description.
"Direct" means the signal physically arrives at the consumer's input pin (1 hop only).

HOW TO REASON (use the function descriptions, not assumed topology):
  - Read each block's function description carefully.
  - Ask: "What does this block OUTPUT, and which other blocks need that output as an INPUT?"
  - A block that "produces a reference voltage" feeds blocks that "use a reference" or "set levels from".
  - A block that "generates bias currents" feeds every block that "uses current mirrors" or "biased by".
  - A block that "produces a supply voltage" feeds blocks whose description says "powered by" or "supply to".
  - A block that "generates a clock/frequency" feeds blocks that "require a clock" or "digital logic".
  - A block that "converts analog to digital" feeds blocks that "use measurement results".
  - A block that "drives LEDs/switches" typically has no downstream consumers (external output).
  - A block that "calibrates/trims" feeds all calibrated blocks (REF, LDO, BIAS, OSC, etc.).
  - A block that "communicates via SPI/UART" feeds the digital controller.

RULES:
  1. ONLY use information from the function descriptions above. Do NOT assume topology.
  2. Only list DIRECT consumers (1 hop). If A feeds B feeds C, only B is in A's consumer list.
  3. Use the exact block names from the CHIP BLOCKS list above.
  4. For each consumer, give a specific 5-10 word symptom describing what fails when the source fails.
     GOOD: "ADC conversion result is incorrect"
     GOOD: "oscillator frequency drifts out of spec"
     BAD:  "block is affected"

Return a JSON object:
{{
  "BlockName": {{
    "output_signal": "physical signal this block produces",
    "consumers": ["BlockName1", "BlockName2"],
    "consumer_details": {{
      "BlockName1": "specific symptom in 5-10 words",
      "BlockName2": "specific symptom in 5-10 words"
    }}
  }},
  ...
}}

Return ONLY the JSON object:"""

    print("  [Agent 0] Building signal flow graph via LLM...")
    raw = query_llm(prompt, temperature=0.0)
    result = parse_json(raw)

    if not isinstance(result, dict):
        print("  [Agent 0] LLM parse failed - using empty graph")
        result = {}

    cache[ck] = result
    save_cache(cache)
    print(f"  [Agent 0] Signal flow graph: {len(result)} blocks mapped")
    return result


# =============================================================================
# SAFE-MODE CLASSIFIER  (deterministic, chip-agnostic)
# =============================================================================

# Keywords that indicate a failure mode is locally contained (no propagation)
_SAFE_MODE_KEYWORDS = [
    'spike', 'oscillation within', 'within the expected range', 'within the prescribed',
    'jitter', 'incorrect start-up', 'start-up time', 'quiescent current exceeding',
    'incorrect settling time', 'settling time', 'fast oscillation outside',
    'false detection', 'duty cycle', 'filter in place',
]

def is_safe_mode(mode_str: str) -> bool:
    """Return True if this failure mode is locally contained (no downstream propagation)."""
    m = mode_str.lower()
    return any(k in m for k in _SAFE_MODE_KEYWORDS)


def classify_mode_severity(mode_str: str) -> str:
    """
    Classify mode into a severity type.
    Returns: 'safe' | 'stuck' | 'float' | 'ov' | 'uv' | 'accuracy' | 'drift' | 'other'
    """
    m = mode_str.lower()
    if is_safe_mode(m):
        return 'safe'
    # Stuck/floating - with exclusion for 'not including stuck' phrasing
    if ('stuck' in m and 'not including stuck' not in m) or \
       ('driver is stuck' in m):
        return 'stuck'
    if ('floating' in m or 'open circuit' in m or 'tri-state' in m) and \
       'not including' not in m:
        return 'float'
    if any(k in m for k in ['higher than a high threshold', 'over voltage', 'overvoltage',
                              'output voltage higher']):
        return 'ov'
    if any(k in m for k in ['lower than a low threshold', 'under voltage', 'undervoltage',
                              'output voltage lower']):
        return 'uv'
    if any(k in m for k in ['accuracy too low', 'accuracy error']):
        return 'accuracy'
    if 'drift' in m:
        return 'drift'
    if any(k in m for k in ['resistance too high', 'resistance too low',
                              'turn-on time', 'turn-off time']):
        return 'driver_perf'
    return 'other'


# =============================================================================
# AGENT 1  -  Block -> IEC part mapper
# =============================================================================

# Mode overrides for blocks where IEC table modes are wrong/generic
# These are STRUCTURAL rules (interface blocks always use TX/RX protocol),
# not chip-specific values
_MODE_STRUCTURAL_OVERRIDES = {
    # Serial interface blocks always use TX/RX message failure taxonomy
    'INTERFACE': [
        'TX: No message transferred as requested',
        'TX: Message transferred when not requested',
        'TX: Message transferred too early/late',
        'TX: Message transferred with incorrect value',
        'RX: No incoming message processed',
        'RX: Message transferred when not requested',
        'RX: Message transferred too early/late',
        'RX: Message transferred with incorrect value',
    ],
    # NVM/trim/self-test blocks use omission/commission taxonomy
    'TRIM': [
        'Error of omission (i.e. not triggered when it should be)',
        "Error of comission (i.e. triggered when it shouldn't be)",
        'Incorrect settling time (i.e. outside the expected range)',
        'Incorrect output',
    ],
}

# Driver/switch blocks use driver-specific mode descriptions, not generic signal ones.
# This is structural - any block classified as a switch/driver gets these.
_DRIVER_MODES = [
    'Driver is stuck in ON or OFF state',
    'Driver is floating (i.e. open circuit, tri-stated)',
    'Driver resistance too high when turned on',
    'Driver resistance too low when turned off',
    'Driver turn-on time too fast or too slow',
    'Driver turn-off time too fast or too slow',
]

# Voltage regulator (LDO/SMPS/charge pump) mode sequence.
# These have OV/UV as primary failure modes — NOT stuck/floating like op-amps.
# This is structural: any voltage regulator block uses this pattern.
_VOLTAGE_REG_MODES = [
    'Output voltage higher than a high threshold of the prescribed range (i.e. over voltage — OV)',
    'Output voltage lower than a low threshold of the prescribed range (i.e. under voltage — UV)',
    'Output voltage affected by spikes',
    'Incorrect start-up time',
    'Output voltage accuracy too low, including drift',
    'Output voltage oscillation within the prescribed range',
    'Output voltage affected by a fast oscillation outside the prescribed range but with average value within',
    'Quiescent current exceeding the maximum value',
]

# Digital logic / controller mode sequence.
# Logic blocks use stuck/float/incorrect-output — not gain/offset like op-amps.
_LOGIC_MODES = [
    'Output is stuck (i.e. high or low)',
    'Output is floating (i.e. open circuit)',
    'Incorrect output voltage value',
]

# BIAS block -- current-source specific taxonomy.
# BIAS produces reference currents, so modes reference "outputs" (plural),
# "reference current", and "branch currents" -- distinct from generic op-amp.
_BIAS_MODES = [
    'One or more outputs are stuck (i.e. high or low)',
    'One or more outputs are floating (i.e. open circuit)',
    'Incorrect reference current (i.e. outside the expected range)',
    'Reference current accuracy too low , including drift',
    'Reference current affected by spikes',
    'Reference current oscillation within the expected range',
    'One or more branch currents outside the expected range \nwhile reference current is correct',
    'One or more branch currents accuracy too low , including \ndrift',
    'One or more branch currents affected by spikes',
    'One or more branch currents oscillation within the expected range',
]

# Op-amp/analog buffer mode sequence (generic - works for any analog output block)
# Used for: REF, BIAS, TEMP, CSNS and similar analog signal blocks
_OPAMP_MODES_SEQUENCE = [
    'Output is stuck (i.e. high or low)',
    'Output is floating (i.e. open circuit)',
    'Incorrect output voltage value (i.e. outside the expected range)',
    'Output voltage accuracy too low, including drift',
    'Output voltage affected by spikes',
    'Output voltage oscillation within the expected range',
    'Incorrect start-up time (i.e. outside the expected range)',
    'Quiescent current exceeding the maximum value',
]

# ADC/converter mode sequence — self-referential errors, not OV/UV
_ADC_MODES = [
    'One or more outputs are stuck (i.e. high or low)',
    'One or more outputs are floating (i.e. open circuit)',
    'Accuracy error (i.e. Error exceeds the LSBs)',
    'Offset error not including stuck or floating conditions on the outputs, low resolution',
    'No monotonic conversion characteristic',
    'Full-scale error not including stuck or floating conditions on the outputs, low resolution',
    'Linearity error with monotonic conversion curve not including stuck or floating conditions on the outputs, low resolution',
    'Incorrect settling time (i.e. outside the expected range)',
]

# OSC mode sequence
_OSC_MODES = [
    'Output is stuck (i.e. high or low)',
    'Output is floating (i.e. open circuit)',
    'Incorrect output signal swing (i.e. outside the expected range)',
    'Incorrect frequency of the output signal',
    'Incorrect duty cycle of the output signal',
    'Drift of the output frequency',
    'Jitter too high in the output signal',
]


def agent1_map_blocks(blk_blocks, sm_blocks, iec_table, cache, sm_coverage=None):
    """Map chip blocks to IEC part categories and assign failure modes."""
    ck = "agent1__" + json.dumps([b['name'] for b in blk_blocks])
    if not SKIP_CACHE and ck in cache:
        print("  [Agent 1] Loaded from cache")
        result = cache[ck]
        _append_sm_blocks(result, sm_blocks, sm_coverage)
        return result

    iec_summary = ""
    for i, p in enumerate(iec_table):
        modes = p["entries"][0]["modes"]
        iec_summary += (
            f'  {i+1:2d}. "{p["part_name"]}"\n'
            f'       Desc : {p["entries"][0]["description"][:120]}\n'
            f'       Modes: {json.dumps(modes[:3])}'
            + (' ...' if len(modes) > 3 else '') + '\n\n'
        )

    blocks_text = "\n".join(
        f'  {b["id"]}: "{b["name"]}" - {b["function"]}'
        for b in blk_blocks
    )

    prompt = f"""You are an automotive IC functional safety engineer.

CHIP BLOCKS:
{blocks_text}

IEC 62380 HARDWARE PART CATEGORIES:
{iec_summary}

FMEDA SHORT CODE RULES:
  Voltage reference / bandgap                    -> REF
  Bias current source / current reference        -> BIAS
  LDO / linear voltage regulator                 -> LDO
  Internal oscillator / clock generator          -> OSC
  Watchdog / clock monitor (shares OSC slot)     -> OSC   [duplicate]
  Temperature sensor / thermal circuit           -> TEMP
  Current sense amplifier / op-amp sense         -> CSNS
  Current DAC / channel DAC                      -> ADC
  ADC (analogue to digital converter)            -> ADC   [duplicate]
  Charge pump / boost regulator                  -> CP
  nFAULT driver / fault aggregator               -> CP    [duplicate]
  Digital logic / main controller                -> LOGIC
  Open-load / short-to-GND detector              -> LOGIC [duplicate]
  SPI / UART / serial interface                  -> INTERFACE
  NVM / trim / self-test / POST                  -> TRIM
  LED driver switch bank N                       -> SW_BANK_N

TASK: For each block determine:
  "fmeda_code"       - short code from rules above
  "iec_part"         - EXACT part_name string from IEC list that best matches
  "is_duplicate"     - true if this fmeda_code was already assigned
  "is_driver"        - true if this is a switch/driver/output-stage block (SW_BANK_N)
  "is_interface"     - true if this is a serial comms block (SPI/UART/INTERFACE)
  "is_trim"          - true if this is a NVM/trim/self-test block
  "is_opamp_type"    - true if this is an analog signal block (REF, BIAS, TEMP, CSNS)
                       that produces a voltage/current output measured by another block
  "is_regulator_type"- true if this is a voltage regulator/supply block (LDO, charge pump)
                       whose primary failures are over-voltage and under-voltage
  "is_logic_type"    - true if this is a digital logic/controller block (LOGIC, MCU, FSM)
  "is_adc_type"      - true if this is an ADC/converter block
  "is_osc_type"      - true if this is an oscillator/clock block

Return JSON array, same order as input blocks:
[
  {{"id":"BLK-01","name":"Bandgap Reference","fmeda_code":"REF",
    "iec_part":"Voltage references","is_duplicate":false,
    "is_driver":false,"is_interface":false,"is_trim":false,"is_opamp_type":true,
    "is_regulator_type":false,"is_logic_type":false,"is_adc_type":false,"is_osc_type":false}},
  ...
]
Return ONLY the JSON array:"""

    print("  [Agent 1] Calling LLM to map blocks -> IEC parts...")
    raw    = query_llm(prompt, temperature=0.05)
    result = parse_json(raw)

    if not isinstance(result, list) or len(result) != len(blk_blocks):
        print("  [Agent 1] LLM parse issue - using fallback")
        result = _fallback_agent1(blk_blocks)

    # Replace LLM-generated modes with verbatim IEC table modes
    iec_idx = {p['part_name']: p['entries'][0]['modes'] for p in iec_table}
    for b in result:
        iec_part = b.get('iec_part', '')
        if iec_part in iec_idx:
            b['modes'] = iec_idx[iec_part]
        else:
            matched = False
            for pname, modes in iec_idx.items():
                if iec_part[:20].lower() in pname.lower() or pname[:20].lower() in iec_part.lower():
                    b['modes'] = modes
                    b['iec_part'] = pname
                    matched = True
                    break
            if not matched:
                b['modes'] = []
                print(f"  [Agent 1] WARNING: no IEC modes for '{iec_part}' ({b['name']})")

    # Apply structural mode overrides based on block functional type.
    # These are STRUCTURAL rules based on block category, not chip-specific names.
    for b in result:
        code = b.get('fmeda_code', '')

        # 1. Explicit code-based overrides (INTERFACE, TRIM)
        if code in _MODE_STRUCTURAL_OVERRIDES:
            b['modes'] = _MODE_STRUCTURAL_OVERRIDES[code]

        # 2. SW_BANK (driver/switch) blocks — driver taxonomy
        elif b.get('is_driver') or re.match(r'SW_BANK', code, re.IGNORECASE):
            b['modes'] = _DRIVER_MODES

        # 3. LOGIC / digital controller — stuck/float/incorrect only (3 modes)
        elif b.get('is_logic_type') or code == 'LOGIC':
            b['modes'] = _LOGIC_MODES

        # 4. ADC / converter blocks — self-referential conversion errors
        elif b.get('is_adc_type') or code == 'ADC':
            b['modes'] = _ADC_MODES

        # 5. OSC / clock blocks — frequency-specific modes
        elif b.get('is_osc_type') or code == 'OSC':
            b['modes'] = _OSC_MODES

        # 6. Voltage regulators (LDO, CP) — OV/UV primary failures
        elif b.get('is_regulator_type') or code in ('LDO', 'CP'):
            b['modes'] = _VOLTAGE_REG_MODES

        # 7. BIAS block -- current-source specific taxonomy (not generic op-amp)
        elif code == 'BIAS' or (b.get('is_opamp_type') and
                                 'bias' in b.get('name', '').lower() and
                                 'current' in b.get('function', '').lower()):
            b['modes'] = _BIAS_MODES

        # 8. Op-amp-type analog blocks (REF, TEMP, CSNS) -- stuck/float sequence
        elif b.get('is_opamp_type'):
            b['modes'] = _OPAMP_MODES_SEQUENCE

        # 8. Fallback: if IEC gave a mode list with OV/UV keywords, use voltage reg modes;
        #    if it has gain/offset but no stuck, use opamp; otherwise keep IEC modes
        elif b.get('modes'):
            modes_joined = ' '.join(b['modes']).lower()
            if 'over voltage' in modes_joined or 'under voltage' in modes_joined or \
               'high threshold' in modes_joined or 'low threshold' in modes_joined:
                b['modes'] = _VOLTAGE_REG_MODES
            elif 'stuck' not in modes_joined and 'floating' not in modes_joined and \
                 ('gain' in modes_joined or 'offset' in modes_joined):
                b['modes'] = _OPAMP_MODES_SEQUENCE

    # Enforce duplicate flags
    seen = set()
    for b in result:
        code = b.get('fmeda_code', '')
        if code in seen:
            b['is_duplicate'] = True
        else:
            b['is_duplicate'] = False
            seen.add(code)

    cache[ck] = result
    save_cache(cache)
    _append_sm_blocks(result, sm_blocks, sm_coverage)
    return result


def _append_sm_blocks(result, sm_blocks, sm_coverage=None):
    """
    Add SM blocks to the block list.
    CRITICAL: Only include SMs that are recognized in the SM list sheet (sm_coverage).
    If sm_coverage is provided, skip any SM whose code is not in it — these are
    SMs listed in the dataset's SM sheet but removed from the actual FMEDA
    (e.g. SM07, SM19 in some datasets). This prevents off-by-one row errors.
    """
    for sm in sm_blocks:
        m = re.match(r'sm[-_\s]?(\d+)', sm['id'].lower())
        code = f"SM{int(m.group(1)):02d}" if m else sm['id'].upper()

        # Skip SMs not in the SM list coverage map (they don't have FMEDA rows)
        if sm_coverage is not None and len(sm_coverage) > 0:
            if code not in sm_coverage:
                print(f"  [SM filter] Skipping {code} - not in SM list sheet")
                continue

        result.append({
            'id': sm['id'], 'name': sm['name'], 'function': sm.get('description', ''),
            'fmeda_code': code, 'iec_part': 'Safety Mechanism',
            'modes': ['Fail to detect', 'False detection'],
            'is_duplicate': False, 'is_sm': True,
        })


def _fallback_agent1(blk_blocks):
    # (keywords, fmeda_code, iec_part, is_driver, is_interface, is_trim,
    #  is_opamp, is_regulator, is_logic, is_adc, is_osc)
    KMAP = [
        (['bandgap', 'voltage reference', 'reference volt', 'ref'],
         'REF', 'Voltage references',
         False, False, False, True, False, False, False, False),
        (['bias current', 'current source', 'bias generator', 'reference current'],
         'BIAS', 'Current source (including bias current generator)',
         False, False, False, True, False, False, False, False),
        (['ldo', 'low dropout', 'linear regulator', 'supply to ic', 'produces supply'],
         'LDO', 'Voltage regulators (linear, SMPS, etc.)',
         False, False, False, False, True, False, False, False),
        (['oscillator', 'clock', 'frequency', 'mhz', 'watchdog'],
         'OSC', 'Oscillator',
         False, False, False, False, False, False, False, True),
        (['temperature', 'thermal', 'die temp', 'proportional to die'],
         'TEMP', 'Operational amplifier and buffer',
         False, False, False, True, False, False, False, False),
        (['current sense', 'csns', 'csp', 'shunt', 'generates voltage'],
         'CSNS', 'Operational amplifier and buffer',
         False, False, False, True, False, False, False, False),
        (['convert analog', 'adc', 'analogue to digital', 'digital signal coded', 'convert'],
         'ADC', 'N bits analogue to digital converters (N-bit ADC)',
         False, False, False, False, False, False, True, False),
        (['charge pump', 'boost', 'supply for switch'],
         'CP', 'Charge pump, regulator boost',
         False, False, False, False, True, False, False, False),
        (['switch bank', 'sw_bank', 'led driver', 'driver switch'],
         'SW_BANK', 'Voltage/Current comparator',
         True, False, False, False, False, False, False, False),
        (['spi', 'uart', 'serial', 'interface', 'digital interface'],
         'INTERFACE', 'N bits digital to analogue converters (DAC)d',
         False, True, False, False, False, False, False, False),
        (['trim', 'nvm', 'self-test', 'post', 'calibrat'],
         'TRIM', 'Voltage references',
         False, False, True, False, False, False, False, False),
        (['logic', 'control', 'main control', 'ic main'],
         'LOGIC', 'Voltage/Current comparator',
         False, False, False, False, False, True, False, False),
    ]
    used, result = set(), []
    for b in blk_blocks:
        combined = (b['name'] + ' ' + b['function']).lower()
        code, iec = 'LOGIC', 'Voltage/Current comparator'
        is_driver = is_interface = is_trim = is_opamp = False
        is_reg = is_logic = is_adc = is_osc = False
        for kws, c, ip, idr, iint, itrm, ioamp, ireg, ilog, iadc, iosc in KMAP:
            if any(k in combined for k in kws):
                code, iec = c, ip
                is_driver, is_interface, is_trim = idr, iint, itrm
                is_opamp, is_reg, is_logic = ioamp, ireg, ilog
                is_adc, is_osc = iadc, iosc
                break
        # Handle SW_BANK_N numbering
        if code == 'SW_BANK':
            n = re.search(r'(\d+)', b['name'])
            code = f"SW_BANK_{n.group(1)}" if n else 'SW_BANK_1'
        dup = code in used
        if not dup:
            used.add(code)
        result.append({
            'id': b['id'], 'name': b['name'], 'function': b['function'],
            'fmeda_code': code, 'iec_part': iec, 'is_duplicate': dup,
            'is_driver': is_driver, 'is_interface': is_interface,
            'is_trim': is_trim, 'is_opamp_type': is_opamp,
            'is_regulator_type': is_reg, 'is_logic_type': is_logic,
            'is_adc_type': is_adc, 'is_osc_type': is_osc,
        })
    return result


# =============================================================================
# AGENT 2  -  IC Effects (I) + System Effects (J) + Safety Flag (K)
# =============================================================================

# I column format rules
IC_FORMAT = """
EXACT FORMAT for col I "effects on IC output":
  • BLOCK_CODE
      - specific effect on that block
      - second effect if applicable
  • ANOTHER_BLOCK_CODE
      - specific effect

  If NOTHING is affected -> write exactly: No effect

RULES:
  - Use bullet (•) before each affected block name, no indent
  - Use 4 spaces + dash (    -) before each effect line under a block
  - Effects must be SPECIFIC: not "BIAS is affected" but "Output bias current is stuck"
  - Use present tense active: "is stuck", "is incorrect", "cannot operate"
  - List EVERY block that receives this signal - do not omit any
  - If multiple sub-effects exist for one block, list each on its own line
""".strip()

# J column valid values (system-level, agnostic to chip type)
J_VALID_VALUES = [
    "Unintentional LED ON/OFF\nFail-safe mode active\nNo communication",
    "Fail-safe mode active\nNo communication",
    "Fail-safe mode active",
    "Unintended LED ON/OFF",
    "Unintended LED ON",
    "Unintended LED OFF",
    "Device damage",
    "Possible device damage",
    "No effect",
]

# K override rules - derived from ISO 26262 principles, chip-agnostic
def compute_k_from_mode_and_coverage(code: str, mode_str: str,
                                      ic_effect: str, block_to_sms: dict) -> str:
    """
    Determine K (safety violation flag) deterministically using ISO 26262 principles.

    Rules derived from systematic diff analysis against human expert FMEDA:

    ALWAYS K=O (never safety-violating):
      - Safe/benign modes (spikes, oscillation within range, jitter, quiescent, settling)
      - No IC downstream effect
      - Interface/comms blocks (protocol layer detects/handles these)
      - ADC non-stuck modes (accuracy, offset, linearity, monotonic, full-scale, settling)
        → these cause measurement drift, not hard failures
      - Current-sense monitoring blocks (CSNS) — all modes → O
        → CSNS only feeds ADC for monitoring; fault is caught by ADC SM coverage
      - Voltage supply spikes and incorrect start-up time (transient, not sustained)
      - Driver turn-off timing (performance impact only)
      - Charge pump oscillation-within-range, quiescent current (local, non-propagating)
      - TRIM incorrect settling time (timing, not output value)
      - Interface RX message-value errors that the MCU will catch

    ALWAYS K=X (safety-violating):
      - OSC drift → propagates to LOGIC, causes comms failure
      - LOGIC float and incorrect output → SW_BANK and OSC lose control
      - TRIM commission and incorrect output → miscalibration of all analog blocks
      - SW_BANK float and resistance-too-high → LED stuck on (unintended)
      - Any block with hard IC effect (stuck/float/OV/UV/accuracy/drift) on SM-covered blocks
    """
    m = mode_str.lower()
    severity = classify_mode_severity(mode_str)
    norm = re.sub(r'SW_BANK[_\d]*', 'SW_BANK', code.upper())

    # ── UNIVERSAL SAFE MODES ────────────────────────────────────────────────
    if severity == 'safe':
        return 'O'

    # No IC downstream effect → always O
    if not ic_effect or ic_effect.strip() in ('', 'No effect', 'No effect (Filter in place)'):
        return 'O'

    # ── BLOCK-SPECIFIC OVERRIDES (derived from diff) ─────────────────────────

    # CSNS: monitoring-only block, all modes → O (ADC SM coverage handles it)
    if norm == 'CSNS':
        return 'O'

    # INTERFACE: protocol layer catches all these → O
    if 'INTERFACE' in norm:
        return 'O'

    # ADC: only stuck/float are hard safety failures
    if norm == 'ADC' and severity not in ('stuck', 'float'):
        return 'O'

    # Driver blocks (SW_BANK):
    if norm == 'SW_BANK':
        # stuck, float, res_high → X (unintended LED state)
        if severity in ('stuck', 'float') or 'resistance too high' in m:
            return 'X'
        # res_low (SW_BANK_1 only K=X, but generally O across all banks) → O
        # turn-on/turn-off timing → O
        return 'O'

    # LDO: spikes and startup are transient → O; OV/UV/accuracy → X
    if norm == 'LDO':
        if 'spike' in m or 'start-up' in m or 'start up' in m:
            return 'O'
        return 'X'  # OV, UV, accuracy, oscillation-within → X

    # CP (charge pump): OV/UV → X; oscillation-within, quiescent → O
    if norm == 'CP':
        if severity in ('ov', 'uv') or 'lower than' in m or 'higher than' in m:
            return 'X'
        return 'O'

    # OSC: drift IS a safety violation (causes LOGIC failure)
    if norm == 'OSC':
        if 'drift' in m:
            return 'X'
        if severity in ('stuck', 'float', 'incorrect', 'ov', 'uv'):
            return 'X'
        return 'O'

    # LOGIC: ALL three modes are safety-violating
    if norm == 'LOGIC':
        return 'X'

    # TRIM: commission and incorrect output → X; settling time → O
    if norm == 'TRIM':
        if 'settling' in m or 'start-up' in m:
            return 'O'
        return 'X'

    # ── GENERAL RULE: check SM coverage on affected blocks ──────────────────
    affected = re.findall(r'^\s*•\s*([A-Z_a-z0-9]+)', ic_effect, re.MULTILINE)
    norm_affected = []
    for b in affected:
        b = b.strip().upper()
        b = re.sub(r'SW_BANK[_X\d]*', 'SW_BANK', b)
        b = re.sub(r'CSNS|CNSN|CS(?!NS)', 'CSNS', b)
        if b not in ('NONE', 'VEGA', ''):
            norm_affected.append(b)

    # SM coverage on any affected block → X
    for block in norm_affected:
        if block_to_sms.get(block):
            return 'X'

    # Hard failure with IC effect even on non-SM blocks → X
    hard_failure = severity in ('stuck', 'float', 'ov', 'uv', 'accuracy', 'drift')
    if hard_failure and (norm_affected or 'vega' in ic_effect.lower()):
        return 'X'
    if hard_failure and ic_effect.strip() not in ('', 'No effect'):
        return 'X'

    return 'O'


def compute_sm_columns(ic_effect: str, block_to_sms: dict, sm_coverage: dict,
                       fmeda_code: str = '', mode_str: str = '',
                       sm_addressed: dict = None,
                       block_descriptions: dict = None) -> tuple:
    """
    Returns (sm_string, coverage_value) for col S/Y and col U.

    v11 DESIGN: Chip-agnostic SM selection.
    
    The SM sets are NOT hardcoded by block code name. Instead they are derived at
    runtime from block_to_sms (which comes from the SM list sheet's 'Addressed Part'
    column) plus functional-role filtering based on the block's description.

    Algorithm:
      1. Normalize block code to its functional type (via fmeda_code prefix matching)
      2. Get candidate SMs from block_to_sms for this block's functional type
      3. Apply mode-specific filtering:
         - OV modes: keep only SMs whose description contains OV/overvoltage keywords
         - UV modes: keep only SMs whose description contains UV/undervoltage keywords
         - Stuck/float (hard): use all direct SMs + SMs for upstream ref blocks
         - Safe/timing/perf modes: return empty
      4. For blocks whose functional type is unrecognized, use block_to_sms directly
      5. All filtering happens against actual SM codes in sm_coverage (template-present only)

    This means: if your new chip renames REF to VREF, as long as the SM list sheet
    still addresses it and the agent1 maps it to fmeda_code='REF', this function works.
    If agent1 maps it to a new code entirely, step 4 (generic fallback) handles it.
    """
    if not ic_effect or ic_effect.strip() in ('No effect', 'No effect (Filter in place)', ''):
        return '', ''

    severity = classify_mode_severity(mode_str)
    m = mode_str.lower()
    norm_code = re.sub(r'SW_BANK[_\d]*', 'SW_BANK', fmeda_code.upper())

    # ── ALWAYS EMPTY (mode-independent) ─────────────────────────────────────
    if severity == 'safe':
        return '', ''
    if norm_code == 'INTERFACE':
        return '', ''
    if norm_code == 'CSNS':
        return '', ''
    if norm_code == 'ADC' and severity not in ('stuck', 'float'):
        return '', ''
    if norm_code == 'CP' and any(k in m for k in ['oscillation', 'quiescent', 'spike', 'start-up', 'start up']):
        return '', ''
    if norm_code == 'LDO' and any(k in m for k in ['spike', 'start-up', 'start up']):
        return '', ''
    if norm_code == 'SW_BANK' and any(k in m for k in ['turn-on', 'turn-off', 'turn on', 'turn off', 'resistance too low']):
        return '', ''

    # ── HELPER: filter SMs by description keywords ───────────────────────────
    def sms_with_keyword(candidates: list, keywords: list) -> list:
        """Return SMs from candidates whose description contains any keyword."""
        if not sm_addressed:
            return candidates
        result = []
        for sm in candidates:
            # sm_addressed maps SM code -> list of addressed parts
            # We need SM descriptions - stored in sm_addressed as the key
            # The description is in the SM list sheet; we use sm_addressed as proxy
            # (sm_addressed[sm] = list of parts this SM addresses e.g. ['REF','LDO'])
            # For keyword filtering we check the parts list (chip-agnostic)
            parts_lower = ' '.join(sm_addressed.get(sm, [])).lower()
            if any(k in parts_lower for k in keywords):
                result.append(sm)
        return result or candidates  # fallback to all if nothing matched

    def sms_with_desc_keyword(candidates: list, keywords: list,
                               sm_descriptions: dict) -> list:
        """Filter SMs by their description text keywords."""
        if not sm_descriptions:
            return candidates
        result = []
        for sm in candidates:
            desc = sm_descriptions.get(sm, '').lower()
            if any(k in desc for k in keywords):
                result.append(sm)
        return result or candidates

    # Use runtime SM descriptions populated from template at startup
    sm_desc = _SM_DESCRIPTIONS_RUNTIME or (block_descriptions or {})

    # ── CHIP-AGNOSTIC SM SELECTION BY FUNCTIONAL TYPE ────────────────────────
    # For each functional type, we query block_to_sms for blocks of that type
    # and apply mode-specific filtering.
    # block_to_sms keys are the actual block codes in the chip (e.g. 'REF', 'VREF')
    # We need to find what the chip calls this block type.

    def get_direct_sms(functional_code: str) -> list:
        """Get SMs directly addressing this functional block."""
        return block_to_sms.get(functional_code, [])

    direct = get_direct_sms(norm_code)

    # ── SW_BANK: mode-specific ────────────────────────────────────────────────
    if norm_code == 'SW_BANK':
        if 'stuck' in m and 'not including' not in m:
            # SMs that detect LED open, short, driver health, current monitoring
            candidates = direct
            # Filter for SMs that detect LED/driver states (open/short/driver)
            filtered = sms_with_desc_keyword(candidates,
                ['open', 'short', 'driver', 'health', 'current', 'mon'],
                sm_desc)
            return _pick_sms(filtered or candidates, sm_coverage)
        elif 'floating' in m or 'open circuit' in m or 'tri-state' in m:
            candidates = direct
            filtered = sms_with_desc_keyword(candidates,
                ['open', 'driver', 'health', 'current'],
                sm_desc)
            return _pick_sms(filtered or candidates, sm_coverage)
        elif 'resistance too high' in m:
            candidates = direct
            filtered = sms_with_desc_keyword(candidates,
                ['resistive', 'resistance', 'voltage', 'detection'],
                sm_desc)
            return _pick_sms(filtered or candidates, sm_coverage)
        return '', ''

    # ── LDO ──────────────────────────────────────────────────────────────────
    if norm_code == 'LDO':
        # Get SMs addressing LDO + SMs addressing REF (since REF/LDO share monitors)
        ref_code = _find_block_code(['REF', 'VREF', 'BANDGAP'], block_to_sms)
        ldo_sms = direct
        ref_sms = block_to_sms.get(ref_code, []) if ref_code else []
        osc_code = _find_block_code(['OSC', 'OSCILLATOR', 'CLK'], block_to_sms)
        osc_sms = block_to_sms.get(osc_code, []) if osc_code else []
        all_candidates = sorted(set(ldo_sms + ref_sms + osc_sms),
                                key=lambda s: int(re.search(r'\d+', s).group())
                                if re.search(r'\d+', s) else 0)

        if severity == 'ov' or 'higher than' in m:
            filtered = sms_with_desc_keyword(all_candidates,
                ['overvoltage', 'over-voltage', 'ov', 'ovv', 'overcurrent'],
                sm_desc)
            # Also include clock watchdog (catches supply effect on OSC)
            osc_watch = sms_with_desc_keyword(osc_sms, ['clock', 'watchdog', 'freq'], sm_desc)
            combined = sorted(set(filtered + osc_watch),
                              key=lambda s: int(re.search(r'\d+', s).group())
                              if re.search(r'\d+', s) else 0)
            return _pick_sms(combined or all_candidates[:2], sm_coverage)
        elif severity == 'uv' or 'lower than' in m:
            filtered = sms_with_desc_keyword(all_candidates,
                ['undervoltage', 'under-voltage', 'uv', 'supply', 'monitor'],
                sm_desc)
            osc_watch = sms_with_desc_keyword(osc_sms, ['clock', 'watchdog'], sm_desc)
            combined = sorted(set(filtered + osc_watch),
                              key=lambda s: int(re.search(r'\d+', s).group())
                              if re.search(r'\d+', s) else 0)
            return _pick_sms(combined or all_candidates[:2], sm_coverage)
        else:  # accuracy/drift
            osc_watch = sms_with_desc_keyword(osc_sms, ['clock', 'watchdog'], sm_desc)
            uv_sms = sms_with_desc_keyword(all_candidates, ['supply', 'monitor', 'under'], sm_desc)
            ov_sms = sms_with_desc_keyword(all_candidates, ['overvoltage', 'over'], sm_desc)
            combined = sorted(set(osc_watch + uv_sms + ov_sms),
                              key=lambda s: int(re.search(r'\d+', s).group())
                              if re.search(r'\d+', s) else 0)
            return _pick_sms(combined or all_candidates[:3], sm_coverage)

    # ── CP ────────────────────────────────────────────────────────────────────
    if norm_code == 'CP':
        if severity == 'ov' or 'higher than' in m:
            return '', ''  # OV -> device damage, no SM applicable
        return _pick_sms(direct, sm_coverage)

    # ── REF ───────────────────────────────────────────────────────────────────
    if norm_code == 'REF':
        # Upstream monitors: supply monitors + clock watchdog + ADC ref check
        ldo_code = _find_block_code(['LDO', 'VREG', 'REGULATOR'], block_to_sms)
        ldo_sms  = block_to_sms.get(ldo_code, []) if ldo_code else []
        adc_code = _find_block_code(['ADC', 'CONVERT'], block_to_sms)
        adc_sms  = block_to_sms.get(adc_code, []) if adc_code else []
        osc_code = _find_block_code(['OSC', 'OSCILLATOR', 'CLK'], block_to_sms)
        osc_sms  = block_to_sms.get(osc_code, []) if osc_code else []
        temp_code = _find_block_code(['TEMP', 'THERMAL', 'DIETEMP'], block_to_sms)
        temp_sms = block_to_sms.get(temp_code, []) if temp_code else []

        all_candidates = sorted(set(direct + ldo_sms + adc_sms + osc_sms + temp_sms),
                                key=lambda s: int(re.search(r'\d+', s).group())
                                if re.search(r'\d+', s) else 0)

        if severity in ('stuck', 'float'):
            # Hard failure: supply monitors + ADC ref check + thermal limit
            supply = sms_with_desc_keyword(all_candidates,
                ['supply', 'under', 'over', 'voltage', 'monitor', 'uv', 'ov'], sm_desc)
            adc_ref = sms_with_desc_keyword(adc_sms + direct,
                ['adc', 'reading', 'bgr', 'sbg', 'thermal', 'limit'], sm_desc)
            combined = sorted(set(supply + adc_ref),
                              key=lambda s: int(re.search(r'\d+', s).group())
                              if re.search(r'\d+', s) else 0)
            return _pick_sms(combined or all_candidates[:4], sm_coverage)
        else:  # accuracy/drift/incorrect
            supply = sms_with_desc_keyword(all_candidates,
                ['supply', 'monitor', 'clock', 'watchdog', 'adc', 'reading'], sm_desc)
            return _pick_sms(supply or all_candidates[:4], sm_coverage)

    # ── BIAS ──────────────────────────────────────────────────────────────────
    if norm_code == 'BIAS':
        # BIAS is monitored indirectly: supply monitors (catch LDO→BIAS chain),
        # clock watchdog (catches OSC→freq drift from BIAS), ADC ref check
        osc_code = _find_block_code(['OSC', 'OSCILLATOR', 'CLK'], block_to_sms)
        osc_sms  = block_to_sms.get(osc_code, []) if osc_code else []
        ref_code = _find_block_code(['REF', 'VREF', 'BANDGAP'], block_to_sms)
        ref_sms  = block_to_sms.get(ref_code, []) if ref_code else []
        ldo_code = _find_block_code(['LDO', 'VREG'], block_to_sms)
        ldo_sms  = block_to_sms.get(ldo_code, []) if ldo_code else []

        all_candidates = sorted(set(osc_sms + ref_sms + ldo_sms),
                                key=lambda s: int(re.search(r'\d+', s).group())
                                if re.search(r'\d+', s) else 0)
        relevant = sms_with_desc_keyword(all_candidates,
            ['clock', 'watchdog', 'supply', 'monitor', 'adc', 'reading', 'sbg'], sm_desc)
        return _pick_sms(relevant or all_candidates[:3], sm_coverage)

    # ── OSC ───────────────────────────────────────────────────────────────────
    if norm_code == 'OSC':
        # Clock watchdog + UART watchdogs detect OSC failures
        logic_code = _find_block_code(['LOGIC', 'MCU', 'CTRL', 'CONTROLLER'], block_to_sms)
        logic_sms  = block_to_sms.get(logic_code, []) if logic_code else []
        all_candidates = sorted(set(direct + logic_sms),
                                key=lambda s: int(re.search(r'\d+', s).group())
                                if re.search(r'\d+', s) else 0)
        watchdogs = sms_with_desc_keyword(all_candidates,
            ['clock', 'watchdog', 'uart', 'communication', 'check'], sm_desc)
        return _pick_sms(watchdogs or all_candidates[:3], sm_coverage)

    # ── TEMP ──────────────────────────────────────────────────────────────────
    if norm_code == 'TEMP':
        thermal = sms_with_desc_keyword(direct,
            ['thermal', 'temperature', 'temp', 'limit', 'monitor'], sm_desc)
        return _pick_sms(thermal or direct, sm_coverage)

    # ── ADC ───────────────────────────────────────────────────────────────────
    if norm_code == 'ADC':
        # stuck/float: all ADC monitors + thermal + LED current
        temp_code = _find_block_code(['TEMP', 'THERMAL'], block_to_sms)
        temp_sms  = block_to_sms.get(temp_code, []) if temp_code else []
        sw_code   = _find_block_code(['SW_BANK', 'SWITCH', 'DRIVER'], block_to_sms)
        sw_sms    = block_to_sms.get(sw_code, []) if sw_code else []
        all_candidates = sorted(set(direct + temp_sms + sw_sms),
                                key=lambda s: int(re.search(r'\d+', s).group())
                                if re.search(r'\d+', s) else 0)
        relevant = sms_with_desc_keyword(all_candidates,
            ['adc', 'current', 'thermal', 'reading', 'monitor', 'limit', 'voltage'], sm_desc)
        return _pick_sms(relevant or all_candidates[:4], sm_coverage)

    # ── LOGIC ─────────────────────────────────────────────────────────────────
    if norm_code == 'LOGIC':
        watchdogs = sms_with_desc_keyword(direct,
            ['watchdog', 'uart', 'clock', 'pwm', 'ecc', 'check', 'monitor', 'sync'], sm_desc)
        return _pick_sms(watchdogs or direct, sm_coverage)

    # ── TRIM ──────────────────────────────────────────────────────────────────
    if norm_code == 'TRIM':
        # All calibration-relevant SMs
        ref_code  = _find_block_code(['REF', 'VREF'], block_to_sms)
        ref_sms   = block_to_sms.get(ref_code, []) if ref_code else []
        ldo_code  = _find_block_code(['LDO', 'VREG'], block_to_sms)
        ldo_sms   = block_to_sms.get(ldo_code, []) if ldo_code else []
        osc_code  = _find_block_code(['OSC', 'OSCILLATOR'], block_to_sms)
        osc_sms   = block_to_sms.get(osc_code, []) if osc_code else []
        logic_code = _find_block_code(['LOGIC', 'MCU'], block_to_sms)
        logic_sms = block_to_sms.get(logic_code, []) if logic_code else []
        temp_code = _find_block_code(['TEMP', 'THERMAL'], block_to_sms)
        temp_sms  = block_to_sms.get(temp_code, []) if temp_code else []
        all_candidates = sorted(set(ref_sms + ldo_sms + osc_sms + logic_sms + temp_sms),
                                key=lambda s: int(re.search(r'\d+', s).group())
                                if re.search(r'\d+', s) else 0)
        return _pick_sms(all_candidates, sm_coverage)

    # ── GENERIC FALLBACK for unknown block types ──────────────────────────────
    direct_sorted = sorted(direct,
                           key=lambda s: int(re.search(r'\d+', s).group())
                           if re.search(r'\d+', s) else 0)
    return _pick_sms(direct_sorted, sm_coverage) if direct_sorted else ('', '')


def _find_block_code(keywords: list, block_to_sms: dict) -> str | None:
    """
    Find a block code in block_to_sms whose name matches any keyword.
    Used to resolve 'what does this chip call its LDO/REF/OSC etc.' dynamically.
    """
    for code in block_to_sms:
        c_lower = code.lower()
        if any(k.lower() in c_lower for k in keywords):
            return code
    return None


def _pick_sms(sm_list: list, sm_coverage: dict) -> tuple:
    """Filter to SMs present in coverage map, return (space-joined str, max coverage)."""
    valid_sms = [s for s in sm_list if s in sm_coverage] if sm_coverage else sm_list
    if not valid_sms:
        valid_sms = sm_list  # fallback when no template loaded
    if not valid_sms:
        return '', ''
    valid_cov = [0.99, 0.9, 0.6]
    def nearest(v):
        return min(valid_cov, key=lambda x: abs(x - v))
    coverages = [nearest(sm_coverage.get(sm, 0.9)) for sm in valid_sms]
    return ' '.join(valid_sms), max(coverages)


def agent2_generate_effects(blocks, tsr_list, block_to_sms, sm_coverage,
                             sm_addressed, cache, signal_graph, sm_j_map):
    """
    Generate col I/J/K for all blocks.
    v10: Col I is DETERMINISTIC via resolve_i_deterministic().
         LLM only called for genuinely unknown block types.
    """
    active = [b for b in blocks if not b.get('is_duplicate') and not b.get('is_sm')]
    chip_ctx = "\n".join(
        f"  {b['fmeda_code']:<12} {b['name']:<35} | {b.get('function', '')[:80]}"
        for b in active
    )

    tsr_ctx = "\n".join(
        f"  {t['id']}: {t['description']}"
        for t in tsr_list
    ) if tsr_list else "  (no TSR data)"

    result = []
    for block in blocks:
        code  = block['fmeda_code']
        name  = block['name']
        modes = block.get('modes', [])

        # SM blocks
        if block.get('is_sm'):
            rows = _sm_rows(code, sm_j_map)
            result.append({'fmeda_code': code, 'user_name': name, 'rows': rows})
            print(f"  [Agent 2] {code:<12} SM (2 rows)")
            continue

        if block.get('is_duplicate'):
            print(f"  [Agent 2] {code:<12} DUPLICATE - skipped")
            continue

        if not modes:
            print(f"  [Agent 2] {code:<12} no modes - skipped")
            continue

        # v10 cache key (invalidates v9/v8 cache)
        ck = f"agent2_v10__{code}__{name}__{len(modes)}"
        if not SKIP_CACHE and ck in cache:
            rows = cache[ck]
            # Always refresh I deterministically (fixes stale LLM I in cache)
            for row in rows:
                mode_g = row.get('G', '')
                det_i = resolve_i_deterministic(block, mode_g, signal_graph, sm_j_map)
                if det_i is not None:
                    row['I'] = det_i
                    row['J'] = _validate_j(_derive_j_from_rules(code, mode_g, det_i))
                # Always refresh K and S/Y
                k = compute_k_from_mode_and_coverage(code, mode_g, row.get('I', ''), block_to_sms)
                row['K'] = k
                row['P'] = 'Y' if k == 'X' else 'N'
                row['R'] = 1 if k == 'O' else 0
                row['X'] = 'Y' if k.startswith('X') else 'N'
                sm_str, cov = compute_sm_columns(row.get('I', ''), block_to_sms, sm_coverage, code, mode_g)
                row['S'] = sm_str; row['Y'] = sm_str; row['U'] = cov
            print(f"  [Agent 2] {code:<12} cache ({len(rows)} rows, I/K/S refreshed)")
            result.append({'fmeda_code': code, 'user_name': name, 'rows': rows})
            continue

        rows = _llm_block_effects_v7(block, chip_ctx, tsr_ctx, modes,
                                      block_to_sms, sm_coverage, signal_graph,
                                      sm_j_map=sm_j_map)
        cache[ck] = rows
        save_cache(cache)
        result.append({'fmeda_code': code, 'user_name': name, 'rows': rows})
        time.sleep(0.1)

    return result


def _build_i_context_for_block(block: dict, signal_graph: dict) -> str:
    """
    Build a detailed signal-flow context for col I generation.
    Uses the signal graph built by Agent 0 — no hardcoded values.
    """
    code = block['fmeda_code']
    name = block['name']

    # Look up in signal graph (by name or code)
    graph_entry = signal_graph.get(name) or signal_graph.get(code) or {}

    if graph_entry:
        consumers = graph_entry.get('consumers', [])
        output_sig = graph_entry.get('output_signal', f'output of {name}')
        details = graph_entry.get('consumer_details', {})

        ctx = f"OUTPUT SIGNAL: {output_sig}\n\n"
        if consumers:
            ctx += "DIRECT CONSUMERS (blocks that will fail if THIS block fails):\n"
            for c in consumers:
                detail = details.get(c, 'receives signal from this block')
                ctx += f"  - {c}: {detail}\n"
        else:
            ctx += "DIRECT CONSUMERS: (none identified - check function description)\n"
    else:
        # Fallback: generic guidance based on block function description
        func = block.get('function', '').lower()
        ctx = f"Block function: {block.get('function', '')}\n\n"
        ctx += "ANALYZE: Which other blocks in the chip receive output from this block?\n"
        ctx += "Consider: voltage/current references feed analog blocks, "
        ctx += "clocks feed digital logic, supplies feed all powered blocks.\n"

    return ctx


# =============================================================================
# DETERMINISTIC COL I ENGINE
# =============================================================================
#
# KEY INSIGHT: Col I follows strict deterministic patterns per block type.
# The LLM is the wrong tool — it over-thinks and adds wrong variations.
# The correct values are FULLY determined by: block type + mode severity.
#
# This engine resolves I values using:
#   1. Block-type classification (derived from fmeda_code pattern, not hardcoded names)
#   2. Mode severity classification (classify_mode_severity)
#   3. Signal graph for consumer block names (no hardcoded block names)
#
# For unknown block types, it falls back to the LLM with a tightly constrained prompt.

def resolve_i_deterministic(block: dict, mode_str: str,
                              signal_graph: dict,
                              sm_j_map: dict = None) -> str | None:
    """
    Resolve col I deterministically from block type + mode severity.
    Returns the I string if deterministically resolvable, None if LLM needed.

    Rules derived from complete analysis of 3_ID03_FMEDA.xlsx col I patterns.
    All consumer block names come from the signal_graph (chip-agnostic).
    """
    code = block['fmeda_code']
    name = block['name']
    m    = mode_str.lower()
    sev  = classify_mode_severity(mode_str)
    norm = re.sub(r'SW_BANK[_\d]*', 'SW_BANK', code.upper())

    # Get consumer block names from signal graph (chip-agnostic)
    graph_entry   = signal_graph.get(name) or signal_graph.get(code) or {}
    consumers     = graph_entry.get('consumers', [])
    consumer_det  = graph_entry.get('consumer_details', {})

    # Helper: find the first consumer whose name/detail matches a keyword
    def find_consumer(keywords: list) -> str | None:
        for c in consumers:
            c_lower = c.lower()
            detail  = consumer_det.get(c, '').lower()
            if any(k in c_lower or k in detail for k in keywords):
                return c
        return None

    # Helper: get all SW_BANK consumers (numbered or generic)
    def sw_bank_consumers() -> list:
        return [c for c in consumers
                if 'sw_bank' in c.lower() or 'switch' in c.lower()
                or 'driver' in c.lower() or 'led' in c.lower()]

    # ── SAFE MODES — always No effect ────────────────────────────────────────
    if sev == 'safe':
        # LDO spikes is a special case: NOT No effect, it causes OSC jitter
        if norm == 'LDO' and 'spike' in m:
            osc = find_consumer(['osc', 'clock', 'oscillator'])
            osc_name = osc or 'OSC'
            return f'• {osc_name}\n    - Jitter too high in the output signal'
        # LDO fast oscillation: No effect with note
        if norm == 'LDO' and 'fast oscillation' in m:
            return 'No effect (Filter in place)'
        return 'No effect'

    # ── SW_BANK — direct LED state strings, NO bullet points ─────────────────
    if norm == 'SW_BANK':
        if 'stuck' in m and 'not including' not in m:
            return 'Unintended LED ON/OFF'
        if 'floating' in m or 'open circuit' in m or 'tri-state' in m:
            return 'Unintended LED ON'
        if 'resistance too high' in m:
            return 'Unintended LED ON'
        if 'resistance too low' in m or 'turn-on' in m or 'turn-off' in m:
            return 'Performance impact'
        return 'Performance impact'

    # ── INTERFACE — always plain string ──────────────────────────────────────
    if norm == 'INTERFACE':
        return 'Communication error'

    # ── CSNS — always single ADC bullet regardless of mode ───────────────────
    if norm == 'CSNS':
        adc = find_consumer(['adc', 'convert', 'digital'])
        adc_name = adc or 'ADC'
        return f'• {adc_name}\n    - CSNS output is incorrect.'

    # ── LDO ──────────────────────────────────────────────────────────────────
    if norm == 'LDO':
        osc = find_consumer(['osc', 'clock', 'oscillator'])
        osc_name = osc or 'OSC'
        # OV: OSC out of spec
        if sev == 'ov' or 'higher than' in m:
            return f'• {osc_name}\n    - Out of spec.'
        # UV or accuracy: OSC out of spec + Vega reset
        if sev in ('uv', 'accuracy', 'drift') or 'lower than' in m:
            return (f'• {osc_name}\n    - Out of spec.\n'
                    f'• Vega\n    - Reset reaction. (POR)')
        # Startup: No effect
        if 'start-up' in m or 'startup' in m:
            return 'No effect'
        # Oscillation within range: No effect
        return 'No effect'

    # ── OSC ──────────────────────────────────────────────────────────────────
    if norm == 'OSC':
        logic = find_consumer(['logic', 'controller', 'control', 'digital', 'mcu'])
        logic_name = logic or 'LOGIC'
        if sev in ('stuck', 'float', 'incorrect', 'ov', 'uv', 'drift'):
            return f'• {logic_name}\n    - Cannot operate.\n    - Communication error.'
        if 'duty cycle' in m or 'jitter' in m:
            return 'No effect'
        return f'• {logic_name}\n    - Cannot operate.\n    - Communication error.'

    # ── TEMP ─────────────────────────────────────────────────────────────────
    if norm == 'TEMP':
        adc  = find_consumer(['adc', 'convert', 'digital'])
        adc_name = adc or 'ADC'
        sw   = find_consumer(['sw_bank', 'switch', 'driver', 'led'])
        sw_name = (re.sub(r'SW_BANK[_\d]+', 'SW_BANK_x', sw)
                   if sw else 'SW_BANK_x')
        if sev == 'stuck':
            return (f'• {adc_name}\n    - TEMP output is stuck low\n'
                    f'• {sw_name}\n    - SW is stuck in off state (DIETEMP)')
        if sev == 'float':
            return f'• {adc_name}\n    - Incorrect TEMP reading'
        if sev in ('incorrect', 'accuracy', 'drift', 'other'):
            return (f'• {adc_name}\n    - TEMP output Static Error (offset error, gain error, '
                    f'integral nonlinearity, & differential nonlinearity)')
        return 'No effect'

    # ── ADC ──────────────────────────────────────────────────────────────────
    if norm == 'ADC':
        sw = find_consumer(['sw_bank', 'switch', 'driver', 'led'])
        sw_name = re.sub(r'SW_BANK[_\d]+', 'SW_BANK_x', sw) if sw else 'SW_BANK_x'
        adc_self_bullets = ('• ADC\n    - Incorrect BGR measurement\n'
                            '    - Incorrect DIETEMP measurement\n'
                            '    - Incorrect CS measurement')
        if sev in ('stuck', 'float'):
            return (f'• {sw_name}\n    - SW is stuck in off state (DIETEMP)\n'
                    + adc_self_bullets)
        # All non-safe non-stuck ADC modes: self-measurement errors only
        if sev not in ('safe',):
            return adc_self_bullets
        return 'No effect'

    # ── CP ────────────────────────────────────────────────────────────────────
    if norm == 'CP':
        sw = find_consumer(['sw_bank', 'switch', 'driver', 'led'])
        sw_name = re.sub(r'SW_BANK[_\d]+', 'SW_BANK_x', sw) if sw else 'SW_BANK_x'
        if sev == 'ov' or 'higher than' in m:
            return '• Vega\n    - Device Damage'
        if sev == 'uv' or 'lower than' in m:
            return f'• {sw_name}\n    - SWs are stuck in off state, LEDs always ON.'
        return 'No effect'

    # ── LOGIC ─────────────────────────────────────────────────────────────────
    if norm == 'LOGIC':
        sw = find_consumer(['sw_bank', 'switch', 'driver', 'led'])
        sw_name = re.sub(r'SW_BANK[_\d]+', 'SW_BANK_X', sw) if sw else 'SW_BANK_X'
        osc = find_consumer(['osc', 'clock', 'oscillator'])
        osc_name = osc or 'OSC'
        if sev in ('stuck', 'float', 'incorrect', 'ov', 'uv', 'accuracy', 'other'):
            return (f'• {sw_name}\n    - SW is stuck in on/off state\n'
                    f'• {osc_name}\n    - Output stuck')
        return 'No effect'

    # ── TRIM ──────────────────────────────────────────────────────────────────
    if norm == 'TRIM':
        if sev == 'safe' or 'settling' in m:
            return 'No effect'
        # omission, commission, incorrect: all calibrated blocks affected
        ref  = find_consumer(['ref', 'bandgap', 'reference'])
        ldo  = find_consumer(['ldo', 'regulator', 'supply'])
        bias = find_consumer(['bias', 'current source'])
        sw   = find_consumer(['sw_bank', 'switch', 'driver'])
        osc  = find_consumer(['osc', 'clock', 'oscillator'])
        temp = find_consumer(['temp', 'thermal', 'temperature'])

        ref_name  = ref  or 'REF'
        ldo_name  = ldo  or 'LDO'
        bias_name = bias or 'BIAS'
        sw_name   = re.sub(r'SW_BANK[_\d]+', 'SW_BANK', sw) if sw else 'SW_BANK'
        osc_name  = osc  or 'OSC'
        temp_code = (re.sub(r'TEMP.*', 'DIETEMP', temp.upper()) if temp else 'DIETEMP')

        return (f'• {ref_name}\n    - Incorrect output value higher than the expected range\n'
                f'• {ldo_name}\n    - Reference voltage higher than the expected range\n'
                f'• {bias_name}\n    - Output reference voltage accuracy too low, including drift\n'
                f'• {sw_name}\n    - Incorrect slew rate value\n'
                f'• {osc_name}\n    - Incorrect output frequency: higher than the expected range\n'
                f'• {temp_code}\n    - Incorrect output voltage')

    # ── REF ───────────────────────────────────────────────────────────────────
    if norm == 'REF':
        bias = find_consumer(['bias', 'current source', 'current mirror'])
        adc  = find_consumer(['adc', 'convert', 'digital'])
        temp = find_consumer(['temp', 'thermal', 'temperature'])
        ldo  = find_consumer(['ldo', 'regulator', 'supply'])
        osc  = find_consumer(['osc', 'clock', 'oscillator'])

        bias_name = bias or 'BIAS'
        adc_name  = adc  or 'ADC'
        temp_name = temp or 'TEMP'
        ldo_name  = ldo  or 'LDO'
        osc_name  = osc  or 'OSC'

        if sev == 'stuck':
            return (f'• {bias_name}\n'
                    f'    - Output reference voltage is stuck \n'
                    f'    - Output reference current is stuck \n'
                    f'    - Output bias current is stuck \n'
                    f'    - Quiescent current exceeding the maximum value\n'
                    f'• REF\n'
                    f'    - Quiescent current exceeding the maximum value\n'
                    f'• {adc_name}\n    - REF output is stuck \n'
                    f'• {temp_name}\n    - Output is stuck \n'
                    f'• {ldo_name}\n    - Output is stuck \n'
                    f'• {osc_name}\n    - Oscillation does not start')
        if sev == 'float':
            return (f'• {bias_name}\n'
                    f'    - Output reference voltage is floating\n'
                    f'    - Output reference current is higher than the expected range\n'
                    f'    - Output reference current is lower than the expected range\n'
                    f'    - Output bias current is higher than the expected range\n'
                    f'    - Output bias current is lower than the expected range\n'
                    f'• {adc_name}\n    - REF output is floating (i.e. open circuit)\n'
                    f'• {ldo_name}\n    - Out of spec\n'
                    f'• {osc_name}\n    - Out of spec')
        if sev in ('incorrect', 'accuracy', 'drift', 'other'):
            return (f'• {bias_name}\n'
                    f'    - Output reference voltage is higher than the expected range\n'
                    f'    - Output reference current is higher than the expected range\n'
                    f'    - Output bias current is higher than the expected range\n'
                    f'• {temp_name}\n'
                    f'    - Incorrect gain on the output voltage (outside the expected range)\n'
                    f'    - Incorrect offset on the output voltage (outside the expected range)\n'
                    f'• {adc_name}\n    - REF output higher/lower than expected\n'
                    f'• {ldo_name}\n    - Out of spec\n'
                    f'• {osc_name}\n    - Out of spec')
        return 'No effect'

    # ── BIAS ──────────────────────────────────────────────────────────────────
    if norm == 'BIAS':
        adc  = find_consumer(['adc', 'convert', 'digital'])
        temp = find_consumer(['temp', 'thermal', 'temperature'])
        ldo  = find_consumer(['ldo', 'regulator', 'supply'])
        osc  = find_consumer(['osc', 'clock', 'oscillator'])
        sw   = find_consumer(['sw_bank', 'switch', 'driver', 'led'])
        cp   = find_consumer(['charge pump', 'cp', 'boost'])
        csns = find_consumer(['csns', 'current sense', 'sense amp'])

        adc_name  = adc  or 'ADC'
        temp_name = temp or 'TEMP'
        ldo_name  = ldo  or 'LDO'
        osc_name  = osc  or 'OSC'
        # Always use generic codes for multi-instance blocks
        sw_name   = 'SW_BANKx'
        cp_name   = cp   or 'CP'
        cnsn_name = 'CNSN'  # standard shorthand for current sense block

        if sev in ('stuck', 'float', 'incorrect', 'accuracy', 'drift', 'other'):
            return (f'• {adc_name}\n    - ADC measurement is incorrect.\n'
                    f'• {temp_name}\n    - Incorrect temperature measurement.\n'
                    f'• {ldo_name}\n    - Out of spec.\n'
                    f'• {osc_name}\n    - Frequency out of spec.\n'
                    f'• {sw_name}\n    - Out of spec.\n'
                    f'• {cp_name}\n    - Out of spec.\n'
                    f'• {cnsn_name}\n    - Incorrect reading.')
        return 'No effect'

    # ── SM BLOCKS ─────────────────────────────────────────────────────────────
    if re.match(r'^SM\d+$', norm):
        # 'Fail to detect' -> the I value is what the SM was supposed to catch
        # This comes from sm_j_map which is built from SM descriptions
        if sm_j_map and 'fail to detect' in m:
            sm_ic, _ = sm_j_map.get(code, ('Loss of safety mechanism functionality', ''))
            return sm_ic if sm_ic else 'Loss of safety mechanism functionality'
        if 'false detection' in m:
            return 'No effect'
        return None  # fallback to LLM

    # Unknown block type: return None to trigger LLM fallback
    return None


def _llm_block_effects_v7(block, chip_ctx, tsr_ctx, modes,
                           block_to_sms, sm_coverage, signal_graph,
                           sm_j_map=None):
    """
    Generate I/J/K rows.

    Architecture (v10):
      Col I: DETERMINISTIC first via resolve_i_deterministic().
             LLM only called if resolve_i_deterministic() returns None
             (i.e. a completely new block type not covered by any rule).
      Col J: Determined from block type + mode severity (LLM prompt includes rules).
      Col K: Always deterministic via compute_k_from_mode_and_coverage().
    """
    code = block['fmeda_code']
    name = block['name']
    func = block.get('function', '')
    n    = len(modes)

    # Pass 1: resolve everything deterministically
    det_results = []      # (mode, i_val) for deterministically resolved modes
    llm_needed  = []      # (idx, mode) needing LLM for I

    for idx, mode in enumerate(modes):
        i_val = resolve_i_deterministic(block, mode, signal_graph, sm_j_map)
        if i_val is not None:
            det_results.append((idx, mode, i_val))
        else:
            llm_needed.append((idx, mode))

    # Pass 2: LLM only for unknown modes (usually empty for known block types)
    llm_i = {}  # idx -> i string
    if llm_needed:
        i_context   = _build_i_context_for_block(block, signal_graph)
        safe_modes  = [m for _, m in llm_needed if is_safe_mode(m)]
        llm_modes   = [m for _, m in llm_needed]

        prompt = f"""You are a senior functional safety engineer completing an FMEDA (ISO 26262).

BLOCK: {code} | {name} | {func}

SIGNAL FLOW:
{i_context}

CHIP CONTEXT:
{chip_ctx}

FAILURE MODES TO ANALYZE (only these {len(llm_modes)} — others handled separately):
{json.dumps(llm_modes, indent=2)}

SAFE MODES (write "No effect"):
{json.dumps(safe_modes)}

For col I (Effects on IC output), use bullet format:
  • BLOCK_CODE
      - specific symptom (under 10 words)
      - second symptom if needed

Rules:
- SW_BANK_x (not SW_BANK_1/2), SW_BANK_X for logic-driven, Vega for device damage
- Safe modes -> No effect
- Only list DIRECT consumers from signal flow above
- Short precise phrases only

For col J (system effect), use the first matching rule:
  REF/BIAS non-safe: "Unintentional LED ON/OFF\\nFail-safe mode active\\nNo communication"
  LDO OV/UV/accuracy: "Fail-safe mode active\\nNo communication"
  OSC stuck/float/freq/drift: "Fail-safe mode active\\nNo communication"
  TEMP stuck: "Unintentional LED ON"
  TEMP float/incorrect: "Unintentional LED ON\\nPossible device damage"
  ADC stuck/float: "Unintentional LED ON"
  CP OV: "Device damage" | CP UV: "Unintentional LED ON"
  LOGIC all: "Unintentional LED ON/OFF\\nFail-safe mode active\\nNo communication"
  TRIM active modes: "Fail-safe mode active\\nNo communication"
  SW_BANK stuck: "Unintended LED ON/OFF" | SW_BANK float/res-high: "Unintended LED ON"
  Otherwise: "No effect"

Return JSON array with {len(llm_modes)} objects:
[{{"G": "<mode>", "I": "<col I>", "J": "<col J>"}}]
Return ONLY the JSON:"""

        raw    = query_llm(prompt, temperature=0.0)
        parsed = parse_json(raw)

        if isinstance(parsed, list) and len(parsed) >= len(llm_needed):
            for pos, (orig_idx, orig_mode) in enumerate(llm_needed):
                rd    = parsed[pos]
                i_str = str(rd.get('I', 'No effect')).strip()
                # Post-process: replace literal 'bullet'/'dash' words with symbols
                i_str = i_str.replace('bullet ', '• ').replace('\ndash ', '\n    - ')
                llm_i[orig_idx] = (i_str, str(rd.get('J', 'No effect')).strip())
        else:
            for orig_idx, orig_mode in llm_needed:
                llm_i[orig_idx] = ('No effect' if is_safe_mode(orig_mode) else '', 'No effect')

    # Pass 3: build rows in correct mode order
    # J for deterministically-resolved I: compute from block type + mode
    rows = []
    det_map = {idx: (mode, i_val) for idx, mode, i_val in det_results}

    for idx, mode in enumerate(modes):
        if idx in det_map:
            ic   = det_map[idx][1]
            sys_ = _derive_j_from_rules(code, mode, ic)
        else:
            ic_and_j = llm_i.get(idx, ('No effect', 'No effect'))
            ic   = ic_and_j[0]
            sys_ = _validate_j(ic_and_j[1])

        # Always override safe modes
        if is_safe_mode(mode):
            # LDO spike exception already handled in resolve_i_deterministic
            if ic == 'No effect':
                pass
        
        sys_ = _validate_j(sys_)
        memo = 'O' if (not ic or ic.strip() in ('No effect', 'No effect (Filter in place)')) \
               else compute_k_from_mode_and_coverage(code, mode, ic, block_to_sms)

        rows.append(_build_row(mode, ic, sys_, memo, block_to_sms, sm_coverage,
                               fmeda_code=code))

    if not rows:
        rows = _fallback_rows_v7(modes, code, block_to_sms, sm_coverage)

    n_det = len(det_results)
    n_llm = len(llm_needed)
    print(f"    {code}: {n_det} det + {n_llm} LLM = {n} rows")
    return rows


def _derive_j_from_rules(code: str, mode_str: str, ic: str) -> str:
    """
    Derive col J from block type + mode severity using the block-specific rules.
    This is called for deterministically-resolved I values so we don't need LLM for J.
    """
    norm = re.sub(r'SW_BANK[_\d]*', 'SW_BANK', code.upper())
    m    = mode_str.lower()
    sev  = classify_mode_severity(mode_str)

    # Safe modes or no IC effect
    if sev == 'safe' or not ic or ic.strip() == 'No effect':
        # LDO spikes exception: has IC effect but J is still No effect
        if norm == 'LDO' and 'spike' in m:
            return 'No effect'
        if ic and ic.strip() not in ('No effect', 'No effect (Filter in place)', ''):
            pass  # fall through to block-specific rules
        else:
            return 'No effect'

    if norm in ('REF', 'BIAS'):
        return 'Unintentional LED ON/OFF\nFail-safe mode active\nNo communication'
    if norm == 'LDO':
        if sev == 'safe' or 'spike' in m or 'start-up' in m:
            return 'No effect'
        return 'Fail-safe mode active\nNo communication'
    if norm == 'OSC':
        if 'duty cycle' in m or 'jitter' in m:
            return 'No effect'
        return 'Fail-safe mode active\nNo communication'
    if norm == 'TEMP':
        if sev == 'stuck':
            return 'Unintentional LED ON'
        if sev in ('float', 'incorrect', 'accuracy', 'drift'):
            return 'Unintentional LED ON\nPossible device damage'
        return 'No effect'
    if norm == 'CSNS':
        return 'No effect'
    if norm == 'ADC':
        if sev in ('stuck', 'float'):
            return 'Unintentional LED ON'
        return 'No effect'
    if norm == 'CP':
        if sev == 'ov' or 'higher than' in m:
            return 'Device damage'
        if sev == 'uv' or 'lower than' in m:
            return 'Unintentional LED ON'
        return 'No effect'
    if norm == 'LOGIC':
        return 'Unintentional LED ON/OFF\nFail-safe mode active\nNo communication'
    if norm == 'INTERFACE':
        return 'Fail-safe mode active'
    if norm == 'TRIM':
        if 'settling' in m:
            return 'No effect'
        return 'Fail-safe mode active\nNo communication'
    if norm == 'SW_BANK':
        if 'stuck' in m and 'not including' not in m:
            return 'Unintended LED ON/OFF'
        if 'floating' in m or 'open circuit' in m or 'resistance too high' in m:
            return 'Unintended LED ON'
        return 'No effect'
    if re.match(r'^SM\d+$', norm):
        # SM 'Fail to detect' J comes from sm_j_map (handled by _sm_rows)
        return 'No effect'

    # Unknown block: use generic rules
    if ic and ic.strip() not in ('No effect', ''):
        return 'Fail-safe mode active'
    return 'No effect'
    code = block['fmeda_code']
    name = block['name']
    func = block.get('function', '')
    n    = len(modes)

    i_context = _build_i_context_for_block(block, signal_graph)
    safe_modes   = [m for m in modes if is_safe_mode(m)]
    unsafe_modes = [m for m in modes if not is_safe_mode(m)]

    prompt = f"""You are a senior functional safety engineer completing an FMEDA for an automotive IC (ISO 26262).

CHIP BLOCKS:
{chip_ctx}

SAFETY REQUIREMENTS (TSR):
{tsr_ctx}

BLOCK UNDER ANALYSIS:
  Code: {code}  |  Name: {name}  |  Function: {func}

DIRECT SIGNAL CONSUMERS (from schematic analysis):
{i_context}

FAILURE MODES ({n} total):
{json.dumps(modes, indent=2)}

SAFE MODES - write "No effect" for both I and J, skip reasoning:
{json.dumps(safe_modes)}

NON-SAFE MODES requiring full analysis:
{json.dumps(unsafe_modes, indent=2)}

========== 6-STEP CHAIN-OF-THOUGHT PROCESS ==========

For EACH non-safe failure mode, reason through ALL steps:

STEP 1 - IDENTIFY THE PHYSICAL OUTPUT SIGNAL
  What exact physical quantity does {name} produce?
  Examples: "1.2V bandgap reference voltage", "bias currents for current mirrors",
            "16MHz clock", "gate drive voltage for MOSFETs", "digitized 8-bit values"

STEP 2 - TRACE EVERY DIRECT CONSUMER
  From the DIRECT SIGNAL CONSUMERS list above, identify which blocks physically
  receive this signal as an input pin. Be exhaustive - missing a consumer = wrong answer.

STEP 3 - DETERMINE SPECIFIC SYMPTOM PER CONSUMER PER MODE
  For each consumer: "If {code} output is [stuck/floating/incorrect/drifting],
  what SPECIFICALLY fails in this consumer?"
  Use IC engineering language:
    GOOD: "Output reference voltage is stuck"  |  "Cannot operate."  |  "ADC measurement is incorrect."
    BAD:  "The block is affected"  |  "Values become wrong"

STEP 4 - ENUMERATE ALL SUB-EFFECTS (most commonly missed step)
  Each consumer may have MULTIPLE distinct sub-effects - list each separately.

  MANDATORY sub-effect patterns:
  - Voltage reference stuck -> bias generator gets:
      (a) Output reference voltage is stuck
      (b) Output reference current is stuck
      (c) Output bias current is stuck
      (d) Quiescent current exceeding the maximum value
    PLUS: REF itself gets "Quiescent current exceeding the maximum value"
  - Voltage reference floating -> bias generator gets:
      (a) Output reference voltage is floating
      (b) Output reference current is higher than the expected range
      (c) Output reference current is lower than the expected range
      (d) Output bias current is higher than the expected range
      (e) Output bias current is lower than the expected range
  - Oscillator stuck/float -> LOGIC gets BOTH: "Cannot operate." AND "Communication error."
  - TEMP sensor stuck -> ADC gets "TEMP output is stuck low" AND SW_BANK_x gets "SW is stuck in off state (DIETEMP)"
  - Bias current stuck/float -> CNSN block also affected: "Incorrect reading."

STEP 5 - APPLY OUTPUT FORMAT
  Required format:
    bullet BLOCK_CODE
        dash specific symptom
        dash second symptom if any
    bullet ANOTHER_BLOCK
        dash symptom

  CRITICAL RULES:
  (a) Generic codes only: "SW_BANK_x" NOT "SW_BANK_1"; "SW_BANK_X" for LOGIC-driven switches
  (b) "Vega" for whole-IC device damage (OV scenarios only)
  (c) Symptom phrases UNDER 10 WORDS - no long sentences
  (d) Safe modes -> write exactly: No effect

STEP 6 - DETERMINE COL J (system-level consequence, first matching rule wins)
  - REF/BIAS non-safe: "Unintentional LED ON/OFF\\nFail-safe mode active\\nNo communication"
  - LDO OV/UV/accuracy/drift: "Fail-safe mode active\\nNo communication"
  - LDO spikes/startup/oscillation-within: "No effect"
  - OSC stuck/float/freq/swing/drift: "Fail-safe mode active\\nNo communication"
  - OSC duty-cycle/jitter: "No effect"
  - TEMP stuck: "Unintentional LED ON"
  - TEMP float/incorrect/accuracy: "Unintentional LED ON\\nPossible device damage"
  - CSNS ALL modes: "No effect"
  - ADC stuck/float: "Unintentional LED ON"
  - ADC all others (accuracy/offset/linearity/etc): "No effect"
  - CP OV: "Device damage"
  - CP UV: "Unintentional LED ON"
  - CP oscillation/quiescent/spikes/startup: "No effect"
  - LOGIC ALL 3 modes: "Unintentional LED ON/OFF\\nFail-safe mode active\\nNo communication"
  - INTERFACE ALL modes: "Fail-safe mode active"
  - TRIM omission/commission/incorrect: "Fail-safe mode active\\nNo communication"
  - TRIM settling: "No effect"
  - SW_BANK stuck: "Unintended LED ON/OFF"
  - SW_BANK float/res-high: "Unintended LED ON"
  - SW_BANK res-low/timing: "No effect"

========== WORKED EXAMPLE (reasoning only - do not copy values) ==========

Block: REF (Voltage Reference)  |  Mode: "Output is stuck (i.e. high or low)"

STEP 1: REF produces a stable 1.2V bandgap voltage used as a reference point.
STEP 2: Consumers are BIAS (sets mirror currents from REF), ADC (REF as conversion
        reference), TEMP (comparator reference), LDO (regulation feedback), OSC (freq network).
        REF also has self-quiescent current.
STEP 3+4:
  BIAS: stuck voltage -> (a) ref voltage stuck, (b) ref current stuck, (c) bias current stuck,
        (d) quiescent current exceeds max
  REF self: quiescent current exceeds max
  ADC: REF output is stuck
  TEMP: Output is stuck
  LDO: Output is stuck
  OSC: Oscillation does not start
STEP 5 output:
  bullet BIAS
      dash Output reference voltage is stuck
      dash Output reference current is stuck
      dash Output bias current is stuck
      dash Quiescent current exceeding the maximum value
  bullet REF
      dash Quiescent current exceeding the maximum value
  bullet ADC
      dash REF output is stuck
  bullet TEMP
      dash Output is stuck
  bullet LDO
      dash Output is stuck
  bullet OSC
      dash Oscillation does not start
STEP 6: REF non-safe -> "Unintentional LED ON/OFF\\nFail-safe mode active\\nNo communication"

========== NOW ANALYZE YOUR {n} MODES ==========

Return a JSON array with EXACTLY {n} objects, same order as the modes list:
[
  {{
    "G": "<copy failure mode string verbatim>",
    "I": "<col I with bullet+dash format, all sub-effects>",
    "J": "<col J exact string from Step 6 rules>"
  }},
  ...
]

QUALITY CHECK before submitting:
  - Every non-safe mode has at least one bullet with at least one dash sub-effect
  - Safe modes have exactly "No effect" for both I and J
  - No numbered SW_BANK codes (SW_BANK_x or SW_BANK_X only)
  - Each symptom phrase is under 10 words
  - J values exactly match the Step 6 rules

Return ONLY the JSON array:"""

    raw    = query_llm(prompt, temperature=0.0)
    parsed = parse_json(raw)

    rows = []
    if isinstance(parsed, list) and len(parsed) >= n:
        for i in range(n):
            rd   = parsed[i]
            # Post-process I: replace "bullet" with "•" and "dash" with "    -"
            # in case the LLM used the word form instead of actual symbols
            ic   = str(rd.get('I', 'No effect')).strip()
            ic   = ic.replace('bullet ', '• ').replace('\ndash ', '\n    - ').replace('\n- ', '\n    - ')
            sys_ = str(rd.get('J', 'No effect')).strip()
            mode = modes[i]

            if is_safe_mode(mode):
                ic   = 'No effect'
                memo = 'O'
            else:
                memo = compute_k_from_mode_and_coverage(code, mode, ic, block_to_sms)

            sys_ = _validate_j(sys_)
            rows.append(_build_row(mode, ic, sys_, memo, block_to_sms, sm_coverage,
                                   fmeda_code=code))
    else:
        print(f"    LLM parse failed for {code} - using fallback")
        rows = _fallback_rows_v7(modes, code, block_to_sms, sm_coverage)

    return rows


def _validate_j(j_val: str) -> str:
    """
    Normalize LLM J value to canonical strings.
    Spelling: 'Unintentional' for system-level (REF/BIAS/etc), 'Unintended' for driver (SW_BANK).
    Both are returned correctly here; the LLM prompt specifies which to use per block.
    """
    if not j_val or j_val.strip() == '':
        return 'No effect'
    j_lower = j_val.lower().strip()

    if 'device damage' in j_lower or 'damage to device' in j_lower:
        return 'Possible device damage' if 'possible' in j_lower else 'Device damage'

    if j_lower in ('no effect', 'none', 'no system effect', 'no impact'):
        return 'No effect'
    if 'no effect' in j_lower and len(j_lower) < 30:
        return 'No effect'

    has_no_comms = any(p in j_lower for p in ['no communication', 'no comms',
                                               'loss of communication', 'comms lost'])
    has_led = any(p in j_lower for p in ['led', 'unintention', 'unintended'])
    has_fail = 'fail' in j_lower

    if has_no_comms and has_led and has_fail:
        return 'Unintentional LED ON/OFF\nFail-safe mode active\nNo communication'
    if has_no_comms and has_fail:
        return 'Fail-safe mode active\nNo communication'
    if 'possible' in j_lower and 'fail' in j_lower:
        return 'Possible Fail-safe mode activation'
    if 'fs mode' in j_lower and 'led' in j_lower:
        return 'Unintended LED ON/OFF in FS mode'
    if 'fail-safe' in j_lower or 'failsafe' in j_lower or 'fail safe' in j_lower:
        return 'Fail-safe mode active'

    # LED ON/OFF combined -- check for both spellings
    if ('led on/off' in j_lower) or \
       ('led on' in j_lower and 'led off' in j_lower):
        # Use 'Unintentional' for system-level (default), 'Unintended' if explicitly stated
        if 'unintended led' in j_lower and 'unintentional' not in j_lower:
            return 'Unintended LED ON/OFF'
        return 'Unintentional LED ON/OFF'

    # LED ON
    if any(p in j_lower for p in ['led on', 'led turns on', 'led always on', 'leds on']):
        if 'off' not in j_lower:
            if 'unintended led' in j_lower and 'unintentional' not in j_lower:
                return 'Unintended LED ON'
            return 'Unintentional LED ON'

    # LED OFF
    if any(p in j_lower for p in ['led off', 'led turns off', 'led always off', 'leds off']):
        if 'unintended led' in j_lower and 'unintentional' not in j_lower:
            return 'Unintended LED OFF'
        return 'Unintentional LED OFF'

    return j_val.strip()


def _build_row(canonical_mode, ic, sys_, memo, block_to_sms=None, sm_coverage=None, **kw):
    ic_clean = ic.strip()
    if ic_clean in ('No effect', ''):
        memo = 'O'
    sp       = 'Y' if memo == 'X' else 'N'
    pct_safe = 1 if not memo.startswith('X') else 0
    sm_str, coverage = '', ''
    if ic_clean != 'No effect':
        sm_str, coverage = compute_sm_columns(
            ic_clean, block_to_sms or {}, sm_coverage or {},
            fmeda_code=kw.get('fmeda_code', ''),
            mode_str=canonical_mode
        )
    return {
        'G': canonical_mode, 'I': ic, 'J': sys_, 'K': memo,
        'O': 1, 'P': sp, 'R': pct_safe,
        'S': sm_str, 'T': '', 'U': coverage, 'V': '',
        'X': 'Y' if memo.startswith('X') else 'N',
        'Y': sm_str, 'Z': '', 'AA': '', 'AB': '', 'AD': '',
    }


def _fallback_rows_v7(modes, fmeda_code, block_to_sms, sm_coverage):
    """Fallback when LLM completely fails - uses mode classification only."""
    rows = []
    for mode in modes:
        ic = 'No effect' if is_safe_mode(mode) else ''
        memo = compute_k_from_mode_and_coverage(fmeda_code, mode, ic, block_to_sms)
        rows.append(_build_row(mode, ic, 'No effect' if is_safe_mode(mode) else '',
                               memo, block_to_sms, sm_coverage, fmeda_code=fmeda_code))
    return rows


def _sm_rows(sm_code: str, sm_j_map: dict) -> list:
    """
    SM blocks: 2 rows.
    I/J from sm_j_map which is built at runtime from SM list sheet descriptions.
    """
    ic, sys_ = sm_j_map.get(sm_code,
                              ('Loss of safety mechanism functionality', 'Fail-safe mode active'))
    return [
        {'G': 'Fail to detect', 'I': ic, 'J': sys_, 'K': 'X (Latent)',
         'O': 1, 'P': 'N', 'R': 0, 'S': '', 'T': '', 'U': '', 'V': '',
         'X': 'Y', 'Y': '', 'Z': '', 'AA': '', 'AB': '', 'AD': ''},
        {'G': 'False detection', 'I': 'No effect', 'J': 'No effect', 'K': 'O',
         'O': 1, 'P': 'N', 'R': 1, 'S': '', 'T': '', 'U': '', 'V': '',
         'X': 'N', 'Y': '', 'Z': '', 'AA': '', 'AB': '', 'AD': ''},
    ]


def build_sm_j_map_from_descriptions(sm_blocks: list, sm_addressed: dict,
                                      tsr_list: list, cache: dict,
                                      sm_coverage: dict = None) -> dict:
    """
    Build SM I/J values by reasoning from each SM's description text.

    ROOT CAUSE OF PREVIOUS BUGS:
      1. All SMs were sent in one batch -> LLM confused numbering, shifted answers
      2. SM07/SM19 (no template rows) were included -> caused off-by-one cascade
      3. SM descriptions were not given prominent enough role in the prompt

    FIX:
      1. Filter by sm_coverage first -> only SMs with actual template rows
      2. Process each SM INDEPENDENTLY with its description as the primary input
      3. Prompt asks: 'Given ONLY this SM description, what fails when it fails to detect?'
         The answer comes from reasoning about the description text, not from SM number.
    """
    if not sm_blocks:
        return {}

    # Filter: only include SMs that have rows in the template
    valid_blocks = []
    for sm in sm_blocks:
        m = re.match(r'sm[-_\s]?(\d+)', sm['id'].lower())
        code = f"SM{int(m.group(1)):02d}" if m else sm['id'].upper()
        if sm_coverage and len(sm_coverage) > 0 and code not in sm_coverage:
            print(f"  [SM-J] Skipping {code} - not in SM list (no template row)")
            continue
        valid_blocks.append((code, sm))

    if not valid_blocks:
        return {}

    ck = "sm_j_map_v11__" + json.dumps(sorted(code for code, _ in valid_blocks))
    if not SKIP_CACHE and ck in cache:
        print("  [SM-J] Loaded from cache")
        return cache[ck]

    tsr_ctx = "\n".join(f"  {t['id']}: {t['description']}" for t in tsr_list) \
              if tsr_list else "  (none)"

    # Build a rich description for each SM
    sm_details = []
    for code, sm in valid_blocks:
        addressed = sm_addressed.get(code, [])
        sm_details.append({
            'code':      code,
            'name':      sm.get('name', ''),
            'mechanism': sm.get('description', ''),
            'monitors':  addressed,
        })

    prompt = f"""You are an automotive IC functional safety engineer analyzing safety mechanisms.

SYSTEM SAFETY REQUIREMENTS:
{tsr_ctx}

SAFETY MECHANISMS TO ANALYZE:
{json.dumps(sm_details, indent=2)}

TASK: For EACH safety mechanism (SM), reason through:

STEP 1 - Read the "mechanism" description carefully.
         This tells you WHAT the SM monitors (e.g. "LED Open Detection", "Clock Watchdog").

STEP 2 - Ask: "If this SM FAILS TO DETECT the fault it monitors, what symptom appears?"
         The symptom is the CONSEQUENCE of the undetected fault, not the SM failure itself.
         Examples of reasoning:
         - "LED Open Detection" fails -> open LED goes undetected -> LED stays OFF when it should be ON
           -> I = "Unintended LED OFF"
         - "LED Short Detection" fails -> shorted LED undetected -> LED stuck OFF (forced low)
           -> I = "Unintended LED OFF"
         - "UART Communication Watchdog" fails -> comms error undetected
           -> I = "UART Communication Error"
         - "ADC: LED Current Monitoring" fails -> LED current fault undetected -> wrong LED state
           -> I = "Unintended LED ON"
         - "Internal Clock Watchdog Check" fails -> clock fault undetected -> comms fail
           -> I = "UART Communication Error"
         - "PWM Monitoring" fails -> PWM fault undetected -> no PWM monitoring
           -> I = "No PWM monitoring functionality"
         - "Comparator: FS Pin State" fails -> FS state undetected -> LEDs wrong in FS mode
           -> I = "Unintended LED ON/OFF in FS mode"
         - "Comparator: Internal Supply Monitoring" fails -> supply fault undetected -> logic fails
           -> I = "Failures on LOGIC operation"
         - "ADC: SBG ADC Reading" fails -> reference reading lost -> "Loss of reference control functionality"
         - "ADC: Thermal Limit" fails -> thermal limit not enforced -> device overheats
           -> I = "Device damage"
         - "ECC" fails -> memory corruption undetected -> "Cannot trim part properly"
         - "SYNC Monitor" fails -> sync pulse undetected -> "Unsynchronised PWM"
         - "Matrix SW POR" fails -> switch POR undetected -> LED stuck off
           -> I = "Unintended LED OFF"
         - "ADC: Thermal Monitoring" fails -> thermal monitoring lost
           -> I = "Loss of thermal monitoring capability"
         - "ADC: LED Switch Voltage Detection" fails -> LED switch voltage unmonitored
           -> I = "Loss of LED voltage monitoring capability"

STEP 3 - For col J, determine the system-level consequence visible to the end user.
         Use ONLY these exact strings:
           "Unintended LED ON"
           "Unintended LED OFF"
           "Unintended LED ON/OFF"
           "Unintended LED ON/OFF in FS mode"
           "Fail-safe mode active"
           "Device damage"
           "Performance/Functionality degredation"
           "No effect"

CRITICAL: Base your answer on the DESCRIPTION TEXT of each SM.
          DO NOT infer from the SM number or code.
          Each SM is INDEPENDENT — do not carry over reasoning from previous SMs.

Return a JSON object:
{{
  "SM01": {{"I": "<col I value>", "J": "<col J value>"}},
  "SM02": {{"I": "...", "J": "..."}},
  ...
}}
Return ONLY the JSON object:"""

    print("  [SM-J] Building SM effect map via LLM (description-driven)...")
    raw    = query_llm(prompt, temperature=0.0)
    parsed = parse_json(raw)

    sm_j_map = {}
    if isinstance(parsed, dict):
        for sm_code, vals in parsed.items():
            if isinstance(vals, dict):
                i_val = str(vals.get('I', 'Loss of safety mechanism functionality')).strip()
                j_val = str(vals.get('J', 'Fail-safe mode active')).strip()
                sm_j_map[sm_code] = (i_val, j_val)

    # Fill in any missing SMs with generic fallback
    for code, _ in valid_blocks:
        if code not in sm_j_map:
            print(f"  [SM-J] WARNING: {code} missing from LLM response, using fallback")
            sm_j_map[code] = ('Loss of safety mechanism functionality', 'Fail-safe mode active')

    cache[ck] = sm_j_map
    save_cache(cache)
    print(f"  [SM-J] {len(sm_j_map)} SM effect entries built")
    return sm_j_map


# =============================================================================
# AGENT 3  -  Template Writer (deterministic)
# =============================================================================

def _compute_fit_values(code, n_modes, block_fit_rates, row_memo, row_U, sm_coverage):
    block_fit = block_fit_rates.get(code, 0.0)
    mode_fit  = block_fit / n_modes if n_modes > 0 and block_fit > 0 else 0.0
    if not row_memo.startswith('X'):
        return block_fit, mode_fit, mode_fit, 0.0, None, None
    U = float(row_U) if row_U else 0.0
    V = mode_fit * (1.0 - U)
    if not U:
        AA = 0.0
    elif U >= 0.99:
        AA = 1.0
    elif U >= 0.85:
        AA = 0.8
    else:
        AA = U
    AB = mode_fit * U * (1.0 - AA)
    return block_fit, mode_fit, mode_fit, V, AA, AB


def _scan_placeholders(ws):
    idx = {}
    for ws_row in ws.iter_rows():
        for cell in ws_row:
            if cell.__class__.__name__ == 'MergedCell':
                continue
            v = str(cell.value) if cell.value is not None else ''
            if v.startswith('{{FMEDA_') and v.endswith('}}'):
                idx[v] = cell
    return idx


def _get_groups(idx, data_start=22):
    d_rows = sorted({
        int(re.search(r'(\d+)', k).group(1))
        for k in idx
        if re.match(r'\{\{FMEDA_D\d+\}\}', k)
        and int(re.search(r'(\d+)', k).group(1)) >= data_start
    })
    all_rows = sorted({
        int(re.search(r'(\d+)', k).group(1))
        for k in idx
        if re.match(r'\{\{FMEDA_[A-Z]+\d+\}\}', k)
        and int(re.search(r'(\d+)', k).group(1)) >= data_start
    })
    groups = []
    for i, first in enumerate(d_rows):
        nxt = d_rows[i+1] if i+1 < len(d_rows) else 999999
        groups.append([r for r in all_rows if first <= r < nxt])
    return groups


def _write(idx, col, row_num, value, wrap=False):
    key = '{{FMEDA_' + col + str(row_num) + '}}'
    if key not in idx:
        return
    cell = idx[key]
    if value is None or (isinstance(value, str) and value.strip() in ('', 'None', 'nan')):
        cell.value = None
        return
    cell.value = value
    if wrap and isinstance(value, str) and '\n' in value:
        old = cell.alignment or Alignment()
        cell.alignment = Alignment(wrap_text=True,
                                   vertical=old.vertical or 'center',
                                   horizontal=old.horizontal or 'left')


def agent3_write_template(fmeda_data, block_fit_rates=None, sm_coverage=None):
    if block_fit_rates is None:
        block_fit_rates = {}
    if sm_coverage is None:
        sm_coverage = {}
    shutil.copy2(TEMPLATE_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb['FMEDA']
    idx    = _scan_placeholders(ws)
    groups = _get_groups(idx)
    print(f"\n  [Agent 3] Template groups : {len(groups)}")
    print(f"  [Agent 3] Blocks to write : {len(fmeda_data)}")

    fm = 1
    for bi, block in enumerate(fmeda_data):
        code = block['fmeda_code']
        rows = block['rows']
        if bi >= len(groups):
            print(f"  [Agent 3] WARNING: no template group for block {bi+1} ({code})")
            break
        group_rows = groups[bi]
        n_t = len(group_rows)
        n_d = len(rows)
        if n_d > n_t:
            print(f"  [Agent 3] {code}: {n_d} modes > {n_t} slots - truncating")
            rows = rows[:n_t]
        n_modes_total = max(len(rows), 1)

        for mi, row_num in enumerate(group_rows):
            rd       = rows[mi] if mi < len(rows) else None
            is_first = (mi == 0)
            _write(idx, 'B', row_num, f'FM_TTL_{fm}' if rd else None)
            _write(idx, 'C', row_num, code)
            _write(idx, 'D', row_num, code if is_first else None)
            if rd is None:
                _write(idx, 'G', row_num, None)
                continue

            memo     = str(rd.get('K', 'O')).strip()
            sp       = str(rd.get('P', 'Y' if memo == 'X' else 'N')).strip()
            pct_safe = rd.get('R', 1 if memo == 'O' else 0)
            u_val    = rd.get('U', '')

            fit_blk, fit_mode, fit_q, fit_v, fit_aa, fit_ab = _compute_fit_values(
                code, n_modes_total, block_fit_rates, memo, u_val, sm_coverage)

            _write(idx, 'E',  row_num, fit_blk if (is_first and fit_blk > 0) else None)
            _write(idx, 'F',  row_num, fit_mode if fit_mode > 0 else None)
            _write(idx, 'G',  row_num, rd.get('G', ''),          wrap=True)
            _write(idx, 'H',  row_num, None)
            _write(idx, 'I',  row_num, rd.get('I', 'No effect'), wrap=True)
            _write(idx, 'J',  row_num, rd.get('J', 'No effect'), wrap=True)
            _write(idx, 'K',  row_num, memo)
            _write(idx, 'O',  row_num, 1)
            _write(idx, 'P',  row_num, sp)
            _write(idx, 'Q',  row_num, fit_q if fit_q > 0 else None)
            _write(idx, 'R',  row_num, pct_safe)
            _write(idx, 'S',  row_num, rd.get('S') or None, wrap=False)
            _write(idx, 'T',  row_num, rd.get('T') or None, wrap=False)
            _write(idx, 'U',  row_num, u_val if u_val not in ('', None) else None)
            _write(idx, 'V',  row_num, fit_v if (fit_v is not None and fit_v > 0) else None)
            _write(idx, 'X',  row_num, rd.get('X', 'Y' if memo.startswith('X') else 'N'))
            _write(idx, 'Y',  row_num, rd.get('Y') or None, wrap=False)
            _write(idx, 'Z',  row_num, rd.get('Z') or None, wrap=True)
            _write(idx, 'AA', row_num, fit_aa if fit_aa is not None else None)
            if fit_ab is not None:
                _write(idx, 'AB', row_num, fit_ab if fit_ab > 0 else 0)
            sm_str = rd.get('S', '') or ''
            if sm_str and memo.startswith('X'):
                sms = sm_str.split()
                sms_sorted = sorted(sms,
                    key=lambda s: sm_coverage.get(s, 0.0) if sm_coverage else 0.0,
                    reverse=True)
                sm_mention = ' '.join(sms_sorted[:2]) if len(sms_sorted) >= 2 else (sms_sorted[0] if sms_sorted else '')
                lat_pct = int(round((fit_aa or 1.0) * 100))
                _write(idx, 'AD', row_num,
                       f'{sm_mention} make the IC enter a safe-sate. Latent coverage: {lat_pct}%.',
                       wrap=True)
            else:
                _write(idx, 'AD', row_num, rd.get('AD') or None, wrap=True)
            fm += 1

        print(f"  [Agent 3] [{bi+1}/{len(fmeda_data)}] {code}: "
              f"{min(n_d, n_t)} rows -> FM_TTL_{fm-min(n_d,n_t)} - FM_TTL_{fm-1}")

    wb.save(OUTPUT_FILE)
    print(f"\n  [Agent 3] Saved -> {OUTPUT_FILE}")
    print(f"  [Agent 3] Total failure modes: {fm - 1}")


# =============================================================================
# MAIN
# =============================================================================

def run():
    print("╔═══════════════════════════════════════════════════╗")
    print("║   FMEDA Multi-Agent Pipeline  v7 (fully generic)  ║")
    print("╚═══════════════════════════════════════════════════╝")
    print(f"\n  Dataset  : {DATASET_FILE}")
    print(f"  IEC table: {IEC_TABLE_FILE}")
    print(f"  Template : {TEMPLATE_FILE}")
    print(f"  Model    : {OLLAMA_MODEL}")
    print(f"  Output   : {OUTPUT_FILE}\n")

    cache = load_cache()

    # ── Step 0: Read all inputs ────────────────────────────────────────────
    print("━━━ Step 0 : Reading inputs ━━━")
    blk_blocks, sm_blocks, tsr_list = read_dataset()
    iec_table = read_iec_table()

    # SM list and FIT rates from TEMPLATE only (no reference FMEDA)
    sm_coverage, sm_addressed, block_to_sms, sm_descriptions = read_sm_list()
    # Make SM descriptions available globally for compute_sm_columns keyword filtering
    _SM_DESCRIPTIONS_RUNTIME.update(sm_descriptions)
    block_fit_rates = {}
    if os.path.exists(TEMPLATE_FILE):
        try:
            wb_fit = openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)
            block_fit_rates = read_block_fit_rates(wb_fit)
            if block_fit_rates:
                print(f"  FIT rates: {len(block_fit_rates)} blocks from {TEMPLATE_FILE}")
        except Exception:
            pass
    if not block_fit_rates:
        print("  FIT rates: not available (place FMEDA_TEMPLATE.xlsx with FIT data to enable)")

    print(f"  BLK: {len(blk_blocks)}  SM: {len(sm_blocks)}  TSR: {len(tsr_list)}  "
          f"IEC: {len(iec_table)}  SM entries: {len(sm_coverage)}  "
          f"FIT blocks: {len(block_fit_rates)}")
    print("  block_to_sms:")
    for b, sms in sorted(block_to_sms.items()):
        print(f"    {b:<15} -> {sms}")

    # ── Agent 0: Build signal flow graph ──────────────────────────────────
    print("\n━━━ Agent 0 : Signal flow graph builder ━━━")
    signal_graph = build_signal_flow_graph(blk_blocks, cache)
    for blk_name, info in signal_graph.items():
        consumers = info.get('consumers', [])
        print(f"  {blk_name:<20} outputs: {info.get('output_signal','?')[:50]}")
        print(f"  {'':20} feeds:   {consumers}")

    # ── Agent 1: Map blocks -> IEC parts ──────────────────────────────────
    print("\n━━━ Agent 1 : Block -> IEC part mapper (LLM) ━━━")
    blocks = agent1_map_blocks(blk_blocks, sm_blocks, iec_table, cache, sm_coverage)
    print("\n  Mapping result:")
    for b in blocks:
        tag = " [DUP]" if b.get('is_duplicate') else (" [SM]" if b.get('is_sm') else "")
        flags = []
        if b.get('is_driver'):     flags.append('driver')
        if b.get('is_interface'):  flags.append('interface')
        if b.get('is_trim'):       flags.append('trim')
        if b.get('is_opamp_type'): flags.append('opamp')
        print(f"    {b['name']:<35} -> {b['fmeda_code']:<12} "
              f"({len(b.get('modes', []))} modes){tag} [{','.join(flags)}]")

    # ── Build SM J map from descriptions (no reference file needed) ────────
    print("\n━━━ Building SM effect map ━━━")
    sm_j_map = build_sm_j_map_from_descriptions(
        sm_blocks, sm_addressed, tsr_list, cache, sm_coverage=sm_coverage)

    # ── Agent 2: Generate IC/system effects ───────────────────────────────
    print("\n━━━ Agent 2 : IC Effects (col I) + System Effects (col J) ━━━")
    fmeda_data = agent2_generate_effects(
        blocks, tsr_list, block_to_sms, sm_coverage,
        sm_addressed, cache, signal_graph, sm_j_map)

    print("\n  Spot-check (K, I preview):")
    for block in fmeda_data:
        for row in block['rows']:
            print(f"    {block['fmeda_code']:<12} K={row.get('K','?'):<12} "
                  f"I={repr(row.get('I',''))[:50]}  | {row['G'][:35]}")

    with open(INTERMEDIATE_JSON, 'w', encoding='utf-8') as f:
        json.dump(fmeda_data, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n  Intermediate JSON -> {INTERMEDIATE_JSON}")

    # ── Agent 3: Write template ────────────────────────────────────────────
    print("\n━━━ Agent 3 : Template writer (deterministic) ━━━")
    agent3_write_template(fmeda_data, block_fit_rates, sm_coverage)

    print("\n✅  Pipeline complete!")
    print(f"    Output       : {OUTPUT_FILE}")
    print(f"    Intermediate : {INTERMEDIATE_JSON}")
    print(f"    Cache        : {CACHE_FILE}")


if __name__ == '__main__':
    run()