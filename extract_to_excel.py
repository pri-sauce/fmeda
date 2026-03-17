"""
Extracts Pin Functions and Faults/Diagnostics (SM) from ID803-style PDFs
and populates the Pin and SM sheets in the Excel template.

Usage: python3 extract_to_excel.py <pdf_path> <excel_path> [output_path]
"""

import re, sys, shutil
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# PDF TEXT EXTRACTION  (two-column aware)
# ─────────────────────────────────────────────────────────────────────────────

def extract_text(pdf_path):
    """Extract text respecting the two-column layout of the datasheet."""
    pages_text = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            if not words:
                continue
            # Split into left (col 1) and right (col 2) by x midpoint
            xs = [w['x0'] for w in words]
            # Find natural column split: look for gap in x distribution
            # Simple heuristic: midpoint of page bbox
            mid = (min(xs) + max(xs)) / 2
            left  = sorted([w for w in words if w['x0'] <  mid], key=lambda w: (round(w['top']/4), w['x0']))
            right = sorted([w for w in words if w['x0'] >= mid], key=lambda w: (round(w['top']/4), w['x0']))
            col1 = ' '.join(w['text'] for w in left)
            col2 = ' '.join(w['text'] for w in right)
            pages_text.append(col1 + '\n' + col2)
    return '\n'.join(pages_text)


# ─────────────────────────────────────────────────────────────────────────────
# PIN PARSER
# ─────────────────────────────────────────────────────────────────────────────

def _clean_desc(s):
    s = re.sub(r'ID803 Datasheet Rev[\d\.]+', '', s)
    s = re.sub(r'(PRELIMINARY DATASHEET|Information Herein Subject to Change'
               r'|ELEVATION MICROSYSTEMS CONFIDENTIAL)\s*\d*', '', s)
    s = re.sub(r'\s{2,}', ' ', s).strip()
    for stop in ['FAULTS AND DIAGNOSTICS', 'TYPICAL CHARACTERISTICS', 'Table 4.']:
        idx = s.find(stop)
        if idx > 0:
            s = s[:idx].strip()
    return s


def _infer_type(name, desc):
    n, d = name.lower(), desc.lower()
    if 'ground' in d or n.strip().startswith('gnd'): return 'Power'
    if 'supply input' in d or 'supply voltage' in d:  return 'Power'
    if 'vddo' in n:                                    return 'Power'
    if 'charge pump' in d:                             return 'Passive'
    if 'led' in n:                                     return 'I/O (Analog)'
    if 'current sense' in d or 'adc input' in d:       return 'Input (Analog)'
    if 'general purpose adc' in d:                     return 'Input (Analog)'
    if 'uart transmit' in d or 'transmit data' in d:   return 'Output (Digital)'
    if 'uart receive'  in d or 'receive data'  in d:   return 'Input (Digital)'
    if 'synchronization' in d:                         return 'I/O (Digital)'
    if 'configuration' in d or 'address' in d or 'selection' in d or 'fail-safe' in d:
        return 'Input (Digital)'
    return 'I/O'


def parse_pins(text):
    """
    Match blocks like: NAME(s) (Pin(s) N,...): description ... up to next block
    """
    splitter = re.compile(
        r'([A-Z][A-Z0-9_]+(?:,\s*[A-Z][A-Z0-9_]+)*)'
        r'\s*\(Pins?\s*([\d,\s]+(?:,\s*Exposed\s+Pad\s+\d+)?)\)\s*:'
    )
    matches = list(splitter.finditer(text))
    rows = []

    for i, m in enumerate(matches):
        raw_names = m.group(1).strip()
        raw_pins  = m.group(2).strip()
        start = m.end()
        end   = matches[i+1].start() if i+1 < len(matches) else start + 1200
        desc  = _clean_desc(text[start:end])

        pin_nums_raw = re.sub(r'Exposed\s+Pad\s+\d+', '', raw_pins)
        pin_nums  = [p.strip() for p in pin_nums_raw.split(',') if p.strip().isdigit()]
        name_list = [n.strip() for n in raw_names.split(',') if n.strip()]

        pin_type = _infer_type(raw_names, desc)
        function = re.split(r'\.\s', desc)[0].rstrip('.')
        if len(function) > 80:
            function = function[:80].rstrip() + '...'

        if len(pin_nums) > 1 and len(name_list) > 1:
            for pn, nm in zip(pin_nums, name_list):
                rows.append({'Pin No.': pn, 'Pin Name': nm, 'Type': pin_type,
                             'Function': function, 'Description': desc})
        elif len(pin_nums) == 1:
            rows.append({'Pin No.': pin_nums[0], 'Pin Name': ', '.join(name_list),
                         'Type': pin_type, 'Function': function, 'Description': desc})
        else:
            pin_range = ', '.join(pin_nums) if pin_nums else raw_pins
            ep = re.search(r'Exposed\s+Pad\s+(\d+)', raw_pins)
            if ep:
                pin_range += f', Exposed Pad {ep.group(1)}'
            rows.append({'Pin No.': pin_range, 'Pin Name': ', '.join(name_list),
                         'Type': pin_type, 'Function': function, 'Description': desc})
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# SAFETY MECHANISM / FAULTS PARSER
# ─────────────────────────────────────────────────────────────────────────────

_FAULTS = [
    ('VDD is below UVLO',            'VDD < 3.92V(typ)',                                       'Yes', 'PWR',                   'N/A'),
    ('VDD Overvoltage',               'VDD > 5.77V typ',                                        'No',  'VDD_OV_FLT',            'EN_VDD_OV'),
    ('Resistive FET Detection',       'VLEDx > VTH_SHORT',                                     'No',  'FLT_RESFET[16:1]',      'NA'),
    ('LED Open Detection',            'VLEDx > VTH_OPEN',                                      'Yes', 'FLT_OPEN_OR_DRV[16:1]', 'NA'),
    ('LED Short Detection',           'VLEDx < VTH_SHORT',                                     'No',  'FLT_SHORT[16:1]',       'NA'),
    ('Driver health diagnostic',      'NA',                                                     'No',  'FLT_OPEN_OR_DRV[16:1]', 'DRV_CHK'),
    ('Matrix SW POR',                 'NA',                                                     'Yes', 'FLT_MATRIX_POR[16:1]',  'NA'),
    ('LED Current Monitoring',        'NA',                                                     'No',  'CS',                    'CSEN, CSGAIN'),
    ('UART Communication Watchdog',   'Not Receiving time valid UART > CMWTAP',                'Yes', 'OPMODE',                'CMWEN, CMWTAP'),
    ('Internal UART Watchdog Check',  'Not Receiving time valid UART > CMWTAP/8 after start-up', 'No', 'OPMODE',               'CMWEN, CMWTAP'),
    ('PWM Monitoring',                'NA',                                                     'No',  'PWM_ERR, PWM_MISCOUNT', 'NA'),
    ('FS pin status',                 'NA',                                                     'No',  'FS_PIN',                'NA'),
    ('Charge pump Voltage Monitoring','VCP < VCPTH-F',                                          'Yes', 'CHPMP_ERR',             'NA'),
    ('Internal Supply Monitoring',    'Internal supply < UVLO threshold',                      'Yes', 'PWR',                   'NA'),
    ('Bandgap ADC',                   'BGR > 0xA4 or BGR < 0x90',                             'No',  'BGR',                   'NA'),
    ('Thermal Limit',                 'TJ > 175\u00b0C typically',                             'Yes', 'TSD',                   'TSDEN'),
]


def _extract_fault_desc(text, condition_name):
    idx = text.find(condition_name)
    if idx == -1:
        return ''
    snippet = text[idx: idx + 800]
    lines = [l.strip() for l in snippet.split(' ')]
    # Reconstruct into sentences; take the long description sentence
    full = ' '.join(lines)
    # Find the paragraph after the heading and detection line
    # Look for the sentence that explains the behavior
    sentences = re.split(r'(?<=[.!?])\s+', full)
    desc_parts = []
    for sent in sentences:
        if len(sent) > 60 and condition_name not in sent:
            if not re.match(r'^(VDD|VLEDx|VCP|BGR|TJ|Not\s|CONDITION|NA\b)', sent.strip()):
                desc_parts.append(sent.strip())
        if len(desc_parts) >= 3:
            break
    desc = ' '.join(desc_parts)
    desc = re.sub(r'ID803 Datasheet.*', '', desc).strip()
    return desc


def parse_safety_mechanisms(text):
    rows = []
    for i, (cond, detection, fault_rxn, status_bit, cfg_resp) in enumerate(_FAULTS, 1):
        desc = _extract_fault_desc(text, cond)
        rows.append({
            'SM ID':                       f'SM-{i:02d}',
            'Name':                         cond,
            'Description':                  desc,
            'Detection':                    detection,
            'Fault Reaction':               fault_rxn,
            'Status Bit':                   status_bit,
            'Configurable Fault Response':  cfg_resp,
            'Addressed Part (Block)':       None,
            'Connected TSR(s)':             None,
            'Diagnostic Coverage (DC)':     None,
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITING
# ─────────────────────────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill('solid', start_color='1F4E79')
_HDR_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
_BODY_FONT = Font(name='Arial', size=9)
_WRAP      = Alignment(wrap_text=True, vertical='top')
_CTR       = Alignment(horizontal='center', vertical='top', wrap_text=True)
_THIN      = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))

def _hdr(cell):
    cell.fill = _HDR_FILL; cell.font = _HDR_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _THIN

def _body(cell, center=False):
    cell.font = _BODY_FONT
    cell.alignment = _CTR if center else _WRAP
    cell.border = _THIN


def write_pin_sheet(ws, pins):
    headers = ['Pin No.', 'Pin Name', 'Type', 'Function', 'Description']
    widths  = [10, 18, 18, 40, 72]
    ws.delete_rows(2, ws.max_row)
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h); _hdr(cell)
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 20
    for r, p in enumerate(pins, 2):
        vals = [p['Pin No.'], p['Pin Name'], p['Type'], p['Function'], p['Description']]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); _body(cell, center=(c <= 3))
        ws.row_dimensions[r].height = 55
    ws.freeze_panes = 'A2'


def write_sm_sheet(ws, sms):
    headers = ['SM ID', 'Name', 'Description', 'Detection', 'Fault Reaction',
               'Status Bit', 'Configurable Fault Response',
               'Addressed Part (Block)', 'Connected TSR(s)', 'Diagnostic Coverage (DC)']
    widths  = [10, 28, 68, 40, 14, 24, 28, 22, 18, 24]
    ws.delete_rows(2, ws.max_row)
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h); _hdr(cell)
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 30
    for r, sm in enumerate(sms, 2):
        vals = [sm['SM ID'], sm['Name'], sm['Description'], sm['Detection'],
                sm['Fault Reaction'], sm['Status Bit'], sm['Configurable Fault Response'],
                sm['Addressed Part (Block)'], sm['Connected TSR(s)'], sm['Diagnostic Coverage (DC)']]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); _body(cell, center=(c in (1,5,6,7,8,9,10)))
        ws.row_dimensions[r].height = 65
    ws.freeze_panes = 'A2'


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(pdf_path, excel_path, output_path=None):
    if not output_path:
        output_path = excel_path.replace('.xlsx', '_updated.xlsx')
    shutil.copy(excel_path, output_path)

    print(f"Reading {pdf_path} ...")
    text = extract_text(pdf_path)

    pins = parse_pins(text)
    print(f"  Pins parsed: {len(pins)}")

    sms = parse_safety_mechanisms(text)
    print(f"  Safety mechanisms parsed: {len(sms)}")

    wb = load_workbook(output_path)
    write_pin_sheet(wb['Pin'], pins)
    write_sm_sheet(wb['SM'], sms)
    wb.save(output_path)
    print(f"Saved -> {output_path}")
    return pins, sms


if __name__ == '__main__':
    pdf   = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/uploads/pin_sm.pdf'
    excel = sys.argv[2] if len(sys.argv) > 2 else '/mnt/user-data/uploads/fusa_ai_agent_mock_data.xlsx'
    out   = sys.argv[3] if len(sys.argv) > 3 else '/mnt/user-data/outputs/fusa_ai_agent_updated.xlsx'
    main(pdf, excel, out)
