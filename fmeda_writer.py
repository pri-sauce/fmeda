"""
fmeda_writer.py

Fills FMEDA_TEMPLATE.xlsx with data from the JSON pipeline output.
Writes ONLY the FMEDA sheet. All other sheets and formulas are preserved.

Flow:
  llm_output_with_effects.json → fmeda_writer.py → FMEDA_filled.xlsx

The template uses {{FMEDA_B22}} style placeholders for header cells (rows 1-21).
Data rows start at row 22 and follow the same column layout as the real FMEDA.

Column mapping (from real FMEDA analysis):
  B = Failure Mode Number
  C = Component Name (same as block name)
  D = Block Name (only on first row of each block)
  E = Block Failure rate [FIT] (only on first row, formula-driven)
  F = Failure mode separate fault rate (formula)
  G = Standard failure mode
  H = Failure Mode (same as G for now)
  I = effects on the IC output
  J = effects on the system
  K = memo (X or O)
  O = Failure distribution (always 1)
  P = Single Point Failure mode (Y/N)
  Q = Failure rate [FIT] (formula)
  R = Percentage of Safe Faults (0 or 1)
  S = Safety mechanism(s) IC
  T = Safety mechanism(s) System
  U = Failure mode coverage SPF
  V = Residual/SPF failure rate (formula)
  X = Latent failure mode (Y/N)
  Y = Safety mechanism(s) IC latent
  Z = Safety mechanism(s) System latent
  AA = Failure mode coverage latent
  AB = Latent Multiple Point Fault rate (formula)
  AD = Comment
"""

import json
import shutil
import openpyxl
from openpyxl.styles import Alignment
from copy import copy

# ─── CONFIG ──────────────────────────────────────────────────────────────────
TEMPLATE_FILE  = 'FMEDA_TEMPLATE.xlsx'
JSON_INPUT     = 'llm_output_with_effects.json'
OUTPUT_FILE    = 'FMEDA_filled.xlsx'
DATA_START_ROW = 22   # first data row in template
# ─────────────────────────────────────────────────────────────────────────────


# ─── COLUMN MAP ──────────────────────────────────────────────────────────────
# Maps field name → Excel column letter
COL = {
    'failure_mode_number': 'B',
    'component_name':      'C',
    'block_name':          'D',   # only first row per block
    'block_fit':           'E',   # only first row per block
    'mode_fit':            'F',
    'standard_failure_mode': 'G',
    'failure_mode':        'H',   # same as G
    'ic_output_effect':    'I',
    'system_effect':       'J',
    'memo':                'K',   # X or O
    'system_comment':      'L',
    'system_request':      'M',
    'failure_distribution':'O',   # always 1
    'single_point':        'P',   # Y or N
    'failure_rate_fit':    'Q',
    'pct_safe_faults':     'R',
    'sm_ic':               'S',
    'sm_system':           'T',
    'coverage_spf':        'U',
    'residual_fit':        'V',
    'latent_yesno':        'X',
    'sm_ic_latent':        'Y',
    'sm_system_latent':    'Z',
    'coverage_latent':     'AA',
    'latent_fit':          'AB',
    'comment':             'AD',
}

def col_letter(letter):
    """Convert column letter to 1-based index."""
    result = 0
    for c in letter.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result


# ─── STYLE HELPERS ───────────────────────────────────────────────────────────

def copy_style(src_cell, dst_cell):
    """Copy cell style from source to destination."""
    if src_cell.has_style:
        dst_cell.font      = copy(src_cell.font)
        dst_cell.fill      = copy(src_cell.fill)
        dst_cell.border    = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def set_cell(ws, row, col_letter_str, value, wrap=False, src_style_cell=None):
    """Set a cell value and optionally copy style."""
    cell = ws.cell(row=row, column=col_letter(col_letter_str))
    cell.value = value
    if src_style_cell:
        copy_style(src_style_cell, cell)
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    elif cell.alignment:
        cell.alignment = Alignment(
            wrap_text=cell.alignment.wrap_text,
            vertical=cell.alignment.vertical or 'top',
            horizontal=cell.alignment.horizontal
        )


# ─── TEMPLATE FILLER ─────────────────────────────────────────────────────────

def fill_template(template_path, json_path, output_path):
    # Load JSON
    with open(json_path, 'r', encoding='utf-8-sig') as f:
        data = json.load(f)

    # Copy template to output
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb['FMEDA']

    # Clear placeholder values in header rows (rows 1-21) — keep {{}} cells as-is
    # (they'll be populated by the application layer, not us)

    # Get style from template row 22 to use for new rows
    template_row_22 = {
        col: ws.cell(row=22, column=col_letter(col))
        for col in ['B','C','D','E','F','G','H','I','J','K','O','P','Q','R','S','T','U','V','X','Y','Z','AA','AB','AD']
    }

    # Clear existing data rows (row 22 onwards) — skip merged cells
    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.__class__.__name__ != 'MergedCell':
                cell.value = None

    # Write data
    current_row = DATA_START_ROW
    fm_counter  = 1  # global failure mode number counter

    for block in data:
        block_name = block['block_name']
        function   = block.get('function', '')
        rows       = block.get('rows', [])

        if not rows:
            continue

        for idx, row in enumerate(rows):
            is_first = (idx == 0)
            mode = row.get('Standard failure mode', '')
            effect_ic = row.get('effects on the IC output', 'No effect')
            effect_sys = row.get('effects on the system', '')
            memo = row.get('memo', '')
            fd = row.get('Failure distribution', 1)
            sp = row.get('Single Point Failure mode', '')
            fit = row.get('Failure rate [FIT]', '')
            pct_safe = row.get('Percentage of Safe Faults', '')
            sm_ic = row.get('Safety mechanism(s) (IC) allowing to prevent the violation of the safety goal', '')
            sm_sys = row.get('Safety mechanism(s) (System) allowing to prevent the violation of the safety goal', '')
            cov_spf = row.get('Failure mode coverage wrt. violation of safety goal', '')
            res_fit = row.get('Residual or Single Point Fault failure rate [FIT]', '')
            latent = row.get('Latent Failure mode', '')
            sm_ic_lat = row.get('Safety mechanism(s) (IC) to prevent latent faults', '')
            sm_sys_lat = row.get('Safety mechanism(s) (System) to prevent latent faults', '')
            cov_lat = row.get('Failure mode coverage wrt. Latent failures', '')
            lat_fit = row.get('Latent Multiple Point Fault failure rate [FIT]', '')
            comment = row.get('comment', '')
            block_fit = row.get('Block Failure rate [FIT]', '')

            # Get style reference from template row 22
            style_src = template_row_22

            # B: Failure Mode Number
            set_cell(ws, current_row, 'B', f'FM_TTL_{fm_counter}',
                     src_style_cell=style_src.get('B'))

            # C: Component Name (same as block name)
            set_cell(ws, current_row, 'C', block_name,
                     src_style_cell=style_src.get('C'))

            # D: Block Name — only on first row
            if is_first:
                set_cell(ws, current_row, 'D', block_name,
                         src_style_cell=style_src.get('D'))

            # E: Block FIT — only on first row
            if is_first and block_fit:
                set_cell(ws, current_row, 'E', block_fit,
                         src_style_cell=style_src.get('E'))

            # F: Mode FIT (leave empty — formula driven in real file)
            if fit:
                set_cell(ws, current_row, 'F', fit,
                         src_style_cell=style_src.get('F'))

            # G: Standard failure mode
            set_cell(ws, current_row, 'G', mode, wrap=True,
                     src_style_cell=style_src.get('G'))

            # H: Failure Mode (same as G)
            set_cell(ws, current_row, 'H', mode, wrap=True,
                     src_style_cell=style_src.get('H'))

            # I: Effects on IC output
            set_cell(ws, current_row, 'I', effect_ic, wrap=True,
                     src_style_cell=style_src.get('I'))

            # J: Effects on system
            set_cell(ws, current_row, 'J', effect_sys, wrap=True,
                     src_style_cell=style_src.get('J'))

            # K: Memo (X or O)
            set_cell(ws, current_row, 'K', memo,
                     src_style_cell=style_src.get('K'))

            # O: Failure distribution
            set_cell(ws, current_row, 'O', fd if fd != '' else 1,
                     src_style_cell=style_src.get('O'))

            # P: Single Point
            if sp:
                set_cell(ws, current_row, 'P', sp,
                         src_style_cell=style_src.get('P'))

            # Q: Failure rate
            if fit:
                set_cell(ws, current_row, 'Q', fit,
                         src_style_cell=style_src.get('Q'))

            # R: Pct safe faults
            if pct_safe != '':
                set_cell(ws, current_row, 'R', pct_safe,
                         src_style_cell=style_src.get('R'))

            # S: SM IC
            if sm_ic:
                set_cell(ws, current_row, 'S', sm_ic, wrap=True,
                         src_style_cell=style_src.get('S'))

            # T: SM System
            if sm_sys:
                set_cell(ws, current_row, 'T', sm_sys, wrap=True,
                         src_style_cell=style_src.get('T'))

            # U: Coverage SPF
            if cov_spf != '':
                set_cell(ws, current_row, 'U', cov_spf,
                         src_style_cell=style_src.get('U'))

            # V: Residual FIT
            if res_fit != '':
                set_cell(ws, current_row, 'V', res_fit,
                         src_style_cell=style_src.get('V'))

            # X: Latent Y/N
            if latent:
                set_cell(ws, current_row, 'X', latent,
                         src_style_cell=style_src.get('X'))

            # Y: SM IC latent
            if sm_ic_lat:
                set_cell(ws, current_row, 'Y', sm_ic_lat, wrap=True,
                         src_style_cell=style_src.get('Y'))

            # Z: SM System latent
            if sm_sys_lat:
                set_cell(ws, current_row, 'Z', sm_sys_lat, wrap=True,
                         src_style_cell=style_src.get('Z'))

            # AA: Coverage latent
            if cov_lat != '':
                set_cell(ws, current_row, 'AA', cov_lat,
                         src_style_cell=style_src.get('AA'))

            # AB: Latent FIT
            if lat_fit != '':
                set_cell(ws, current_row, 'AB', lat_fit,
                         src_style_cell=style_src.get('AB'))

            # AD: Comment
            if comment:
                set_cell(ws, current_row, 'AD', comment, wrap=True,
                         src_style_cell=style_src.get('AD'))

            current_row += 1
            fm_counter  += 1

    total_rows = current_row - DATA_START_ROW
    print(f"  Written {total_rows} rows across {len(data)} blocks")
    print(f"  Last row used: {current_row - 1}")

    wb.save(output_path)
    print(f"  Saved to {output_path}")


if __name__ == '__main__':
    print(f"=== Filling FMEDA Template ===")
    print(f"  Template : {TEMPLATE_FILE}")
    print(f"  Data     : {JSON_INPUT}")
    print(f"  Output   : {OUTPUT_FILE}")
    print()

    fill_template(TEMPLATE_FILE, JSON_INPUT, OUTPUT_FILE)
    print("\n✅ Done")
