import pdfplumber
import requests
import json
from openpyxl import load_workbook

# =========================
# CONFIG (HARDCODED)
# =========================

MODEL_NAME = "qwen2.5:7b-instruct-q4_K_M"
OLLAMA_URL = "http://localhost:11434/api/generate"

PDF_FILE = "datasheet.pdf"
TEMPLATE_FILE = "fusa_ai_agent_mock_data.xlsx"
OUTPUT_FILE = "generated_fusa.xlsx"

CHUNK_SIZE = 4000


# =========================
# SYSTEM PROMPT
# =========================

SYSTEM_PROMPT = """
You are an automotive functional safety engineer.

Your task is to extract structured information from semiconductor datasheets
and convert it into ISO 26262 safety model components.

Return ONLY JSON.

Schema:

{
 "chip_info":{
   "name":"",
   "function":"",
   "asil_level":"",
   "application":""
 },

 "blocks":[
   {
    "block_name":"",
    "description":""
   }
 ],

 "pins":[
   {
     "pin":"",
     "name":"",
     "description":""
   }
 ],

 "safety_mechanisms":[
   {
     "mechanism":"",
     "description":""
   }
 ]
}
"""


# =========================
# PDF TEXT EXTRACTION
# =========================

def extract_pdf_text(path):

    text = ""

    with pdfplumber.open(path) as pdf:

        for page in pdf.pages:

            t = page.extract_text()

            if t:
                text += t + "\n"

    return text


# =========================
# TEXT CHUNKING
# =========================

def chunk_text(text):

    chunks = []

    for i in range(0, len(text), CHUNK_SIZE):
        chunks.append(text[i:i + CHUNK_SIZE])

    return chunks


# =========================
# CALL OLLAMA
# =========================

def call_ollama(prompt):

    payload = {
        "model": MODEL_NAME,
        "prompt": prompt,
        "stream": False,
        "temperature": 0
    }

    r = requests.post(OLLAMA_URL, json=payload)

    response = r.json()["response"]

    return response


# =========================
# RUN EXTRACTION
# =========================

def extract_with_llm(text):

    chunks = chunk_text(text)

    merged = {
        "blocks": [],
        "pins": [],
        "safety_mechanisms": []
    }

    chip_info = None

    for c in chunks:

        prompt = SYSTEM_PROMPT + "\n\nDATASHEET TEXT:\n" + c

        result = call_ollama(prompt)

        try:
            data = json.loads(result)

            if not chip_info and "chip_info" in data:
                chip_info = data["chip_info"]

            merged["blocks"].extend(data.get("blocks", []))
            merged["pins"].extend(data.get("pins", []))
            merged["safety_mechanisms"].extend(data.get("safety_mechanisms", []))

        except:
            pass

    return chip_info, merged


# =========================
# WRITE EXCEL
# =========================

def write_excel(info, merged):

    wb = load_workbook(TEMPLATE_FILE)

    # INFO SHEET
    sheet = wb["Info"]

    sheet["A2"] = "Chip Name"
    sheet["B2"] = info.get("name","")

    sheet["A3"] = "Function"
    sheet["B3"] = info.get("function","")

    sheet["A4"] = "ASIL"
    sheet["B4"] = info.get("asil_level","")

    sheet["A5"] = "Application"
    sheet["B5"] = info.get("application","")


    # BLOCK SHEET
    sheet = wb["BLK"]

    row = 2

    for b in merged["blocks"]:

        sheet.cell(row=row, column=1).value = row-1
        sheet.cell(row=row, column=2).value = b["block_name"]
        sheet.cell(row=row, column=3).value = b["description"]

        row += 1


    # PIN SHEET
    sheet = wb["Pin"]

    row = 2

    for p in merged["pins"]:

        sheet.cell(row=row, column=1).value = p["pin"]
        sheet.cell(row=row, column=2).value = p["name"]
        sheet.cell(row=row, column=5).value = p["description"]

        row += 1


    # SAFETY MECHANISMS
    sheet = wb["SM"]

    row = 2

    for s in merged["safety_mechanisms"]:

        sheet.cell(row=row, column=1).value = s["mechanism"]
        sheet.cell(row=row, column=2).value = s["description"]

        row += 1


    wb.save(OUTPUT_FILE)


# =========================
# MAIN
# =========================

def main():

    print("Reading PDF...")
    text = extract_pdf_text(PDF_FILE)

    print("Running LLM extraction...")
    info, merged = extract_with_llm(text)

    print("Writing Excel...")
    write_excel(info, merged)

    print("Done →", OUTPUT_FILE)


if __name__ == "__main__":
    main()