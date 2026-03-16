import json
from pathlib import Path
from openpyxl import load_workbook

SOURCE_JSON = Path("C:\\Users\\KIIT01\\Downloads\\convert\\convert\\formulas.json")
TEMPLATE_XLSX = Path("C:\\Users\\KIIT01\\Downloads\\convert\\convert\\FMEDA_TEMPLATE.xlsx")
OUTPUT_XLSX = Path("formulas_injected.xlsx")


def load_json_any_encoding(path: Path):
    """Load JSON trying common encodings (UTF-8, UTF-16, UTF-8-SIG)."""
    for enc in ("utf-8", "utf-16", "utf-8-sig"):
        try:
            with open(path, "r", encoding=enc) as f:
                return json.load(f)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("utf-8", b"", 0, 1, "Unable to decode JSON file")


def inject_formulas_into_template(formulas_data, template_path: Path, output_path: Path):
    """
    Inject formulas into Excel template using direct cell references.
    
    JSON structure:
    {
      "Cover": {
        "I3": "=F18",
        "I4": "=F19",
        ...
      },
      "Block Diagram": {
        "L2": "=Cover!$F$18",
        ...
      }
    }
    
    Logic: For each sheet, write each formula to the specified cell reference
    """
    wb = load_workbook(template_path)
    
    total_formulas = 0
    
    for sheet_name, formulas in formulas_data.items():
        # Get existing sheet or create new one
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            print(f"  ⚠ Created new sheet: {sheet_name}")
        
        formulas_written = 0
        
        # Write each formula to its cell
        for cell_ref, formula in formulas.items():
            try:
                ws[cell_ref] = formula
                formulas_written += 1
                total_formulas += 1
            except Exception as e:
                print(f"  ⚠ Error writing formula to {sheet_name}!{cell_ref}: {e}")
        
        print(f"  ✓ {sheet_name}: {formulas_written} formulas written")
    
    wb.save(output_path)
    print(f"\n✅ Total formulas written: {total_formulas}")


def main():
    print("Loading formulas JSON...")
    formulas = load_json_any_encoding(SOURCE_JSON)
    
    print(f"Loading template: {TEMPLATE_XLSX}")
    inject_formulas_into_template(formulas, TEMPLATE_XLSX, OUTPUT_XLSX)
    
    print(f"✅ Excel file created successfully: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()