import openpyxl
import json
import sys

def extract_formulas_from_excel(excel_file, output_json='formulas_output.json'):
    """
    Extract all formulas from an Excel file and save to JSON.
    
    Args:
        excel_file: Path to the Excel file
        output_json: Path to output JSON file (default: formulas_output.json)
    """
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_file, data_only=False)
    
    # Dictionary to store all formulas
    all_formulas = {}
    
    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_formulas = {}
        
        # Iterate through all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                # Check if cell contains a formula
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_reference = cell.coordinate
                    sheet_formulas[cell_reference] = cell.value
        
        # Only add sheet to output if it has formulas
        if sheet_formulas:
            all_formulas[sheet_name] = sheet_formulas
    
    # Save to JSON file
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(all_formulas, f, indent=2, ensure_ascii=False)
    
    # Print summary
    total_formulas = sum(len(formulas) for formulas in all_formulas.values())
    print(f"✓ Extracted {total_formulas} formulas from {len(all_formulas)} sheets")
    print(f"✓ Output saved to: {output_json}")
    
    # Print sheet-by-sheet breakdown
    print("\nBreakdown by sheet:")
    for sheet_name, formulas in all_formulas.items():
        print(f"  - {sheet_name}: {len(formulas)} formulas")
    
    return all_formulas

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_excel_formulas.py <excel_file> [output_json]")
        print("\nExample:")
        print("  python extract_excel_formulas.py myfile.xlsx")
        print("  python extract_excel_formulas.py myfile.xlsx output.json")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_json = sys.argv[2] if len(sys.argv) > 2 else 'formulas_output.json'
    
    try:
        extract_formulas_from_excel(excel_file, output_json)
    except FileNotFoundError:
        print(f"Error: File '{excel_file}' not found")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
